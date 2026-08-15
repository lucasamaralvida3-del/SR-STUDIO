from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ALIASES = {
    "code": {"CODIGO", "COD", "ID", "COD PRODUTO", "CODIGO PRODUTO"},
    "ean": {"EAN", "COD BARRAS", "CODIGO BARRAS", "CODBARRAS", "GTIN"},
    "name": {"PRODUTO", "ITEM", "DESCRICAO", "NOME", "DESCRICAO PRODUTO"},
    "promo_price": {"PROMOCAO", "PRECO PROMOCAO", "PRECO PROMO", "PRECO", "OFERTA"},
    "app_price": {"APP", "PRECO APP", "CLUBE", "PRECO CLUBE"},
    "retail_price": {"VAREJO", "PRECO VAREJO", "VENDA"},
    "wholesale_price": {"ATACADO", "PRECO ATACADO"},
    "quantity": {
        "QUANTIDADE",
        "QTD",
        "QTD ATACADO",
        "QUANTIDADE ATACADO",
        "A PARTIR DE",
        "MINIMO ATACADO",
        "QUANTIDADE MINIMA",
    },
    "unit": {"UNIDADE", "UN", "ENTRADA", "TIPO VENDA"},
    "limit": {"LIMITE", "LIMITE CPF", "LIMITE POR CPF"},
    "category": {"CATEGORIA", "SETOR", "DEPARTAMENTO", "SECAO"},
    "validity": {"VALIDADE", "VALIDO ATE", "FIM PROMOCAO"},
}


@dataclass(slots=True)
class ImportIssue:
    row: int
    field: str
    message: str
    severity: str = "warning"


@dataclass(slots=True)
class ExcelImportResult:
    products: list[dict[str, Any]] = field(default_factory=list)
    issues: list[ImportIssue] = field(default_factory=list)
    column_map: dict[str, int] = field(default_factory=dict)
    sheet_name: str = ""
    header_row: int = 0


class ExcelImporter:
    """Importador tolerante a variações de cabeçalho e planilhas operacionais."""

    def import_file(self, path: str | Path, sheet: str | None = None) -> ExcelImportResult:
        file_path = Path(path)
        wb = load_workbook(file_path, data_only=True, read_only=True)
        ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
        header_row, mapping = self._detect_header(ws)
        result = ExcelImportResult(column_map=mapping, sheet_name=ws.title, header_row=header_row)
        if "name" not in mapping:
            result.issues.append(
                ImportIssue(header_row, "name", "Não foi possível identificar a coluna de produto.", "critical")
            )
            return result

        seen: set[str] = set()
        for row_index, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            values = {field: row[index] if index < len(row) else None for field, index in mapping.items()}
            name = self._text(values.get("name"))
            if not name:
                continue
            code = self._text(values.get("code"))
            ean = self._text(values.get("ean"))
            unit = self._normalize_unit(values.get("unit"), name)
            identity = ean or code or self._norm(name)
            product = {
                "code": code,
                "ean": ean,
                "name": name,
                "promo_price": values.get("promo_price"),
                "app_price": values.get("app_price"),
                "retail_price": values.get("retail_price"),
                "wholesale_price": values.get("wholesale_price"),
                "quantity": self._text(values.get("quantity")),
                "unit": unit,
                "limit": self._text(values.get("limit")),
                "category": self._text(values.get("category")),
                "validity": values.get("validity"),
                "source_row": row_index,
            }
            if identity in seen:
                result.issues.append(ImportIssue(row_index, "product", f"Produto duplicado: {name}"))
            seen.add(identity)
            if not any(
                product.get(k) not in (None, "")
                for k in ("promo_price", "app_price", "retail_price", "wholesale_price")
            ):
                result.issues.append(ImportIssue(row_index, "price", f"Produto sem preço: {name}"))
            result.products.append(product)
        return result

    def _detect_header(self, ws) -> tuple[int, dict[str, int]]:
        best_row = 1
        best_mapping: dict[str, int] = {}
        best_score = -1
        for row_index, row in enumerate(
            ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20), values_only=True),
            start=1,
        ):
            mapping: dict[str, int] = {}
            for idx, value in enumerate(row):
                normalized = self._norm(value)
                if not normalized:
                    continue
                for field, aliases in ALIASES.items():
                    if normalized in aliases and field not in mapping:
                        mapping[field] = idx
                        break
            score = len(mapping) + (3 if "name" in mapping else 0) + (
                2 if any(k in mapping for k in ("promo_price", "retail_price", "app_price")) else 0
            )
            if score > best_score:
                best_row, best_mapping, best_score = row_index, mapping, score
        return best_row, best_mapping

    @classmethod
    def _norm(cls, value: Any) -> str:
        text = cls._text(value).upper()
        text = unicodedata.normalize("NFD", text)
        text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
        return re.sub(r"[^A-Z0-9]+", " ", text).strip()

    @staticmethod
    def _text(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    @classmethod
    def _normalize_unit(cls, raw: Any, name: str) -> str:
        unit = cls._norm(raw)
        if any(token in unit for token in ("KG", "KILO", "PESO")):
            return "KG"
        if any(token in unit for token in ("UN", "UNIDADE", "PC", "PCT")):
            return "UN"
        normalized_name = cls._norm(name)
        if "A GRANEL" in normalized_name:
            return "KG"
        return "UN"
