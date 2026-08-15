from __future__ import annotations

from dataclasses import dataclass, field
from decimal import Decimal
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from srstudio.core.models import Product, to_decimal
from srstudio.posters.core import PosterKind
from srstudio.posters.importers import PromotionWorkbookImporter


@dataclass(frozen=True, slots=True)
class CommercialIssue:
    code: str
    severity: str
    group: str
    message: str
    field: str = ""


@dataclass(slots=True)
class CommercialStatus:
    overall: str = "OK"
    issues: list[CommercialIssue] = field(default_factory=list)
    groups: dict[str, str] = field(default_factory=dict)

    @property
    def tooltip(self) -> str:
        if not self.issues:
            return "Nenhuma inconsistência comercial encontrada."
        return "\n".join(f"• {issue.message}" for issue in self.issues)

    def group_tooltip(self, group: str, labels: dict[str, str] | None = None) -> str:
        matches = [issue.message for issue in self.issues if issue.group == group]
        if matches:
            return "\n".join(f"• {message}" for message in matches)
        known = {"cost": "Custo", "sale": "Venda", "club": "Clube", "unit": "Unidade"}
        if labels:
            known.update(labels)
        return f"{known.get(group, group.title())}: nenhuma inconsistência encontrada."


class PosterCommercialValidator:
    """Commercial checks ported from the proven Stable poster workflow."""

    ERROR = "ERRO"
    WARNING = "ALERTA"
    OK = "OK"

    def evaluate(self, product: Product, kind: PosterKind = PosterKind.PROMOTION) -> CommercialStatus:
        issues: list[CommercialIssue] = []
        cost = self.cost(product)
        sale = product.retail_price

        if kind == PosterKind.WHOLESALE:
            promo = product.wholesale_price
            club = None
            price_specs = (("ATACADO", promo, "wholesale_price", "club"),)
            # Report 782 does not carry cost. Only flag missing cost when another import
            # explicitly claimed it should exist; otherwise wholesale would be all-yellow.
            expects_cost = bool(product.metadata.get("expects_cost"))
        else:
            poster_type = int(product.metadata.get("promotion_type", 0) or 0)
            if poster_type == 3:
                promo = None
                club = product.price
            else:
                promo = product.price
                club = product.app_price
            price_specs = (
                ("PROMOÇÃO", promo, "price", "sale"),
                ("CLUBE", club, "app_price", "club"),
            )
            expects_cost = True

        if cost is None and expects_cost:
            issues.append(
                CommercialIssue(
                    "CUSTO_AUSENTE",
                    self.WARNING,
                    "cost",
                    "Custo não informado na planilha; não foi possível validar margem/abaixo do custo.",
                    "cost",
                )
            )
        if sale is None:
            issues.append(
                CommercialIssue(
                    "VENDA_AUSENTE",
                    self.WARNING,
                    "sale",
                    "Preço de venda não informado; a comparação com o preço normal não pôde ser feita.",
                    "retail_price",
                )
            )

        for label, price, field_name, group in price_specs:
            if price is None:
                continue
            if price <= 0:
                issues.append(
                    CommercialIssue(
                        "PRECO_INVALIDO",
                        self.ERROR,
                        group,
                        f"{label} está zerado ou inválido.",
                        field_name,
                    )
                )
                continue
            if cost is not None and price < cost:
                diff = cost - price
                issues.append(
                    CommercialIssue(
                        "ABAIXO_CUSTO",
                        self.ERROR,
                        "cost" if label == "PROMOÇÃO" else group,
                        f"{label} {self.money(price)} está abaixo do custo {self.money(cost)} em {self.money(diff)}.",
                        field_name,
                    )
                )
            if sale is not None and sale > 0:
                ratio = price / sale
                if ratio < Decimal("0.20") or ratio > Decimal("1.20"):
                    issues.append(
                        CommercialIssue(
                            "PRECO_FORA_PADRAO",
                            self.WARNING,
                            group,
                            f"{label} {self.money(price)} está muito distante do preço de venda {self.money(sale)}.",
                            field_name,
                        )
                    )
                elif price > sale:
                    issues.append(
                        CommercialIssue(
                            "PROMO_ACIMA_VENDA",
                            self.WARNING,
                            group,
                            f"{label} {self.money(price)} está acima do preço de venda {self.money(sale)}.",
                            field_name,
                        )
                    )

        if kind == PosterKind.PROMOTION and promo is not None and club is not None and club > promo:
            issues.append(
                CommercialIssue(
                    "CLUBE_MAIOR_PROMO",
                    self.WARNING,
                    "club",
                    f"Preço Clube {self.money(club)} está maior que a Promoção {self.money(promo)}.",
                    "app_price",
                )
            )

        unit_warning = self.unit_warning(product)
        if unit_warning:
            issues.append(
                CommercialIssue(
                    "UNIDADE_SUSPEITA",
                    self.WARNING,
                    "unit",
                    unit_warning,
                    "unit",
                )
            )

        groups = {name: self.OK for name in ("cost", "sale", "club", "unit")}
        for issue in issues:
            previous = groups.get(issue.group, self.OK)
            if issue.severity == self.ERROR or previous != self.ERROR and issue.severity == self.WARNING:
                groups[issue.group] = issue.severity
        overall = self.ERROR if any(x.severity == self.ERROR for x in issues) else (
            self.WARNING if issues else self.OK
        )
        return CommercialStatus(overall=overall, issues=issues, groups=groups)

    def evaluate_batch(
        self,
        products: Iterable[Product],
        kind: PosterKind = PosterKind.PROMOTION,
    ) -> dict[str, CommercialStatus]:
        product_list = list(products)
        statuses = {product.id: self.evaluate(product, kind) for product in product_list}
        seen: dict[str, Product] = {}
        for product in product_list:
            identity = str(product.code or product.ean or "").strip() or product.name.casefold().strip()
            if not identity:
                continue
            if identity in seen:
                issue = CommercialIssue(
                    "DUPLICADO",
                    self.WARNING,
                    "sale",
                    f"Produto/código duplicado na geração; também aparece como '{seen[identity].name}'.",
                    "name",
                )
                status = statuses[product.id]
                status.issues.append(issue)
                if status.overall != self.ERROR:
                    status.overall = self.WARNING
                if status.groups.get("sale") != self.ERROR:
                    status.groups["sale"] = self.WARNING
            else:
                seen[identity] = product
        return statuses

    @staticmethod
    def cost(product: Product) -> Decimal | None:
        return to_decimal(product.metadata.get("cost") or product.metadata.get("custo"))

    @staticmethod
    def unit_warning(product: Product) -> str:
        name = " ".join(product.name.upper().split())
        unit = str(product.unit or "").upper()
        if "A GRANEL" in name and unit != "KG":
            return "Produto indica 'A GRANEL', mas a unidade do cartaz não está como KG."
        if unit == "KG" and __import__("re").search(r"\b\d+(?:[.,]\d+)?\s*(?:ML|L)\b", name):
            return "Produto está marcado como KG, mas a descrição contém medida em ML/L."
        return ""

    @staticmethod
    def money(value: Decimal | None) -> str:
        return "—" if value is None else f"R$ {value:.2f}".replace(".", ",")


def enrich_promotion_commercial_data(path: str | Path, products: Iterable[Product]) -> None:
    """Attach Stable-compatible CUSTO/VENDA source values without changing importer contracts."""

    source = Path(path)
    source_key = str(source.resolve()).casefold()
    product_list = []
    for product in products:
        raw_source = str(product.metadata.get("source_file") or "")
        try:
            product_source = str(Path(raw_source).resolve()).casefold() if raw_source else ""
        except Exception:
            product_source = raw_source.casefold()
        if product_source == source_key:
            product_list.append(product)
    if not product_list or source.suffix.lower() not in {".xlsx", ".xlsm"}:
        return
    workbook = load_workbook(source, data_only=True, read_only=False)
    grouped: dict[str, list[Product]] = {}
    for product in product_list:
        grouped.setdefault(str(product.metadata.get("sheet") or ""), []).append(product)

    for sheet_name, items in grouped.items():
        if sheet_name not in workbook.sheetnames:
            continue
        worksheet = workbook[sheet_name]
        header_rows = PromotionWorkbookImporter.detect_header_rows(worksheet)
        if not header_rows:
            continue
        for product in items:
            row = int(product.metadata.get("source_row") or 0)
            header_row = max((candidate for candidate in header_rows if candidate < row), default=header_rows[0])
            headers: dict[str, int] = {}
            for column in range(1, worksheet.max_column + 1):
                normalized = PromotionWorkbookImporter.norm(worksheet.cell(header_row, column).value)
                if normalized:
                    headers[normalized] = column
            cost_col = next(
                (
                    headers[key]
                    for key in (
                        "CUSTO",
                        "CUSTO_UNITARIO",
                        "PRECO_DE_CUSTO",
                        "CUSTO_GERENCIAL",
                    )
                    if key in headers
                ),
                None,
            )
            sale_col = next(
                (
                    headers[key]
                    for key in ("VENDA", "PRECO_VENDA", "VAREJO", "PRECO_VAREJO")
                    if key in headers
                ),
                None,
            )
            cost_raw = worksheet.cell(row, cost_col).value if cost_col else None
            sale_raw = worksheet.cell(row, sale_col).value if sale_col else None
            cost = to_decimal(cost_raw)
            if cost is not None:
                product.metadata["cost"] = str(cost)
            if product.retail_price is None and sale_raw not in (None, ""):
                product.retail_price = to_decimal(sale_raw)
            ensure_imported_snapshot(product)


def ensure_imported_snapshot(product: Product) -> None:
    product.metadata.setdefault(
        "imported_snapshot",
        {
            "display_name": product.display_name,
            "price": None if product.price is None else str(product.price),
            "app_price": None if product.app_price is None else str(product.app_price),
            "retail_price": None if product.retail_price is None else str(product.retail_price),
            "wholesale_price": None if product.wholesale_price is None else str(product.wholesale_price),
            "unit": product.unit,
            "quantity": product.quantity,
            "cpf_limit": product.cpf_limit,
            "promotion_type": product.metadata.get("promotion_type"),
            "cost": product.metadata.get("cost", ""),
        },
    )


def restore_imported_snapshot(product: Product) -> bool:
    snapshot = product.metadata.get("imported_snapshot")
    if not isinstance(snapshot, dict):
        return False
    product.display_name = str(snapshot.get("display_name") or product.original_name)
    product.price = to_decimal(snapshot.get("price"))
    product.app_price = to_decimal(snapshot.get("app_price"))
    product.retail_price = to_decimal(snapshot.get("retail_price"))
    product.wholesale_price = to_decimal(snapshot.get("wholesale_price"))
    product.unit = str(snapshot.get("unit") or "UN").upper().strip()
    product.quantity = str(snapshot.get("quantity") or "")
    product.cpf_limit = str(snapshot.get("cpf_limit") or "")
    if snapshot.get("promotion_type") is not None:
        product.metadata["promotion_type"] = snapshot.get("promotion_type")
    if snapshot.get("cost") not in (None, ""):
        product.metadata["cost"] = str(snapshot.get("cost"))
    product.metadata.pop("edited", None)
    product.metadata.pop("edited_fields", None)
    return True
