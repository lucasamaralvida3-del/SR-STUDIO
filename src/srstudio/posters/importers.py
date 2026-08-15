from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass, field
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader

from srstudio.core.models import Product, to_decimal


@dataclass(slots=True)
class PosterImportResult:
    products: list[Product] = field(default_factory=list)
    campaigns: list[dict[str, object]] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)
    metadata: dict[str, object] = field(default_factory=dict)


class PromotionWorkbookImporter:
    """Ports the proven Stable 2 promotion-sheet semantics into v5."""

    def import_file(self, path: str | Path) -> PosterImportResult:
        source = Path(path)
        workbook = load_workbook(source, data_only=True)
        result = PosterImportResult(metadata={"source_file": str(source)})
        file_dates = self.extract_date_tokens(source.stem)

        for worksheet in workbook.worksheets:
            header_rows = self.detect_header_rows(worksheet)
            for section_index, header_row in enumerate(header_rows):
                title = self.find_title_for_header(worksheet, header_row)
                campaign = self.clean_campaign_title(title)
                validity = self.period_from(title) or (file_dates[0] if file_dates else "")
                internal_dates = self.extract_date_tokens(title)
                if internal_dates and file_dates:
                    internal = internal_dates[0].replace("-", "/")
                    file_date = file_dates[0].replace("-", "/")
                    if internal != file_date:
                        result.errors.append(
                            f"Data divergente em '{campaign}': enunciado='{internal}' / arquivo='{file_date}'."
                        )

                headers: dict[str, int] = {}
                for column in range(1, worksheet.max_column + 1):
                    key = self.norm(worksheet.cell(header_row, column).value)
                    if key:
                        headers[key] = column
                product_col = headers.get("PRODUTO") or headers.get("PRODUTOS")
                if not product_col:
                    continue
                code_col = headers.get("EAN") or headers.get("CODIGO") or headers.get("COD")
                retail_col = (
                    headers.get("VENDA")
                    or headers.get("PRECO_VENDA")
                    or headers.get("VAREJO")
                    or headers.get("PRECO_VAREJO")
                )
                promo_col = headers.get("PROMOCAO") or headers.get("PRECO_PROMOCAO")
                club_col = headers.get("CLUBE") or headers.get("APP") or headers.get("PRECO_CLUBE")
                entry_col = headers.get("ENTRADA") or headers.get("ENT") or headers.get("UNIDADE")
                limit_col = (
                    headers.get("LIMITE")
                    or headers.get("LIMIT")
                    or headers.get("LIMITE_CPF")
                    or headers.get("LIMITE_CLIENTE")
                )
                next_header = (
                    header_rows[section_index + 1]
                    if section_index + 1 < len(header_rows)
                    else worksheet.max_row + 1
                )
                section_products: list[Product] = []
                type_counts = {1: 0, 2: 0, 3: 0}

                for row in range(header_row + 1, next_header):
                    raw_name = worksheet.cell(row, product_col).value
                    if raw_name is None or not str(raw_name).strip():
                        continue
                    name = " ".join(str(raw_name).strip().split())
                    normalized_name = self.norm(name)
                    if normalized_name in {
                        "PRODUTO",
                        "PRODUTOS",
                        "CERVEJAS",
                        "BEBIDAS",
                        "BEBDAS",
                        "SCANTECH",
                    } or normalized_name.startswith("OFERTA"):
                        continue

                    promo = worksheet.cell(row, promo_col).value if promo_col else None
                    club = worksheet.cell(row, club_col).value if club_col else None
                    promo_present = promo is not None and str(promo).strip() != ""
                    club_present = club is not None and str(club).strip() != ""
                    if promo_present and not club_present:
                        poster_type = 1
                        main_price = promo
                        club_price = None
                    elif promo_present and club_present:
                        if self.same_price(promo, club):
                            poster_type = 1
                            main_price = promo
                            club_price = None
                        else:
                            poster_type = 2
                            main_price = promo
                            club_price = club
                    elif not promo_present and club_present:
                        poster_type = 3
                        main_price = club
                        club_price = None
                    else:
                        continue

                    entry = worksheet.cell(row, entry_col).value if entry_col else None
                    unit, recognized = self.detect_unit_display(entry)
                    if not recognized:
                        result.warnings.append(
                            f"{worksheet.title} linha {row}: unidade '{entry}' não reconhecida; assumido UN."
                        )
                    limit_raw = worksheet.cell(row, limit_col).value if limit_col else None
                    if isinstance(limit_raw, float) and limit_raw.is_integer():
                        limit_text = str(int(limit_raw))
                    else:
                        limit_text = "" if limit_raw is None else str(limit_raw).strip()
                    code_raw = worksheet.cell(row, code_col).value if code_col else None
                    if isinstance(code_raw, float) and code_raw.is_integer():
                        code = str(int(code_raw))
                    else:
                        code = "" if code_raw is None else str(code_raw).strip()
                    retail = worksheet.cell(row, retail_col).value if retail_col else None
                    product = Product(
                        code=code,
                        original_name=name,
                        price=main_price,
                        app_price=club_price,
                        retail_price=retail,
                        unit=unit,
                        cpf_limit=limit_text,
                        campaign="CLUBE EXCLUSIVO" if poster_type == 3 else campaign,
                        validity=validity,
                        source="promotion_workbook",
                        metadata={
                            "poster_kind": "promotion",
                            "promotion_type": poster_type,
                            "source_file": str(source),
                            "sheet": worksheet.title,
                            "source_row": row,
                            "club_price_original": "" if club is None else str(club),
                            "entry_original": "" if entry is None else str(entry),
                        },
                    )
                    section_products.append(product)
                    type_counts[poster_type] += 1
                    if len(name) > 48:
                        result.warnings.append(
                            f"{worksheet.title} linha {row}: nome longo; será ajustado em até 2 linhas."
                        )

                if section_products:
                    result.products.extend(section_products)
                    result.campaigns.append(
                        {
                            "name": campaign,
                            "validity": validity,
                            "total": len(section_products),
                            "one_price": type_counts[1],
                            "two_prices": type_counts[2],
                            "club_only": type_counts[3],
                            "sheet": worksheet.title,
                        }
                    )
        if not result.products and not result.errors:
            result.errors.append(
                "Nenhuma seção de promoção foi reconhecida. A planilha precisa ter PRODUTO(S) e PROMOÇÃO ou CLUBE."
            )
        return result

    @classmethod
    def detect_header_rows(cls, worksheet) -> list[int]:
        rows: list[int] = []
        for row in range(1, worksheet.max_row + 1):
            values = [
                cls.norm(worksheet.cell(row, column).value)
                for column in range(1, min(worksheet.max_column, 20) + 1)
            ]
            if ("PRODUTO" in values or "PRODUTOS" in values) and (
                "PROMOCAO" in values or "CLUBE" in values or "APP" in values
            ):
                rows.append(row)
        return rows

    @classmethod
    def find_title_for_header(cls, worksheet, header_row: int) -> str:
        candidates: list[tuple[int, int, str]] = []
        for row in range(header_row - 1, max(0, header_row - 8), -1):
            values = [worksheet.cell(row, column).value for column in range(1, min(worksheet.max_column, 8) + 1)]
            texts = [str(value).strip() for value in values if value is not None and str(value).strip()]
            if not texts:
                continue
            joined = " ".join(texts)
            score = (2 if len(texts) == 1 else 0) + (6 if cls.is_probable_title(joined) else 0)
            score += 3 if cls.period_from(joined) else 0
            score += max(0, 8 - (header_row - row))
            candidates.append((score, row, joined))
        if candidates:
            candidates.sort(reverse=True)
            return candidates[0][2]
        return worksheet.title

    @classmethod
    def clean_campaign_title(cls, text: str) -> str:
        value = " ".join(str(text or "").split())
        value = re.sub(
            r"\b\d{1,2}[/-]\d{1,2}\s*A\s*\d{1,2}[/-]\d{1,2}[/-]\d{4}\b.*$",
            "",
            value,
            flags=re.I,
        ).strip()
        value = re.sub(
            r"\b\d{1,2}\s*A\s*\d{1,2}[/-]\d{1,2}[/-]\d{4}\b.*$",
            "",
            value,
            flags=re.I,
        ).strip()
        value = re.sub(r"\b\d{1,2}[/-]\d{1,2}[/-]\d{4}\b.*$", "", value, flags=re.I).strip()
        value = re.sub(r"\bSANJU\b|\bSANTA\s+JULIANA\b", "", value, flags=re.I)
        value = " ".join(value.split()).strip(" -–—:") or "OFERTA"
        if cls.norm(value) == "BEBDAS":
            value = "BEBIDAS"
        return value.upper().rstrip("! ") + "!!"

    @staticmethod
    def period_from(text: str) -> str:
        value = str(text or "")
        match = re.search(
            r"(\d{1,2})[/-](\d{1,2})\s*A\s*(\d{1,2})[/-](\d{1,2})[/-](\d{4})",
            value,
            re.I,
        )
        if match:
            return (
                f"{int(match.group(1)):02d}/{int(match.group(2)):02d} A "
                f"{int(match.group(3)):02d}/{int(match.group(4)):02d}/{match.group(5)}"
            )
        match = re.search(r"(\d{1,2})\s*A\s*(\d{1,2})[/-](\d{1,2})[/-](\d{4})", value, re.I)
        if match:
            return f"{int(match.group(1)):02d} A {int(match.group(2)):02d}/{int(match.group(3)):02d}/{match.group(4)}"
        match = re.search(r"(\d{1,2})[/-](\d{1,2})[/-](\d{4})", value)
        if match:
            return f"{int(match.group(1)):02d}/{int(match.group(2)):02d}/{match.group(3)}"
        return ""

    @classmethod
    def extract_date_tokens(cls, text: str) -> list[str]:
        patterns = [
            r"\d{1,2}[/-]\d{1,2}\s*A\s*\d{1,2}[/-]\d{1,2}[/-]\d{4}",
            r"\d{1,2}\s*A\s*\d{1,2}[/-]\d{1,2}[/-]\d{4}",
            r"\d{1,2}[/-]\d{1,2}[/-]\d{4}",
        ]
        hits: list[str] = []
        for pattern in patterns:
            for match in re.findall(pattern, str(text or ""), flags=re.I):
                hits.append(re.sub(r"\s+", " ", match.strip()).replace("-", "/").upper())
        return hits

    @classmethod
    def is_probable_title(cls, text: str) -> bool:
        normalized = cls.norm(text)
        keywords = (
            "OFERTA",
            "VERDE",
            "LIMPEZA",
            "ECONOMIA",
            "CAFE",
            "PAO",
            "FILE",
            "FIM_DE_SEMANA",
            "CERVEJA",
            "BEBIDA",
            "CARNES",
            "VERDURAS",
        )
        return any(token in normalized for token in keywords) or bool(cls.period_from(text))

    @classmethod
    def detect_unit_display(cls, value) -> tuple[str, bool]:
        original = "" if value is None else str(value).strip()
        normalized = cls.norm(original)
        if not normalized:
            return "UN", False
        if "GARRAFA" in normalized:
            return "À GARRAFA", True
        if "LATA" in normalized:
            return "À LATA", True
        if "KG" in normalized or "QUILO" in normalized or "KILO" in normalized or "PESO" in normalized:
            return "KG", True
        if normalized in {"UN", "UND", "UNID", "UNIDADE", "UNIDADES", "CADA"} or "UNIDADE" in normalized:
            return "UN", True
        return "UN", False

    @staticmethod
    def same_price(left, right) -> bool:
        first, second = to_decimal(left), to_decimal(right)
        if first is None or second is None:
            return str(left).strip() == str(right).strip()
        return abs(first - second) < Decimal("0.005")

    @staticmethod
    def norm(value) -> str:
        text = "" if value is None else str(value)
        text = unicodedata.normalize("NFD", text)
        text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
        return re.sub(r"[^A-Z0-9]+", "_", text.upper()).strip("_")


class WholesaleReportImporter:
    """Ports the Stable 2 CISS report 782 Atacado parser into v5."""

    REPORT_MARKER = "782-Listagem de Produtos Atacarejo"

    def import_file(self, path: str | Path) -> PosterImportResult:
        source = Path(path)
        reader = PdfReader(str(source))
        all_text: list[str] = []
        products: list[Product] = []
        pending: tuple[str, str, str] | None = None

        for page_index, page in enumerate(reader.pages, start=1):
            text = page.extract_text() or ""
            all_text.append(text)
            lines = [line.strip() for line in text.splitlines() if line.strip()]
            pending = None
            for line in lines:
                match = re.match(r"^(\d+)\s*-\s*(.*?)\s+(\d*,\d{2})Preço Varejo:\s*$", line, re.I)
                if not match:
                    match = re.match(r"^(\d+)\s*-\s*(.*?)\s+Preço Varejo:\s*(\d*,\d{2})\s*$", line, re.I)
                if match:
                    pending = (match.group(1), match.group(2).strip(), match.group(3))
                    continue
                quantity_match = re.match(
                    r"^A partir de\s+([\d.,]+)\s+([\d.,]+)\s+([\d.,]+)\s*$",
                    line,
                    re.I,
                )
                if quantity_match and pending:
                    code, description, retail = pending
                    quantity, discount, wholesale = quantity_match.groups()
                    unit = self.infer_unit(description, quantity)
                    quantity_display = self.quantity_display(quantity, unit)
                    wholesale_decimal = to_decimal(wholesale) or Decimal("0")
                    quantity_decimal = to_decimal(quantity) or Decimal("0")
                    total = (quantity_decimal * wholesale_decimal).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                    products.append(
                        Product(
                            code=code,
                            original_name=description,
                            retail_price=retail,
                            wholesale_price=wholesale,
                            quantity=quantity_display,
                            unit=unit,
                            campaign="Atacado",
                            source="atacado_report_782",
                            metadata={
                                "poster_kind": "wholesale",
                                "discount": self.money(discount),
                                "total": self.money(total),
                                "source_file": str(source),
                                "source_page": page_index,
                            },
                        )
                    )
                    pending = None

        full_text = "\n".join(all_text)
        result = PosterImportResult(products=products, metadata={"source_file": str(source), "pages": len(reader.pages)})
        if self.REPORT_MARKER not in full_text:
            result.errors.append("O PDF não é o relatório 782 - Listagem de Produtos Atacarejo.")
            result.products.clear()
            return result
        company = re.search(r"Empresa:\s*(\d+)\s*-\s*([^\r\n]+)", full_text, re.I)
        report_date = re.search(r"\b(\d{2}/\d{2}/\d{4}\s+\d{2}:\d{2}:\d{2})\b", full_text)
        if company:
            result.metadata["company_code"] = company.group(1).strip()
            result.metadata["company_name"] = company.group(2).strip()
        if report_date:
            result.metadata["report_date"] = report_date.group(1)
        if not result.products:
            result.errors.append("Nenhum produto Atacarejo foi identificado no relatório.")
        return result

    @classmethod
    def infer_unit(cls, description: str, quantity: str) -> str:
        quantity_text = str(quantity or "").strip()
        if re.fullmatch(r"\d+,\d{3}", quantity_text):
            return "KG"
        normalized = PromotionWorkbookImporter.norm(description).replace("_", " ")
        if re.search(r"\bGARRAFA(S)?\b|\bGFA\b|\bGF\b", normalized):
            return "À GARRAFA"
        if re.search(r"\bLATA(S)?\b", normalized):
            return "À LATA"
        if re.search(r"\bLT\b", normalized) and any(
            token in normalized
            for token in ("CERVEJA", "REFRIGERANTE", "ENERGETICO", "BEBIDA", "AGUA TONICA", "CHA")
        ):
            return "À LATA"
        return "UN"

    @staticmethod
    def quantity_display(value: str, unit: str) -> str:
        decimal = to_decimal(value)
        if unit == "KG" and decimal is not None:
            return f"{decimal:.3f}".replace(".", ",")
        if decimal is not None and decimal == decimal.to_integral_value():
            return str(int(decimal))
        return str(value).strip()

    @staticmethod
    def money(value) -> str:
        decimal = to_decimal(value)
        if decimal is None:
            return str(value or "")
        return f"{decimal.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP):.2f}".replace(".", ",")
