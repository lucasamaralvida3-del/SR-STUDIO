# -*- coding: utf-8 -*-
"""Exportação do Montador usando as estruturas reais das planilhas promocionais do SR."""
import re
from copy import copy
from pathlib import Path
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook

APP_DIR = Path(__file__).resolve().parent
TEMPLATE_DIR = APP_DIR / "modelos_excel"

PROFILE_LABELS = {
    "AUTO": "AUTOMÁTICO",
    "SEGUNDA": "SEGUNDA DA LIMPEZA E ECONOMIA",
    "TERCA": "TERÇA VERDE",
    "QUARTA": "QUARTA CAFÉ COM PÃO",
    "QUINTA": "QUINTA FILÉ",
    "FIM_SEMANA": "FIM DE SEMANA",
    "CARNES_VERDURAS": "CARNES E VERDURAS",
    "CLUBE": "OFERTAS CLUBE",
}

LABEL_TO_PROFILE = {v: k for k, v in PROFILE_LABELS.items()}

TEMPLATES = {
    "SEGUNDA": "SEGUNDA_DA_LIMPEZA_E_ECONOMIA.xlsx",
    "TERCA": "TERCA_VERDE.xlsx",
    "QUARTA": "QUARTA_CAFE_COM_PAO.xlsx",
    "QUINTA": "QUINTA_FILE.xlsx",
    "FIM_SEMANA": "FIM_DE_SEMANA.xlsx",
    "CARNES_VERDURAS": "CARNES_E_VERDURAS.xlsx",
    "CLUBE": "OFERTAS_CLUBE.xlsx",
}


def profile_choices():
    return [PROFILE_LABELS[k] for k in ("AUTO","SEGUNDA","TERCA","QUARTA","QUINTA","FIM_SEMANA","CARNES_VERDURAS","CLUBE")]


def _norm(value):
    s = str(value or "").upper()
    return (s.replace("Á","A").replace("À","A").replace("Â","A").replace("Ã","A")
             .replace("É","E").replace("Ê","E").replace("Í","I").replace("Ó","O")
             .replace("Ô","O").replace("Õ","O").replace("Ú","U").replace("Ç","C"))


def detect_profile(campaign):
    n = _norm(campaign)
    if "CLUBE" in n: return "CLUBE"
    if "TERCA" in n or "TERÇA" in str(campaign).upper() or "VERDE" in n: return "TERCA"
    if "QUARTA" in n or "CAFE COM PAO" in n: return "QUARTA"
    if "QUINTA" in n or "FILE" in n: return "QUINTA"
    if "CARNES E VERDURA" in n or "CARNE E VERDURA" in n: return "CARNES_VERDURAS"
    if "SEGUNDA" in n or "LIMPEZA" in n: return "SEGUNDA"
    if "FIM DE SEMANA" in n or "FINAL DE SEMANA" in n: return "FIM_SEMANA"
    return "FIM_SEMANA"


def resolve_profile(choice, campaign):
    key = LABEL_TO_PROFILE.get(str(choice or "").strip().upper())
    if key and key != "AUTO": return key
    return detect_profile(campaign)


def _num(v):
    if v in (None, ""): return None
    if isinstance(v, (int, float)): return v
    s = str(v).strip().replace("R$", "").replace(" ", "")
    if not s: return None
    if "@" in s: return str(v)
    if "," in s and "." in s: s = s.replace(".", "").replace(",", ".")
    elif "," in s: s = s.replace(",", ".")
    try: return float(Decimal(s))
    except (InvalidOperation, ValueError): return str(v)


def _style_snapshot(ws, row, max_col):
    snap=[]
    for col in range(1,max_col+1):
        c=ws.cell(row,col)
        snap.append({
            "style": copy(c._style),
            "number_format": c.number_format,
            "alignment": copy(c.alignment),
            "protection": copy(c.protection),
        })
    return {"cells":snap,"height":ws.row_dimensions[row].height}


def _apply_row(ws, row, snapshot, values, max_col):
    ws.row_dimensions[row].height = snapshot.get("height")
    for col in range(1,max_col+1):
        c=ws.cell(row,col)
        info=snapshot["cells"][col-1]
        c._style=copy(info["style"])
        c.number_format=info["number_format"]
        c.alignment=copy(info["alignment"])
        c.protection=copy(info["protection"])
        c.value = values[col-1] if col-1 < len(values) else None


def _category(item): return _norm(item.get("categoria") or "")
def _name(item): return _norm(item.get("produto") or "")
def _section_hint(item): return _norm(item.get("secao") or "")

def _is_beer(item): return "CERVEJA" in _name(item)
def _is_cleaning(item):
    if "LIMPEZA" in _category(item): return True
    terms=("DETERGENTE","AMACIANTE","DESINFETANTE","AGUA SANITARIA","SABAO","LIMPADOR","ESPONJA","RODO","VASSOURA","SAPONACEO","SACO LIXO","CERA ")
    return any(t in _name(item) for t in terms)
def _is_hortifruti(item):
    if "HORTIFRUTI" in _category(item): return True
    terms=("BANANA","ABACAXI","ABACATE","ALHO ","BATATA","BETERRABA","CENOURA","CEBOLA","CHUCHU","LARANJA","LIMAO","MACA ","MAMAO","MANGA","MELANCIA","MORANGO","PERA","REPOLHO","TOMATE","UVA ","VAGEM","BROCOLIS","GUARIROBA","COCO VERDE")
    return any(t in _name(item) for t in terms)
def _is_meat(item):
    if "ACOUGUE" in _category(item) or "AÇOUGUE" in str(item.get("categoria") or "").upper(): return True
    terms=("ACEM","ALMONDEGA","BACON","BIFAO","CARNE ","COSTELA","COXA ","COXAO","LINGUICA","LOMBO","MOCOTO","MUSCULO","PANCETA","PERNIL","PICANHA","SALSICHA","FRANGO","FILE DE TILAPIA")
    return any(t in _name(item) for t in terms)


def _row_values(profile, item, section="main"):
    code=item.get("codigo","")
    unit=item.get("unidade","UN") or "UN"
    prod=item.get("produto","")
    custo=_num(item.get("custo"))
    venda=_num(item.get("varejo"))
    promo=_num(item.get("promocao"))
    clube=_num(item.get("clube"))
    limite=item.get("limite","") or None
    if profile=="SEGUNDA":
        return [code,unit,prod,custo,venda,promo,clube,None,limite]
    if profile=="TERCA":
        return [code,unit,prod,venda,promo,None,None,None,limite]
    if profile=="QUARTA":
        return [code,unit,prod,custo,venda,promo]
    if profile=="QUINTA":
        if section=="cervejas": return [code,unit,prod,promo,clube,venda,custo,None,None,None,None,limite]
        return [code,unit,prod,promo,venda,custo,None,None,None,None,limite]
    if profile=="FIM_SEMANA":
        return [code,unit,prod,custo,venda,promo,clube,None,limite]
    if profile=="CARNES_VERDURAS":
        return [code,unit,prod,custo,venda,promo,None,None,None,limite]
    if profile=="CLUBE":
        return [code,unit,prod,custo,venda,clube or promo,limite]
    return [code,unit,prod,custo,venda,promo,clube,None,limite]


def _sections(profile, items):
    items=list(items or [])
    if profile=="SEGUNDA":
        out={"limpeza":[],"economia":[],"cervejas":[],"scantech":[]}
        for x in items:
            hint=_section_hint(x)
            if "SCANTECH" in hint or "SCANTECH" in _name(x): out["scantech"].append(x)
            elif _is_beer(x): out["cervejas"].append(x)
            elif _is_cleaning(x): out["limpeza"].append(x)
            else: out["economia"].append(x)
        return out
    if profile=="QUINTA":
        return {"main":[x for x in items if not _is_beer(x)],"cervejas":[x for x in items if _is_beer(x)]}
    if profile=="FIM_SEMANA":
        main=[];bebidas=[];scantech=[]
        for x in items:
            hint=_section_hint(x)
            if "SCANTECH" in hint or "SCANTECH" in _name(x): scantech.append(x)
            elif _category(x)=="BEBIDAS" or _is_beer(x): bebidas.append(x)
            else: main.append(x)
        return {"main":main,"bebidas":bebidas,"scantech":scantech}
    if profile=="CARNES_VERDURAS":
        verduras=[];carnes=[]
        for x in items:
            (verduras if _is_hortifruti(x) and not _is_meat(x) else carnes).append(x)
        return {"verduras":verduras,"carnes":carnes}
    return {"main":items}


def _title(profile, validity, section="main"):
    v=str(validity or "").strip()
    if profile=="SEGUNDA":
        return {"limpeza":f"OFERTAS DA LIMPEZA SANJU {v}","economia":f"OFERTAS DA ECONOMIA SANJU {v}","cervejas":"CERVEJAS","scantech":"SCANTECH"}[section]
    if profile=="TERCA": return f"OFERTA TERÇA VERDE {v}"
    if profile=="QUARTA": return f"QUARTA CAFÉ COM PÃO SANJU {v}"
    if profile=="QUINTA": return "CERVEJAS" if section=="cervejas" else f"QUINTA FILÉ SANJU  {v}"
    if profile=="FIM_SEMANA":
        return {"main":f"FIM DE SEMANA  SANJU {v}","bebidas":"BEBIDAS","scantech":"SCANTECH"}[section]
    if profile=="CARNES_VERDURAS":
        return f"OFERTAS DAS {'VERDURAS' if section=='verduras' else 'CARNES'} SANJU  {v}"
    if profile=="CLUBE": return f"OFERTAS EXCLUSIVAS CLUBE SANJU {v}"
    return v


def _headers(profile, section="main"):
    if profile=="SEGUNDA": return ["EAN" if section in ("limpeza","economia") else "CÓDIGO","ENTRADA","PRODUTO" if section in ("limpeza","economia") else "PRODUTOS","CUSTO","VENDA","PROMOÇÃO","CLUBE","KD/BZ/SD","LIMITE"]
    if profile=="TERCA": return ["EAN","ENTRADA","PRODUTO","VENDA","PROMOÇÃO","KD","DIA","BZ","LIMITE"]
    if profile=="QUARTA": return ["CÓDIGO","ENT","PRODUTO","CUSTO","VENDA","PROMOÇÃO"]
    if profile=="QUINTA":
        if section=="cervejas": return ["CÓDIGO","ENTRADA","PRODUTO","PROMOÇÃO","CLUBE","VENDA","CUSTO","BZ","KD","PAI","DIA","LIMITE"]
        return ["CÓDIGO","ENTRADA","PRODUTO","PROMOÇÃO","VENDA","CUSTO","BZ","KD","PAI","DIA","LIMITE"]
    if profile=="FIM_SEMANA": return ["CÓDIGO","ENTRADA","PRODUTOS","CUSTO","VENDA","PROMOÇÃO","CLUBE","KD/SD/BZ","LIMITE"]
    if profile=="CARNES_VERDURAS": return ["EAN","ENTRADA","PRODUTO","CUSTO","VENDA","PROMOÇÃO","KD","DIA","BZ/PV" if section=="carnes" else "BZ","LIMITE"]
    if profile=="CLUBE": return ["CÓDIGO","ENTRADA","PRODUTO","CUSTO","VENDA","CLUBE","LIMITE"]
    return []


def export_campaign_xlsx(path, campaign, validity, items, profile_choice="AUTOMÁTICO"):
    profile=resolve_profile(profile_choice,campaign)
    template=TEMPLATE_DIR/TEMPLATES[profile]
    if not template.exists(): raise FileNotFoundError(f"Modelo de planilha não encontrado: {template.name}")
    wb=load_workbook(template)
    ws=wb.worksheets[0]
    max_col={"SEGUNDA":9,"TERCA":9,"QUARTA":6,"QUINTA":12,"FIM_SEMANA":9,"CARNES_VERDURAS":10,"CLUBE":7}[profile]

    # Captura estilos reais antes de reconstruir o conteúdo variável.
    if profile=="QUINTA":
        title_snap=_style_snapshot(ws,7,max_col);header_snap=_style_snapshot(ws,8,max_col);data_snap=_style_snapshot(ws,9,max_col)
        beer_title_snap=_style_snapshot(ws,39,max_col);beer_header_snap=_style_snapshot(ws,40,max_col);beer_data_snap=_style_snapshot(ws,41,max_col)
        start=7
    elif profile=="CARNES_VERDURAS":
        title_snap=_style_snapshot(ws,1,max_col);header_snap=_style_snapshot(ws,2,max_col);data_snap=_style_snapshot(ws,3,max_col)
        beer_title_snap=_style_snapshot(ws,30,max_col);beer_header_snap=_style_snapshot(ws,31,max_col);beer_data_snap=_style_snapshot(ws,32,max_col)
        start=1
    else:
        title_snap=_style_snapshot(ws,1,max_col);header_snap=_style_snapshot(ws,2,max_col);data_snap=_style_snapshot(ws,3,max_col)
        beer_title_snap=title_snap;beer_header_snap=header_snap;beer_data_snap=data_snap;start=1

    if ws.max_row>=start: ws.delete_rows(start, ws.max_row-start+1)
    sections=_sections(profile,items)
    row=start
    order={
        "SEGUNDA":["limpeza","economia","cervejas","scantech"],
        "TERCA":["main"],"QUARTA":["main"],"QUINTA":["main","cervejas"],
        "FIM_SEMANA":["main","bebidas","scantech"],"CARNES_VERDURAS":["verduras","carnes"],"CLUBE":["main"]
    }[profile]
    for si,section in enumerate(order):
        rows=sections.get(section,[])
        # Seções auxiliares sem itens não poluem a planilha; as seções principais permanecem.
        if not rows and section in {"cervejas","bebidas","scantech"}: continue
        ts=beer_title_snap if section in {"cervejas","bebidas","scantech","carnes"} else title_snap
        hs=beer_header_snap if section in {"cervejas","bebidas","scantech","carnes"} else header_snap
        ds=beer_data_snap if section in {"cervejas","bebidas","scantech","carnes"} else data_snap
        _apply_row(ws,row,ts,[_title(profile,validity,section)],max_col);row+=1
        _apply_row(ws,row,hs,_headers(profile,section),max_col);row+=1
        for item in rows:
            _apply_row(ws,row,ds,_row_values(profile,item,section),max_col);row+=1
        if si < len(order)-1: row+=1

    # Garante que o arquivo saia com o mesmo nome/estrutura de aba do modelo.
    wb.save(path)
    return str(path), PROFILE_LABELS[profile]
