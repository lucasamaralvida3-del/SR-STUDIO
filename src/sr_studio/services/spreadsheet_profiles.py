from __future__ import annotations

import hashlib
import re
import unicodedata
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from data.v5_store import connection, json_dumps, json_loads, now_iso, uid

FIELDS = (
    "code",
    "ean",
    "name",
    "promo_price",
    "app_price",
    "retail_price",
    "cost",
    "entry",
    "unit",
    "limit",
    "category",
    "highlight",
)

FIELD_LABELS = {
    "code": "Código",
    "ean": "EAN / Código de barras",
    "name": "Nome / Descrição",
    "promo_price": "Preço promoção",
    "app_price": "Preço APP / Clube",
    "retail_price": "Preço varejo",
    "cost": "Custo",
    "entry": "Entrada",
    "unit": "Unidade",
    "limit": "Limite por CPF",
    "category": "Categoria",
    "highlight": "Destaque",
}

ALIASES = {
    "code": ("CODIGO", "COD", "COD PRODUTO", "CODIGO PRODUTO", "ID PRODUTO", "PLU"),
    "ean": ("EAN", "GTIN", "COD BARRAS", "CODIGO DE BARRAS", "COD BARRA", "BARRAS"),
    "name": ("PRODUTO", "PRODUTOS", "NOME", "DESCRICAO", "DESCRICAO PRODUTO", "ITEM", "MERCADORIA"),
    "promo_price": ("PRECO PROMOCAO", "PROMOCAO", "PRECO PROMOCIONAL", "PRECO OFERTA", "OFERTA"),
    "app_price": ("PRECO APP", "APP", "CLUBE", "PRECO CLUBE", "CLUBE EXCLUSIVO"),
    "retail_price": ("PRECO", "VAREJO", "PRECO VAREJO", "VENDA", "PRECO VENDA"),
    "cost": ("CUSTO", "CUSTO REPOSICAO", "CUSTO GERENCIAL", "ULTIMO CUSTO"),
    "entry": ("ENTRADA", "TIPO ENTRADA", "UN ENTRADA"),
    "unit": ("UNIDADE", "UN", "UND", "UN VENDA", "UNIDADE VENDA"),
    "limit": ("LIMITE", "LIMITE CPF", "LIMITE POR CPF", "QTD LIMITE", "QUANTIDADE LIMITE"),
    "category": ("CATEGORIA", "DEPARTAMENTO", "SECAO", "SEÇÃO", "SETOR", "GRUPO"),
    "highlight": ("DESTAQUE", "DESTACAR", "PRIORIDADE", "PRIORITARIO", "PRIORITÁRIO", "CAPA", "OFERTA DESTAQUE"),
}


def _norm(value: Any) -> str:
    text = unicodedata.normalize("NFD", str(value or ""))
    text = "".join(c for c in text if unicodedata.category(c) != "Mn")
    return re.sub(r"[^A-Z0-9]+", " ", text.upper()).strip()


def _cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _truthy(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    n = _norm(value)
    return n in {"1", "SIM", "S", "YES", "TRUE", "X", "DESTAQUE", "DESTACAR", "CAPA", "PRIORIDADE", "PRIORITARIO"}


def header_signature(headers: list[str]) -> str:
    normalized = "|".join(_norm(x) for x in headers if _norm(x))
    return hashlib.sha1(normalized.encode("utf-8")).hexdigest()


def suggest_mapping(headers: list[str]) -> dict[str, str]:
    normalized = {_norm(h): h for h in headers if _norm(h)}
    mapping: dict[str, str] = {}
    used: set[str] = set()
    for field, aliases in ALIASES.items():
        best = ""
        best_score = -1
        for key, original in normalized.items():
            if original in used:
                continue
            for alias in aliases:
                a = _norm(alias)
                score = 0
                if key == a:
                    score = 100
                elif a in key or key in a:
                    score = 70 - abs(len(key) - len(a))
                else:
                    at = set(a.split())
                    kt = set(key.split())
                    overlap = len(at & kt)
                    if overlap:
                        score = overlap * 20 - abs(len(at) - len(kt)) * 2
                if score > best_score:
                    best_score, best = score, original
        if best and best_score >= 35:
            mapping[field] = best
            used.add(best)
    return mapping


def _header_score(values: list[Any]) -> float:
    score = 0.0
    for value in values:
        n = _norm(value)
        if not n:
            continue
        score += 0.05
        if any(n == _norm(a) for aliases in ALIASES.values() for a in aliases):
            score += 3.0
        elif any(_norm(a) in n for aliases in ALIASES.values() for a in aliases):
            score += 1.0
    return score


def inspect_workbook(path: str | Path) -> dict[str, Any]:
    path = Path(path)
    if not path.is_file():
        raise FileNotFoundError(path)
    wb = load_workbook(path, read_only=True, data_only=True)
    sheets = []
    try:
        for ws in wb.worksheets:
            sample: list[list[Any]] = []
            for row in ws.iter_rows(min_row=1, max_row=min(20, ws.max_row or 20), values_only=True):
                sample.append(list(row))
            if not sample:
                continue
            scores = [_header_score(row) for row in sample]
            index = max(range(len(scores)), key=lambda i: scores[i])
            headers = [_cell(v) for v in sample[index]]
            while headers and not headers[-1]:
                headers.pop()
            mapping = suggest_mapping(headers)
            sheets.append(
                {
                    "name": ws.title,
                    "header_row": index + 1,
                    "headers": headers,
                    "header_signature": header_signature(headers),
                    "suggested_mapping": mapping,
                    "score": round(scores[index], 2),
                    "max_row": int(ws.max_row or 0),
                    "max_column": int(ws.max_column or 0),
                }
            )
    finally:
        wb.close()
    if not sheets:
        raise ValueError("Nenhuma aba utilizável foi encontrada na planilha.")
    sheets.sort(key=lambda x: x["score"], reverse=True)
    return {"path": str(path), "file": path.name, "sheets": sheets, "best": sheets[0]}


def save_profile(name: str, sheet_name: str, header_row: int, headers: list[str], mapping: dict[str, str], profile_id: str = "") -> dict[str, Any]:
    clean_mapping = {k: str(v) for k, v in mapping.items() if k in FIELDS and str(v).strip()}
    if not clean_mapping.get("name") and not clean_mapping.get("code") and not clean_mapping.get("ean"):
        raise ValueError("O perfil precisa mapear pelo menos Nome, Código ou EAN.")
    pid = profile_id or uid("sheet")
    now = now_iso()
    signature = header_signature(headers)
    with connection() as con:
        con.execute(
            """INSERT INTO spreadsheet_profiles(id,name,header_signature,sheet_name,header_row,mapping_json,created_at,updated_at,last_used)
               VALUES(?,?,?,?,?,?,?,?,?)
               ON CONFLICT(id) DO UPDATE SET name=excluded.name,header_signature=excluded.header_signature,
                 sheet_name=excluded.sheet_name,header_row=excluded.header_row,mapping_json=excluded.mapping_json,updated_at=excluded.updated_at""",
            (pid, str(name or "Perfil de planilha").strip(), signature, sheet_name, int(header_row), json_dumps(clean_mapping), now, now, now),
        )
    return get_profile(pid) or {}


def get_profile(profile_id: str) -> dict[str, Any] | None:
    with connection() as con:
        r = con.execute("SELECT * FROM spreadsheet_profiles WHERE id=?", (profile_id,)).fetchone()
    if not r:
        return None
    data = dict(r)
    data["mapping"] = json_loads(data.pop("mapping_json", "{}"), {})
    return data


def list_profiles() -> list[dict[str, Any]]:
    with connection() as con:
        rows = con.execute("SELECT * FROM spreadsheet_profiles ORDER BY last_used DESC,updated_at DESC,name COLLATE NOCASE").fetchall()
    out = []
    for r in rows:
        d = dict(r)
        d["mapping"] = json_loads(d.pop("mapping_json", "{}"), {})
        out.append(d)
    return out


def match_profile(headers: list[str]) -> dict[str, Any] | None:
    sig = header_signature(headers)
    with connection() as con:
        r = con.execute("SELECT * FROM spreadsheet_profiles WHERE header_signature=? ORDER BY last_used DESC LIMIT 1", (sig,)).fetchone()
    if not r:
        return None
    data = dict(r)
    data["mapping"] = json_loads(data.pop("mapping_json", "{}"), {})
    return data


def delete_profile(profile_id: str) -> None:
    with connection() as con:
        con.execute("DELETE FROM spreadsheet_profiles WHERE id=?", (profile_id,))


def _infer_unit(entry: Any, explicit: Any = "") -> str:
    e = _norm(explicit)
    if e:
        if "KG" in e or "KILO" in e:
            return "KG"
        if "UN" in e or "UND" in e:
            return "UN"
        return e[:8]
    n = _norm(entry)
    if "KG" in n or "KILO" in n or "PESO" in n:
        return "KG"
    if "UN" in n or "UND" in n or "UNIDADE" in n:
        return "UN"
    if isinstance(entry, (float, int)) and not float(entry).is_integer():
        return "KG"
    return "UN"


def _mapped(row: dict[str, Any], mapping: dict[str, str]) -> dict[str, Any]:
    result = {field: row.get(header, "") for field, header in mapping.items() if field in FIELDS}
    result["code"] = _cell(result.get("code"))
    result["ean"] = re.sub(r"\D+", "", _cell(result.get("ean")))
    result["name"] = _cell(result.get("name"))
    result["limit"] = _cell(result.get("limit"))
    result["category"] = _cell(result.get("category"))
    result["unit"] = _infer_unit(result.get("entry"), result.get("unit"))
    result["highlight"] = _truthy(result.get("highlight"))
    return result


def read_rows(path: str | Path, profile: dict[str, Any], limit: int = 0) -> list[dict[str, Any]]:
    path = Path(path)
    mapping = profile.get("mapping") or json_loads(profile.get("mapping_json"), {}) or {}
    sheet_name = str(profile.get("sheet_name") or "")
    header_row = max(1, int(profile.get("header_row") or 1))
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
        header_values = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
        headers = [_cell(v) for v in header_values]
        out = []
        for values in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not any(v not in (None, "") for v in values):
                continue
            row = {headers[i]: values[i] if i < len(values) else "" for i in range(len(headers)) if headers[i]}
            item = _mapped(row, mapping)
            if not (item.get("name") or item.get("code") or item.get("ean")):
                continue
            out.append(item)
            if limit and len(out) >= int(limit):
                break
    finally:
        wb.close()
    if profile.get("id"):
        with connection() as con:
            con.execute("UPDATE spreadsheet_profiles SET last_used=? WHERE id=?", (now_iso(), profile["id"]))
    return out


def preview(path: str | Path, profile: dict[str, Any], limit: int = 30) -> dict[str, Any]:
    rows = read_rows(path, profile, limit=limit)
    issues = []
    for i, item in enumerate(rows, 1):
        if not item.get("name"):
            issues.append({"row": i, "field": "name", "message": "Produto sem nome."})
        price = item.get("promo_price")
        if price in (None, ""):
            issues.append({"row": i, "field": "promo_price", "message": "Preço promocional vazio."})
    return {"rows": rows, "issues": issues, "count": len(rows), "profile": profile}
