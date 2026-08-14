# -*- coding: utf-8 -*-
import os
import sys
import re
import json
import shutil
import tempfile
import traceback
import zipfile
import subprocess
import unicodedata
import threading
import queue
import time
import math
from pathlib import Path
from decimal import Decimal, InvalidOperation
from datetime import datetime

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
try:
    from PIL import Image, ImageTk
except Exception:
    Image = None
    ImageTk = None
try:
    from tkinterdnd2 import TkinterDnD
    APP_TK_BASE=TkinterDnD.Tk
except Exception:
    APP_TK_BASE=tk.Tk

import AtacadoModule as _atacado_mod
import ManualModule as _manual_mod
from AtacadoModule import AtacadoPanel, reports_history, run_engine as atacado_run_engine, mark_posters_generated, preload_atacado_catalog
from ManualModule import ManualPanel, preload_sale_catalog
from PromotionLibrary import PromotionLibraryPanel, preload_library_cache
from ProductOrganizer import ProductOrganizerPanel, catalog_counts, rebuild_catalog, preload_product_catalog, invalidate_catalog_cache
from PromotionBuilder import PromotionBuilderPanel, mark_campaign_status_from_jobs, preload_builder_catalog, invalidate_builder_catalog_cache
from Encartes3Engine import Encartes3Panel as EncartePanel, preload_encarte3_data as preload_encarte_data
from SRIAEngine import SRIAPanel, build_sria_settings_card, preload_sria_data
from CISSProductSync import import_report_208 as import_ciss_report_208, last_import_info as ciss_last_import_info, current_product_snapshot as ciss_current_product_snapshot
from UpdateManager import inspect_update, apply_update
from ui_v2 import (choose_palette, install_centered_messageboxes, add_tooltip, ToastManager, add_hover,
                   apply_scaling, default_printer_name, print_pdf, center_toplevel, cache_key, file_signature, install_parented_filedialogs, SRDialog)
from SRSpellCheck import correct_campaign_text
from SRStudio21 import (
    normalize_product_name, apply_learned_correction, learn_correction, corrections,
    validate_promo_jobs, verification_counts, PreGenerationDialog,
    record_product_jobs, search_product_history, product_history,
    queue_add, queue_load, queue_save, queue_update, queue_remove, queue_pending, queue_clear_done,
    smart_queue_jobs, record_reprint, reprint_items, dated_output_dir, smart_pdf_name, unique_path,
    list_printers, load_print_profiles, save_print_profiles, print_with_profile, pdf_with_copies,
    backup_model_version, model_versions, restore_model_version, cleanup_temp, enable_drop,
    default_output_root
)

try:
    from openpyxl import load_workbook
    from pypdf import PdfReader, PdfWriter
except ImportError as e:
    root = tk.Tk(); root.withdraw()
    messagebox.showerror(
        "Dependência não instalada",
        "Execute primeiro CORRIGIR_E_EXECUTAR.bat.\n\n" + str(e)
    )
    raise

APP_DIR = Path(__file__).resolve().parent

def _load_sr_version_info():
    """Usa a versao central do Launcher/instalador e mantem fallback local."""
    info = {
        "product_version": "4.0.4",
        "channel": "beta",
        "release_label": "Beta 5",
        "distribution_version": "4.0.4-hybrid.beta5",
    }
    candidates = []
    try:
        _localapp = Path(os.environ.get("LOCALAPPDATA", ""))
        if str(_localapp):
            candidates.extend([
                _localapp / "SRStudio" / "Config" / "version.json",
                _localapp / "SRStudio" / "Config" / "installed.json",
            ])
    except Exception:
        pass
    candidates.append(APP_DIR / "version.json")
    for _path in candidates:
        try:
            if not _path.exists():
                continue
            _data = json.loads(_path.read_text(encoding="utf-8-sig"))
            _raw_version = str(_data.get("product_version") or _data.get("version") or "").strip()
            _m = re.search(r"\d+\.\d+\.\d+", _raw_version)
            if _m:
                info["product_version"] = _m.group(0)
            _dist = str(_data.get("distribution_version") or _data.get("version") or "").strip()
            if _dist:
                info["distribution_version"] = _dist
            _channel = str(_data.get("channel") or "").strip().lower()
            if _channel in ("beta", "stable"):
                info["channel"] = _channel
            _label = str(_data.get("release_label") or "").strip()
            if _label:
                info["release_label"] = _label
            elif _dist:
                if info["channel"] == "beta":
                    _bm = re.search(r"beta[._-]?(\d+)", _dist, re.I)
                    info["release_label"] = f"Beta {_bm.group(1)}" if _bm else "Beta"
                elif info["channel"] == "stable":
                    _sm = re.search(r"stable[._-]?(\d+)", _dist, re.I)
                    info["release_label"] = f"Stable {_sm.group(1)}" if _sm else "Stable"
            # installed.json e a fonte de verdade apos o Launcher concluir uma atualizacao.
            if _path.name.lower() == "installed.json":
                break
        except Exception:
            continue
    return info

APP_VERSION_INFO = _load_sr_version_info()
APP_VERSION = APP_VERSION_INFO["product_version"]
APP_CHANNEL = APP_VERSION_INFO.get("channel", "")
APP_RELEASE_LABEL = APP_VERSION_INFO.get("release_label", "")
APP_DISPLAY_VERSION = APP_VERSION + (f" • {APP_RELEASE_LABEL}" if APP_RELEASE_LABEL else "")
ASSETS = APP_DIR / "assets"
BRAND_LOGO = ASSETS / "SR_logo.png"
BRAND_ICON = ASSETS / "SR_Studio.ico"

def _brand_photo(master, size):
    """Carrega a logo oficial do SR Studio preservando proporção e suavidade."""
    try:
        if BRAND_LOGO.exists() and Image is not None and ImageTk is not None:
            with Image.open(BRAND_LOGO) as src:
                img=src.convert("RGBA").resize((int(size),int(size)),Image.Resampling.LANCZOS)
            return ImageTk.PhotoImage(img, master=master)
    except Exception:
        pass
    try:
        raw=tk.PhotoImage(master=master,file=str(BRAND_LOGO))
        scale=max(1,math.ceil(max(raw.width(),raw.height())/max(1,int(size))))
        return raw.subsample(scale,scale)
    except Exception:
        return None

MODELS = APP_DIR / "modelos"
ORIGINAL_MODELS = MODELS / "originais"
MODEL1 = MODELS / "SEGUNDA_DA_LIMPEZA_1_PRECO.pptx"
MODEL2 = MODELS / "SEGUNDA_DA_LIMPEZA_2_PRECOS.pptx"
MODEL1_LIMIT = MODELS / "SEGUNDA_DA_LIMPEZA_1_PRECO_COM_LIMITE.pptx"
MODEL2_LIMIT = MODELS / "SEGUNDA_DA_LIMPEZA_2_PRECOS_COM_LIMITE.pptx"
ENGINE = APP_DIR / "PowerPointEngine.ps1"
PREVIEW_ENGINE = APP_DIR / "PreviewEngine.ps1"
VALIDATE_ENGINE = APP_DIR / "ValidateLayout.ps1"
ATACADO_MODEL = MODELS / "ATACADO.pptx"
CLUB_MODEL = MODELS / "CLUBE_EXCLUSIVO.pptx"
CLUB_MODEL_LIMIT = MODELS / "CLUBE_EXCLUSIVO_COM_LIMITE.pptx"
MODEL_SALE = MODELS / "CARTAZ_VENDA.pptx"

# Hybrid 4.0: modelos originais sao dados de referencia derivados dos modelos
# distribuidos. Se a pasta estiver ausente em um computador novo, recrie-a
# localmente sem precisar baixa-la do repositorio publico.
try:
    ORIGINAL_MODELS.mkdir(parents=True, exist_ok=True)
    for _model_path in (MODEL1, MODEL2, MODEL1_LIMIT, MODEL2_LIMIT, ATACADO_MODEL, CLUB_MODEL, CLUB_MODEL_LIMIT, MODEL_SALE):
        if _model_path.exists():
            _original_path = ORIGINAL_MODELS / _model_path.name
            if not _original_path.exists():
                shutil.copy2(_model_path, _original_path)
except Exception:
    pass

LOCAL_ROOT = Path(os.environ.get("LOCALAPPDATA", str(APP_DIR)))
OLD_LOCAL_DATA = LOCAL_ROOT / "SR_Studio_1.0"
LOCAL_DATA = LOCAL_ROOT / "SR_Studio_2.0"
LOCAL_DATA.mkdir(parents=True, exist_ok=True)

# Reaproveita histórico e preferências do 1.x na primeira abertura do 2.0.
for _name in ("historico.json", "sessao.json", "ui_settings.json"):
    _src = OLD_LOCAL_DATA / _name
    _dst = LOCAL_DATA / _name
    if not _dst.exists() and _src.exists():
        try: shutil.copy2(_src, _dst)
        except Exception: pass

HISTORY_FILE = LOCAL_DATA / "historico.json"
SESSION_FILE = LOCAL_DATA / "sessao.json"
UI_SETTINGS_FILE = LOCAL_DATA / "ui_settings.json"
RUNNING_FLAG = LOCAL_DATA / "srstudio_running.flag"
PENDING_STATE_FILE = LOCAL_DATA / "pending_state.json"
UPDATE_HISTORY_FILE = LOCAL_DATA / "update_history.json"

# Dados preparados durante a abertura para evitar consultas lentas ao clicar nas telas.
STARTUP_CACHE = {
    "health": {}, "default_printer": "Impressora padrão do Windows", "printers": [],
    "print_profiles": {}, "corrections_count": 0, "catalog_counts": {},
    "library_counts": {}, "sale_catalog_count": 0, "atacado_catalog_count": 0,
    "builder_catalog_count": 0, "product_catalog_count": 0, "library_cache_count": 0, "encarte_campaign_count": 0,
    "queue": [], "reprints": [], "atacado_reports": [], "ciss_last_import": {},
    "history_count": 0, "model_signatures": {}, "preloaded_at": ""
}


def unlock_sr_model_fields(path):
    """
    Remove a trava interna noTextEdit dos campos SR_* do PowerPoint.
    Alguns PCs/versões do Office ignoram essa trava; outros impedem
    qualquer alteração via COM. A correção não altera o visual do modelo.
    """
    path = Path(path)
    if not path.exists():
        return 0
    temp_path = path.with_suffix(".sr_unlock_tmp.pptx")
    changed = 0
    try:
        with zipfile.ZipFile(path, "r") as zin, zipfile.ZipFile(temp_path, "w", zipfile.ZIP_DEFLATED) as zout:
            shape_pattern = re.compile(rb'(<p:sp\b.*?</p:sp>)', re.S)
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename.startswith("ppt/slides/slide") and item.filename.endswith(".xml"):
                    def _unlock_block(match):
                        nonlocal changed
                        block = match.group(1)
                        if re.search(rb'\bname="SR_[^"]*"', block):
                            new_block, n = re.subn(
                                rb'\s+noTextEdit="(?:1|true)"',
                                b'',
                                block
                            )
                            changed += n
                            return new_block
                        return block
                    data = shape_pattern.sub(_unlock_block, data)
                zout.writestr(item, data)
        if changed:
            temp_path.replace(path)
        else:
            try:
                temp_path.unlink()
            except Exception:
                pass
        return changed
    except Exception:
        try:
            if temp_path.exists():
                temp_path.unlink()
        except Exception:
            pass
        return 0

def ensure_all_models_unlocked():
    total = 0
    for model in [MODEL1, MODEL2, MODEL1_LIMIT, MODEL2_LIMIT, ATACADO_MODEL, CLUB_MODEL, CLUB_MODEL_LIMIT, MODEL_SALE]:
        total += unlock_sr_model_fields(model)
    return total

try:
    _theme_pref=json.loads(UI_SETTINGS_FILE.read_text(encoding="utf-8-sig")).get("theme","Automático") if UI_SETTINGS_FILE.exists() else "Automático"
except Exception:
    _theme_pref="Automático"
_PAL = choose_palette(_theme_pref)
APP_BG=_PAL["APP_BG"]; SIDEBAR=_PAL["SIDEBAR"]; SIDEBAR_HOVER=_PAL["SIDEBAR_HOVER"]
CARD=_PAL["CARD"]; TEXT=_PAL["TEXT"]; MUTED=_PAL["MUTED"]; LINE=_PAL["LINE"]
BLUE=_PAL["BLUE"]; BLUE_2=_PAL["BLUE2"]; GREEN=_PAL["GREEN"]; GREEN_TXT=_PAL["GREEN_TXT"]
ORANGE=_PAL["ORANGE"]; ORANGE_TXT=_PAL["ORANGE_TXT"]; LIGHT_BLUE=_PAL["LIGHT_BLUE"]; LIGHT_BLUE_TXT=_PAL["LIGHT_BLUE_TXT"]
RED=_PAL["RED"]; RED_TXT=_PAL["RED_TXT"]; YELLOW=_PAL["YELLOW"]; YELLOW_TXT=_PAL["YELLOW_TXT"]
PURPLE=_PAL["PURPLE"]; PURPLE_TXT=_PAL["PURPLE_TXT"]; ROW_ALT=_PAL["ROW_ALT"]; SELECT_BG=_PAL["SELECT"]

VALIDITY_OPTIONS = ["VÁLIDO DE", "VÁLIDO SOMENTE"]
UNIT_OPTIONS = ["UN", "KG", "À LATA", "À GARRAFA"]
OUTPUT_OPTIONS = ["PDF ÚNICO", "SEPARADO POR CAMPANHA"]

def norm(v):
    s = "" if v is None else str(v)
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"[^A-Z0-9]+", "_", s.upper()).strip("_")

def to_decimal(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float, Decimal)):
        try:
            return Decimal(str(v))
        except Exception:
            return None
    s = str(v).strip().replace("R$", "").replace(" ", "")
    if not s:
        return None
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return Decimal(s)
    except InvalidOperation:
        return None

def money_str(v):
    d = to_decimal(v)
    if d is None:
        return str(v or "").strip()
    return f"{d:.2f}".replace(".", ",")

def same_price(a, b):
    x, y = to_decimal(a), to_decimal(b)
    if x is None or y is None:
        return str(a).strip() == str(b).strip()
    return abs(x - y) < Decimal("0.005")

def period_from(text):
    text = str(text or "")
    m = re.search(r"(\d{1,2})[/-](\d{1,2})\s*A\s*(\d{1,2})[/-](\d{1,2})[/-](\d{4})", text, re.I)
    if m:
        return f"{int(m.group(1)):02d}/{int(m.group(2)):02d} A {int(m.group(3)):02d}/{int(m.group(4)):02d}/{m.group(5)}"
    m = re.search(r"(\d{1,2})\s*A\s*(\d{1,2})[/-](\d{1,2})[/-](\d{4})", text, re.I)
    if m:
        return f"{int(m.group(1)):02d} A {int(m.group(2)):02d}/{int(m.group(3)):02d}/{m.group(4)}"
    m = re.search(r"(\d{1,2})[/-](\d{1,2})[/-](\d{4})", text)
    if m:
        return f"{int(m.group(1)):02d}/{int(m.group(2)):02d}/{m.group(3)}"
    return ""

def clean_campaign_title(text):
    s = " ".join(str(text or "").split())
    s = re.sub(r"\b\d{1,2}[/-]\d{1,2}\s*A\s*\d{1,2}[/-]\d{1,2}[/-]\d{4}\b.*$", "", s, flags=re.I).strip()
    s = re.sub(r"\b\d{1,2}\s*A\s*\d{1,2}[/-]\d{1,2}[/-]\d{4}\b.*$", "", s, flags=re.I).strip()
    s = re.sub(r"\b\d{1,2}[/-]\d{1,2}[/-]\d{4}\b.*$", "", s, flags=re.I).strip()
    s = re.sub(r"\bSANJU\b", "", s, flags=re.I)
    s = re.sub(r"\bSANTA\s+JULIANA\b", "", s, flags=re.I)
    s = " ".join(s.split()).strip(" -–—:")
    if not s:
        s = "OFERTA"
    # Correção de erro recorrente de digitação observado nas planilhas
    if norm(s) == "BEBDAS":
        s = "BEBIDAS"
    s = correct_campaign_text(s)
    return s.upper().rstrip("! ") + "!!"

def extract_date_tokens(text):
    text = str(text or "")
    patterns = [
        r"\d{1,2}[/-]\d{1,2}\s*A\s*\d{1,2}[/-]\d{1,2}[/-]\d{4}",
        r"\d{1,2}\s*A\s*\d{1,2}[/-]\d{1,2}[/-]\d{4}",
        r"\d{1,2}[/-]\d{1,2}[/-]\d{4}",
    ]
    hits = []
    for p in patterns:
        for m in re.findall(p, text, flags=re.I):
            hits.append(re.sub(r"\s+", " ", m.strip()).replace("-", "/").upper())
    return hits

def is_probable_title(text):
    if not text:
        return False
    n = norm(text)
    keywords = [
        "OFERTA","VERDE","LIMPEZA","ECONOMIA","CAFE","PAO","FILE",
        "FIM_DE_SEMANA","CERVEJA","CERVEJAS","BEBIDA","BEBIDAS","BEBDAS",
        "SCANTECH","CARNES","VERDURAS"
    ]
    return any(k in n for k in keywords) or bool(period_from(text))

def detect_unit_display(value):
    original = "" if value is None else str(value).strip()
    n = norm(original)
    if not n:
        return "UN", False
    if "GARRAFA" in n:
        return "À GARRAFA", True
    if "LATA" in n:
        return "À LATA", True
    if "KG" in n or "QUILO" in n or "KILO" in n or "PESO" in n:
        return "KG", True
    if n in {"UN","UND","UNID","UNIDADE","UNIDADES","CADA"} or "UNIDADE" in n or "CADA" in n:
        return "UN", True
    return "UN", False

def wrap_product_name(text, max_chars=30, max_lines=2):
    """Quebra o nome do produto em no máximo 2 linhas, buscando equilíbrio visual."""
    raw = " ".join(str(text or "").replace("\r", " ").replace("\n", " ").upper().split())
    if not raw:
        return ""
    if max_lines <= 1 or len(raw) <= max_chars or " " not in raw:
        return raw

    words = raw.split()
    if len(words) <= 1:
        return raw

    best = raw
    best_score = None
    for i in range(1, len(words)):
        left = " ".join(words[:i]).strip()
        right = " ".join(words[i:]).strip()
        if not left or not right:
            continue
        score = abs(len(left) - len(right))
        score += max(0, len(left) - max_chars) * 3
        score += max(0, len(right) - max_chars) * 3
        if len(left) < 7 or len(right) < 7:
            score += 8
        right_first = right.split()[0]
        if right_first in {"KG", "G", "GR", "L", "ML", "UN", "UND", "CX", "FD", "PCT"}:
            score += 4
        if right_first in {"DE", "DA", "DO", "DAS", "DOS", "COM", "E"}:
            score += 2
        if best_score is None or score < best_score:
            best_score = score
            best = left + "\r" + right
    return best

def text_fit_indicator(job):
    rendered = job.get("produto_render") or wrap_product_name(job.get("produto", ""))
    lines = max(1, rendered.count("\r") + 1)
    length = len(str(job.get("produto", "")))
    if job.get("layout_status") == "REVISAR":
        return "REVISAR"
    if job.get("layout_status") == "ERRO":
        return "ERRO"
    if lines <= 1:
        return "1 LINHA" if length <= 48 else "1 LINHA + REDUZ"
    if lines == 2:
        return "2 LINHAS" if length <= 64 else "2 LINHAS + REDUZ"
    return "REVISAR"

def merge_pdfs(files, output):
    writer = PdfWriter()
    for f in files:
        reader = PdfReader(str(f))
        for page in reader.pages:
            writer.add_page(page)
    with open(output, "wb") as h:
        writer.write(h)

def find_powershell():
    candidates = [
        shutil.which("powershell"),
        r"C:\Windows\System32\WindowsPowerShell\v1.0\powershell.exe",
    ]
    for c in candidates:
        if c and Path(c).exists():
            return c
    raise RuntimeError("Windows PowerShell não foi encontrado.")

def detect_header_rows(ws):
    rows = []
    for r in range(1, ws.max_row + 1):
        vals = [norm(ws.cell(r,c).value) for c in range(1, min(ws.max_column, 20)+1)]
        if ("PRODUTO" in vals or "PRODUTOS" in vals) and ("PROMOCAO" in vals or "CLUBE" in vals):
            rows.append(r)
    return rows

def find_title_for_header(ws, header_row):
    candidates = []
    for r in range(header_row - 1, max(0, header_row - 8), -1):
        vals = [ws.cell(r,c).value for c in range(1, min(ws.max_column, 8)+1)]
        texts = [str(v).strip() for v in vals if v is not None and str(v).strip()]
        if not texts:
            continue
        joined = " ".join(texts)
        score = 0
        if len(texts) == 1:
            score += 2
        if is_probable_title(joined):
            score += 6
        if period_from(joined):
            score += 3
        score += max(0, 8 - (header_row - r))
        candidates.append((score, r, joined))
    if candidates:
        candidates.sort(reverse=True)
        return candidates[0][2]
    return ws.title

def analyze_workbook(xlsx_path):
    wb = load_workbook(xlsx_path, data_only=True)
    jobs = []
    campaigns = []
    errors = []
    warnings = []
    skips = []
    file_dates = extract_date_tokens(Path(xlsx_path).stem)

    for ws in wb.worksheets:
        header_rows = detect_header_rows(ws)
        for idx, header_row in enumerate(header_rows):
            title_text = find_title_for_header(ws, header_row)
            campaign_name = clean_campaign_title(title_text)
            validity = period_from(title_text) or (file_dates[0] if file_dates else "")
            internal_dates = extract_date_tokens(title_text)

            if internal_dates and file_dates:
                internal = internal_dates[0].replace("-", "/")
                file_date = file_dates[0].replace("-", "/")
                if internal != file_date:
                    errors.append(
                        f"Data divergente em '{campaign_name}': enunciado='{internal}' / arquivo='{file_date}'. "
                        f"Corrija a planilha antes de gerar."
                    )

            headers = {}
            for c in range(1, ws.max_column + 1):
                key = norm(ws.cell(header_row, c).value)
                if key:
                    headers[key] = c

            cp = headers.get("PRODUTO") or headers.get("PRODUTOS")
            ccode = headers.get("EAN") or headers.get("CODIGO") or headers.get("COD")
            ccost = (headers.get("CUSTO") or headers.get("CUSTO UNITARIO") or headers.get("CUSTO_UNITARIO") or
                     headers.get("PRECO DE CUSTO") or headers.get("CUSTO GERENCIAL"))
            cvenda = (headers.get("VENDA") or headers.get("PRECO VENDA") or headers.get("PRECO_VENDA") or
                      headers.get("VAREJO") or headers.get("PRECO VAREJO"))
            cpromo = headers.get("PROMOCAO")
            cclube = headers.get("CLUBE")
            centr = headers.get("ENTRADA") or headers.get("ENT")
            climit = (
                headers.get("LIMITE") or headers.get("LIMIT") or
                headers.get("LIMITE_CPF") or headers.get("LIMITE_CLIENTE")
            )
            next_header = header_rows[idx + 1] if idx + 1 < len(header_rows) else ws.max_row + 1

            section_jobs = []
            sec_units = {"KG":0,"UN":0,"À LATA":0,"À GARRAFA":0}
            sec_one = sec_two = sec_club = 0

            for r in range(header_row + 1, next_header):
                product = ws.cell(r, cp).value if cp else None
                codigo_raw = ws.cell(r, ccode).value if ccode else None
                custo_raw = ws.cell(r, ccost).value if ccost else None
                venda_raw = ws.cell(r, cvenda).value if cvenda else None
                promo = ws.cell(r, cpromo).value if cpromo else None
                clube = ws.cell(r, cclube).value if cclube else None
                entrada = ws.cell(r, centr).value if centr else None
                limite_raw = ws.cell(r, climit).value if climit else None
                if isinstance(limite_raw, float) and limite_raw.is_integer():
                    limite = str(int(limite_raw))
                else:
                    limite = "" if limite_raw is None else str(limite_raw).strip()

                if product is None or not str(product).strip():
                    continue
                product_original=str(product).strip()
                product=apply_learned_correction(product_original)
                np = norm(product)
                if np in {"PRODUTO","PRODUTOS","CERVEJAS","BEBIDAS","BEBDAS","SCANTECH"} or np.startswith("OFERTA"):
                    continue

                promo_present = promo is not None and str(promo).strip()!=""
                clube_present = clube is not None and str(clube).strip()!=""
                if promo_present and not clube_present:
                    tipo = 1
                    promo_out = money_str(promo)
                    clube_out = ""
                elif promo_present and clube_present:
                    tipo = 1 if same_price(promo, clube) else 2
                    promo_out = money_str(promo)
                    clube_out = money_str(clube)
                elif (not promo_present) and clube_present:
                    tipo = 3  # Clube Exclusivo
                    promo_out = ""
                    clube_out = money_str(clube)
                else:
                    continue

                unit, recognized = detect_unit_display(entrada)
                issues = []
                if not recognized:
                    issues.append(f"Unidade '{entrada}' não reconhecida automaticamente; assumido UN.")
                if len(str(product).strip()) > 48:
                    issues.append("Nome longo: o sistema tentará ajustar em até 2 linhas automaticamente.")

                job = {
                    "id": len(jobs) + len(section_jobs) + 1,
                    "campanha": campaign_name,
                    "codigo": "" if codigo_raw is None else str(codigo_raw).strip().replace(".0","") if isinstance(codigo_raw,float) and codigo_raw.is_integer() else str(codigo_raw).strip(),
                    "produto_original": product_original,
                    "produto": normalize_product_name(product),
                    "produto_render": wrap_product_name(normalize_product_name(product)),
                    "custo": money_str(custo_raw) if custo_raw not in (None,"") else "",
                    "varejo": money_str(venda_raw) if venda_raw not in (None,"") else "",
                    "promocao": promo_out,
                    "clube": clube_out,
                    "validade": validity,
                    "tipo": tipo,
                    "entrada_original": "" if entrada is None else str(entrada).strip(),
                    "unidade_exibicao": unit,
                    "unidade_reconhecida": recognized,
                    "limite": limite,
                    "copies": 1,
                    "selected": True,
                    "issues": issues,
                    "status": "ATENÇÃO" if issues else "OK",
                    "layout_status": "",
                    "layout_detail": "",
                    "layout_font": 0,
                    "sheet": ws.title,
                    "linha": r,
                }
                section_jobs.append(job)
                if tipo == 1:
                    sec_one += 1
                elif tipo == 2:
                    sec_two += 1
                else:
                    sec_club += 1
                sec_units[unit] = sec_units.get(unit, 0) + 1

            if section_jobs:
                jobs.extend(section_jobs)
                campaigns.append({
                    "name": campaign_name,
                    "validity": validity,
                    "total": len(section_jobs),
                    "one": sec_one,
                    "two": sec_two,
                    "club": sec_club,
                    "units": sec_units,
                    "sheet": ws.title,
                })

    # Duplicados: código/EAN é o identificador principal; nome é fallback.
    seen = {}
    for job in jobs:
        identity=str(job.get("codigo") or "").strip() or norm(job.get("produto"))
        key=(norm(job.get("campanha")),identity)
        if key in seen:
            msg = f"Possível produto duplicado: {job['produto']} ({job['campanha']})."
            warnings.append(msg)
            job["issues"].append("Possível duplicidade na planilha/código.")
            job["status"] = "ATENÇÃO"
            seen[key]["issues"].append("Possível duplicidade na planilha/código.")
            seen[key]["status"] = "ATENÇÃO"
        else:
            seen[key] = job

    # Verificações comerciais do SR Studio.
    commercial_issues=validate_promo_jobs(jobs)
    for item in commercial_issues:
        j=next((x for x in jobs if x.get("id")==item.get("job_id")),None)
        if j:
            tag=f"[{item['severity']}] {item['message']}"
            if tag not in j["issues"]:j["issues"].append(tag)
            if item["severity"]=="CRÍTICO":j["status"]="CRÍTICO"
            elif j.get("status")!="CRÍTICO":j["status"]="ATENÇÃO"
        warnings.append(f"{item['produto']}: {item['message']}")

    for job in jobs:
        if job["issues"]:
            for issue in job["issues"]:
                if "Nome longo" not in issue:
                    warnings.append(f"{job['produto']}: {issue}")

    def dedupe(seq):
        out, seenx = [], set()
        for x in seq:
            if x not in seenx:
                seenx.add(x); out.append(x)
        return out

    errors = dedupe(errors)
    warnings = dedupe(warnings)
    skips = dedupe(skips)

    if not jobs and not skips:
        raise RuntimeError("Nenhum bloco compatível foi encontrado na planilha.")

    return {
        "jobs": jobs,
        "campaigns": campaigns,
        "errors": errors,
        "warnings": warnings,
        "skips": skips,
        "total": len(jobs),
        "one": sum(1 for j in jobs if j["tipo"] == 1),
        "two": sum(1 for j in jobs if j["tipo"] == 2),
        "club": sum(1 for j in jobs if j["tipo"] == 3),
        "sample": jobs[0] if jobs else None,
        "commercial_issues": commercial_issues,
    }

def _job_payload(job, validity_label):
    x = dict(job)
    rendered = str(job.get("produto_render") or job.get("produto") or "").strip()
    if ("\r" not in rendered) and ("\n" not in rendered):
        rendered = wrap_product_name(rendered)
    x["produto"] = rendered
    x["validade_rotulo"] = validity_label
    return x

def _powershell_base():
    return [find_powershell(), "-NoProfile", "-ExecutionPolicy", "Bypass"]

def generate_preview(job, validity_label):
    if not PREVIEW_ENGINE.exists():
        raise RuntimeError("PreviewEngine.ps1 não encontrado.")
    preview_dir = LOCAL_DATA / "preview"
    preview_dir.mkdir(parents=True, exist_ok=True)
    output_png = preview_dir / f"preview_{job.get('id','x')}.png"
    job_json = preview_dir / f"preview_{job.get('id','x')}.json"
    job_json.write_text(json.dumps(_job_payload(job, validity_label), ensure_ascii=False), encoding="utf-8")

    cmd = _powershell_base() + [
        "-File", str(PREVIEW_ENGINE),
        "-JobJson", str(job_json),
        "-OutputPng", str(output_png),
        "-Model1", str(MODEL1), "-Model2", str(MODEL2),
        "-Model1Limit", str(MODEL1_LIMIT), "-Model2Limit", str(MODEL2_LIMIT),
        "-ClubModel", str(CLUB_MODEL), "-ClubModelLimit", str(CLUB_MODEL_LIMIT),
        "-SaleModel", str(MODEL_SALE),
    ]
    proc = subprocess.run(
        cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
        text=True, encoding="utf-8", errors="replace",
        creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
    )
    if proc.returncode != 0 or not output_png.exists():
        raise RuntimeError("Não foi possível gerar a prévia real.\n\n" + proc.stdout[-1800:])
    return output_png

def validate_layout(jobs, validity_label, progress=None):
    if not VALIDATE_ENGINE.exists():
        raise RuntimeError("ValidateLayout.ps1 não encontrado.")
    if not jobs:
        return {}
    with tempfile.TemporaryDirectory(prefix="srstudio_validate_") as td:
        td = Path(td)
        jobs_json = td / "jobs.json"
        out_json = td / "layout.json"
        jobs_json.write_text(
            json.dumps([_job_payload(j, validity_label) for j in jobs], ensure_ascii=False),
            encoding="utf-8"
        )
        cmd = _powershell_base() + [
            "-File", str(VALIDATE_ENGINE),
            "-JobsJson", str(jobs_json), "-OutputJson", str(out_json),
            "-Model1", str(MODEL1), "-Model2", str(MODEL2),
            "-Model1Limit", str(MODEL1_LIMIT), "-Model2Limit", str(MODEL2_LIMIT),
        "-ClubModel", str(CLUB_MODEL), "-ClubModelLimit", str(CLUB_MODEL_LIMIT),
        ]
        proc = subprocess.Popen(
            cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
            text=True, encoding="utf-8", errors="replace",
            creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
        )
        done=0
        log=[]
        for line in proc.stdout:
            line=line.rstrip(); log.append(line)
            if line.startswith("CHECK|"):
                done += 1
                if progress:
                    progress(done, len(jobs), f"Verificando layout {done} de {len(jobs)}...")
        code=proc.wait()
        if code != 0 or not out_json.exists():
            raise RuntimeError("Falha na verificação visual do PowerPoint.\n\n" + "\n".join(log[-20:]))
        data=json.loads(out_json.read_text(encoding="utf-8-sig"))
        if isinstance(data, dict):
            data=[data]
        return {str(x.get("job_id")): x for x in data}


def _terminate_engine_process(proc, ppt_pid=None):
    """Encerra apenas o PowerShell e a instância de PowerPoint criada por esta geração."""
    try:
        if proc and proc.poll() is None:
            proc.terminate()
            try:
                proc.wait(timeout=1.5)
            except Exception:
                try: proc.kill()
                except Exception: pass
    except Exception:
        pass
    if ppt_pid and os.name == "nt":
        try:
            subprocess.run(
                ["taskkill","/PID",str(int(ppt_pid)),"/T","/F"],
                stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
                creationflags=getattr(subprocess,"CREATE_NO_WINDOW",0),
                timeout=4,
            )
        except Exception:
            pass

def generate_pdf(jobs, output, validity_label, progress=None, cancel_event=None):
    """
    V2.0 - geração estável com watchdog:
    1) tenta primeiro o lote inteiro em UMA sessão do PowerPoint
       (mesma arquitetura da versão que funcionava);
    2) somente se o lote realmente falhar, tenta isolar os produtos
       para descobrir qual cartaz causou a falha.
    """
    required_models = [MODEL1, MODEL2, MODEL1_LIMIT, MODEL2_LIMIT, CLUB_MODEL, CLUB_MODEL_LIMIT, MODEL_SALE]
    missing = [str(x.name) for x in required_models if not x.exists()]
    if missing:
        raise RuntimeError("Modelos PowerPoint não encontrados: " + ", ".join(missing))
    if not ENGINE.exists():
        raise RuntimeError("PowerPointEngine.ps1 não encontrado.")
    if not jobs:
        raise RuntimeError("Nenhum produto selecionado para gerar.")

    payloads=[_job_payload(job, validity_label) for job in jobs]

    def run_engine(payload_list, run_dir, progress_base=0, progress_total=None):
        run_dir=Path(run_dir)
        run_dir.mkdir(parents=True,exist_ok=True)
        pdfdir=run_dir/"pdfs"
        pdfdir.mkdir(exist_ok=True)
        jobs_json=run_dir/"jobs.json"
        jobs_json.write_text(json.dumps(payload_list,ensure_ascii=False),encoding="utf-8")

        cmd=_powershell_base()+[
            "-File",str(ENGINE),
            "-JobsJson",str(jobs_json),
            "-OutputDir",str(pdfdir),
            "-Model1",str(MODEL1),
            "-Model2",str(MODEL2),
            "-Model1Limit",str(MODEL1_LIMIT),
            "-Model2Limit",str(MODEL2_LIMIT),
            "-ClubModel",str(CLUB_MODEL),
            "-ClubModelLimit",str(CLUB_MODEL_LIMIT),
            "-SaleModel",str(MODEL_SALE),
        ]

        proc=subprocess.Popen(
            cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            encoding="utf-8",
            errors="replace",
            creationflags=getattr(subprocess,"CREATE_NO_WINDOW",0),
        )

        completed=0
        log=[]
        ok_pdfs=[]
        all_done_at=None
        batch_done_at=None
        ppt_pid=None
        forced_after_success=False
        line_queue=queue.Queue()
        def _reader():
            try:
                for _line in proc.stdout:
                    line_queue.put(_line.rstrip())
            finally:
                line_queue.put(None)
        threading.Thread(target=_reader,daemon=True).start()
        last_activity=time.time()
        stream_closed=False
        while True:
            if cancel_event and cancel_event.is_set():
                _terminate_engine_process(proc,ppt_pid)
                raise RuntimeError("Geração cancelada pelo usuário.")
            try:
                line=line_queue.get(timeout=0.25)
                if line is None:
                    stream_closed=True
                else:
                    last_activity=time.time()
                    log.append(line)
                    if line.startswith("PPTPID|"):
                        try: ppt_pid=int(line.split("|",1)[1].strip())
                        except Exception: ppt_pid=None
                    elif line.startswith("BATCH_DONE|"):
                        batch_done_at=time.time()
                        if progress:
                            total=progress_total or len(payload_list)
                            progress(progress_base+len(payload_list),total,"PDF criado • finalizando PowerPoint...")
                    elif line.startswith("ENGINE_DONE"):
                        pass
                    elif line.startswith("OK|"):
                        completed+=1
                        parts=line.split("|")
                        if len(parts)>=3:
                            try:
                                _p=Path(parts[2].strip())
                                if _p.exists(): ok_pdfs.append(_p)
                            except Exception: pass
                        if completed>=len(payload_list) and all_done_at is None:
                            all_done_at=time.time()
                        if progress:
                            total=progress_total or len(payload_list)
                            position=progress_base+completed
                            idx=min(position-1,len(jobs)-1)
                            name=jobs[idx]["produto"] if jobs else ""
                            progress(position,total,f"Gerando {position} de {total} • {name}")
            except queue.Empty:
                pass
            if proc.poll() is not None and stream_closed:
                break
            # Todos os PDFs já foram confirmados. Damos alguns segundos para o PowerPoint
            # encerrar a sessão COM; se ele ficar preso apenas no fechamento, não bloqueamos
            # o SR Studio indefinidamente.
            if batch_done_at is not None and proc.poll() is None and time.time()-batch_done_at > 1.5:
                forced_after_success=True
                _terminate_engine_process(proc,ppt_pid)
                break
            if all_done_at is not None and proc.poll() is None and time.time()-all_done_at > 4:
                forced_after_success=True
                _terminate_engine_process(proc,ppt_pid)
                break
            if time.time()-last_activity > 180:
                _terminate_engine_process(proc,ppt_pid)
                raise RuntimeError("O PowerPoint demorou mais que o esperado e foi encerrado com segurança. Tente novamente.")
        try: code=proc.wait(timeout=3)
        except Exception:
            code=proc.poll() if proc.poll() is not None else 1
        manifest=pdfdir/"manifest.txt"
        pdfs=[]
        if manifest.exists():
            pdfs=[Path(x.strip()) for x in manifest.read_text(encoding="utf-8-sig").splitlines() if x.strip()]
        existing=[p for p in pdfs if p.exists()]
        if len(existing)<len(payload_list):
            seen=set()
            existing=[]
            for _p in ok_pdfs:
                key=str(_p).lower()
                if _p.exists() and key not in seen:
                    seen.add(key); existing.append(_p)
        effective_code=0 if completed>=len(payload_list) and len(existing)>=len(payload_list) else code
        if forced_after_success:
            log.append("INFO|PowerPoint encerrado após concluir todos os PDFs; sessão COM não finalizou no tempo de tolerância.")
        return {"code":effective_code,"pdfs":existing,"log":log,"completed":completed}

    with tempfile.TemporaryDirectory(prefix="srstudio_batch_") as td:
        td=Path(td)

        # CAMINHO PRINCIPAL: todos os cartazes em uma única sessão PowerPoint.
        batch=run_engine(payloads,td/"batch",0,len(jobs))

        if batch["code"]==0 and len(batch["pdfs"])==len(jobs):
            merge_pdfs(batch["pdfs"],output)
            return {
                "success_count":len(batch["pdfs"]),
                "failed_jobs":[],
                "error_items":[],
                "output_created":True,
            }

        # Se o lote falhar, não descartamos tudo.
        # Entramos no diagnóstico isolado apenas neste caso.
        successful=[]
        failed=[]

        if progress:
            progress(0,len(jobs),"O lote encontrou uma falha. Identificando o cartaz com problema...")

        for i,(job,payload) in enumerate(zip(jobs,payloads),1):
            if cancel_event and cancel_event.is_set():
                raise RuntimeError("Geração cancelada pelo usuário.")

            single=run_engine([payload],td/f"isolado_{i:04d}",i-1,len(jobs))
            if single["code"]==0 and len(single["pdfs"])==1:
                successful.extend(single["pdfs"])
            else:
                jj=dict(job)
                details="\n".join(single["log"][-30:]).strip()
                if not details:
                    details="O PowerPoint não conseguiu exportar este cartaz e não retornou detalhes."
                jj["generation_error"]=details
                failed.append(jj)

        if successful:
            merge_pdfs(successful,output)

        return {
            "success_count":len(successful),
            "failed_jobs":failed,
            "error_items":[
                {
                    "job_id":str(j.get("id","")),
                    "produto":j.get("produto",""),
                    "campanha":j.get("campanha",""),
                    "message":j.get("generation_error",""),
                }
                for j in failed
            ],
            "output_created":bool(successful),
        }

def load_json(path, default):
    try:
        if path.exists():
            return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        pass
    return default

def save_json(path, data):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

def safe_name(s):
    s = re.sub(r'[<>:"/\\|?*]+', "_", str(s))
    s = re.sub(r"\s+", " ", s).strip()
    return s[:80] or "PROMOCAO"

class ReviewWindow(tk.Toplevel):
    """Revisão V2.0: edição direta, undo/redo, autosave, tabela/galeria e cache de miniaturas."""
    def __init__(self, master, analysis, on_change, focus_job_id=None):
        super().__init__(master)
        self.master_app=master; self.analysis=analysis; self.jobs=analysis["jobs"]; self.on_change=on_change
        self.palette=getattr(master,"palette",_PAL)
        self.title("SR Studio - Revisão de Cartazes")
        self.configure(bg=self.palette["APP_BG"]); self.minsize(920,600)
        center_toplevel(self,master,1280,760)
        self.search_var=tk.StringVar(); self.filter_var=tk.StringVar(value=getattr(master,"review_filter_override","TODOS")); master.review_filter_override="TODOS"; self.campaign_var=tk.StringVar(value="TODAS")
        self.density=tk.StringVar(value="Confortável"); self.view_mode=tk.StringVar(value="Tabela")
        self.counter_var=tk.StringVar(); self.unit_edit=tk.StringVar(value="UN"); self.limit_edit=tk.StringVar()
        self.undo_stack=[];self.redo_stack=[];self._gallery_images=[];self._gallery_token=0;self._thumb_queue=queue.Queue();self._thumb_worker_running=False
        self.thumb_dir=LOCAL_DATA/"thumb_cache";self.thumb_dir.mkdir(parents=True,exist_ok=True)
        self.build(); self.refresh_tree(); self.update_counter()
        if focus_job_id is not None and self.tree.exists(str(focus_job_id)):
            self.tree.selection_set(str(focus_job_id));self.tree.see(str(focus_job_id));self.details()
        self.bind("<Control-z>",lambda e:self.undo());self.bind("<Control-y>",lambda e:self.redo());self.bind("<Control-f>",lambda e:self.search_entry.focus_set())
        self.protocol("WM_DELETE_WINDOW",self.close_review)
    def build(self):
        pal=self.palette
        top=tk.Frame(self,bg=pal["CARD"],height=58,highlightbackground=pal["LINE"],highlightthickness=1);top.pack(fill="x");top.pack_propagate(False)
        tk.Label(top,text="Revisão dos cartazes",bg=pal["CARD"],fg=pal["TEXT"],font=("Segoe UI",15,"bold")).pack(side="left",padx=20)
        tk.Label(top,textvariable=self.counter_var,bg=pal["CARD"],fg=pal["BLUE2"],font=("Segoe UI",9,"bold")).pack(side="right",padx=20)

        controls=tk.Frame(self,bg=pal["APP_BG"]);controls.pack(fill="x",padx=18,pady=(10,7))
        tk.Label(controls,text="Pesquisar",bg=pal["APP_BG"],fg=pal["TEXT"],font=("Segoe UI",8,"bold")).pack(side="left")
        self.search_entry=tk.Entry(controls,textvariable=self.search_var,width=27,bg=pal["CARD"],fg=pal["TEXT"],insertbackground=pal["TEXT"],relief="flat")
        self.search_entry.pack(side="left",padx=(5,10),ipady=5);self.search_entry.bind("<KeyRelease>",lambda e:self.refresh_current())
        filt=ttk.Combobox(controls,textvariable=self.filter_var,state="readonly",width=19,
                          values=["TODOS","SOMENTE PROBLEMAS","CLUBE EXCLUSIVO","PROMO + CLUBE","1 PREÇO","COM LIMITE","SEM LIMITE","EDITADOS","TEXTO LONGO"])
        filt.pack(side="left",padx=(0,8));filt.bind("<<ComboboxSelected>>",lambda e:self.refresh_current())
        camps=["TODAS"]+sorted({j["campanha"] for j in self.jobs})
        camp=ttk.Combobox(controls,textvariable=self.campaign_var,state="readonly",values=camps,width=22);camp.pack(side="left");camp.bind("<<ComboboxSelected>>",lambda e:self.refresh_current())
        tk.Label(controls,text="Densidade",bg=pal["APP_BG"],fg=pal["MUTED"],font=("Segoe UI",8)).pack(side="right",padx=(8,4))
        dens=ttk.Combobox(controls,textvariable=self.density,state="readonly",values=["Confortável","Compacto"],width=12);dens.pack(side="right");dens.bind("<<ComboboxSelected>>",lambda e:self.apply_density())
        tk.Button(controls,text="GALERIA" if self.view_mode.get()=="Tabela" else "TABELA",command=self.toggle_view,bg=pal["LIGHT_BLUE"],fg=pal["LIGHT_BLUE_TXT"],relief="flat",font=("Segoe UI",8,"bold"),padx=10,pady=6).pack(side="right",padx=6)

        quick=tk.Frame(self,bg=pal["APP_BG"]);quick.pack(fill="x",padx=18,pady=(0,7))
        self.undo_btn=tk.Button(quick,text="↶ DESFAZER",command=self.undo,bg=pal["CARD"],fg=pal["TEXT"],relief="flat",font=("Segoe UI",8,"bold"),padx=10,pady=5);self.undo_btn.pack(side="left")
        self.redo_btn=tk.Button(quick,text="↷ REFAZER",command=self.redo,bg=pal["CARD"],fg=pal["TEXT"],relief="flat",font=("Segoe UI",8,"bold"),padx=10,pady=5);self.redo_btn.pack(side="left",padx=5)
        tk.Label(quick,text="Duplo clique para editar • botão direito para ações",bg=pal["APP_BG"],fg=pal["MUTED"],font=("Segoe UI",8)).pack(side="left",padx=10)
        tk.Button(quick,text="SELECIONAR VISÍVEIS",command=lambda:self.set_visible(True),bg=pal["LIGHT_BLUE"],fg=pal["LIGHT_BLUE_TXT"],relief="flat",font=("Segoe UI",8,"bold"),padx=9,pady=5).pack(side="right")
        tk.Button(quick,text="DESMARCAR VISÍVEIS",command=lambda:self.set_visible(False),bg=pal["RED"],fg=pal["RED_TXT"],relief="flat",font=("Segoe UI",8,"bold"),padx=9,pady=5).pack(side="right",padx=5)

        self.paned=tk.PanedWindow(self,orient="horizontal",bg=pal["APP_BG"],sashwidth=6,sashrelief="flat");self.paned.pack(fill="both",expand=True,padx=18,pady=(0,9))
        self.left=tk.Frame(self.paned,bg=pal["CARD"],highlightbackground=pal["LINE"],highlightthickness=1)
        self.right=tk.Frame(self.paned,bg=pal["CARD"],highlightbackground=pal["LINE"],highlightthickness=1,width=320)
        self.paned.add(self.left,stretch="always",minsize=590);self.paned.add(self.right,stretch="never",minsize=285)
        self.build_table();self.build_details()

        bottom=tk.Frame(self,bg=pal["CARD"],height=60,highlightbackground=pal["LINE"],highlightthickness=1);bottom.pack(fill="x");bottom.pack_propagate(False)
        self.validate_btn=tk.Button(bottom,text="VERIFICAR LAYOUT",command=self.validate_powerpoint_layout,bg=pal["ORANGE"],fg=pal["ORANGE_TXT"],relief="flat",font=("Segoe UI",8,"bold"),padx=12,pady=7);self.validate_btn.pack(side="left",padx=18,pady=10)
        tk.Button(bottom,text="FECHAR REVISÃO",command=self.close_review,bg=pal["BLUE"],fg="white",relief="flat",font=("Segoe UI",9,"bold"),padx=18,pady=8).pack(side="right",padx=18,pady=10)
        add_tooltip(self.undo_btn,"Desfaz a última edição feita diretamente na tabela (Ctrl+Z).")
        add_tooltip(self.redo_btn,"Refaz a última edição desfeita (Ctrl+Y).")
    def build_table(self):
        pal=self.palette
        self.table_holder=tk.Frame(self.left,bg=pal["CARD"]);self.table_holder.pack(fill="both",expand=True)
        cols=("sel","tipo","camp","prod","promo","clube","unit","limite","copies","atencao")
        self.tree=ttk.Treeview(self.table_holder,columns=cols,show="headings",selectmode="extended")
        heads={"sel":"✓","tipo":"Tipo","camp":"Campanha","prod":"Produto","promo":"Promoção","clube":"Clube","unit":"Unid.","limite":"Limite","copies":"Cópias","atencao":"Atenção"}
        widths={"sel":34,"tipo":115,"camp":125,"prod":280,"promo":76,"clube":76,"unit":60,"limite":80,"copies":55,"atencao":105}
        for c in cols:self.tree.heading(c,text=heads[c]);self.tree.column(c,width=widths[c],anchor="w" if c in {"tipo","camp","prod"} else "center")
        ys=ttk.Scrollbar(self.table_holder,orient="vertical",command=self.tree.yview);xs=ttk.Scrollbar(self.table_holder,orient="horizontal",command=self.tree.xview)
        self.tree.configure(yscrollcommand=ys.set,xscrollcommand=xs.set)
        self.tree.grid(row=0,column=0,sticky="nsew");ys.grid(row=0,column=1,sticky="ns");xs.grid(row=1,column=0,sticky="ew")
        self.table_holder.grid_rowconfigure(0,weight=1);self.table_holder.grid_columnconfigure(0,weight=1)
        self.tree.bind("<<TreeviewSelect>>",lambda e:self.details());self.tree.bind("<Double-1>",self.on_double_click);self.tree.bind("<Button-3>",self.popup_menu)
        self.tree.tag_configure("club",background=pal["YELLOW"]);self.tree.tag_configure("problem",background=pal["RED"]);self.tree.tag_configure("edited",background=pal["LIGHT_BLUE"])
        self.apply_density()
        self.gallery_holder=tk.Frame(self.left,bg=pal["CARD"])
    def build_details(self):
        pal=self.palette
        tk.Label(self.right,text="Detalhes do cartaz",bg=pal["CARD"],fg=pal["TEXT"],font=("Segoe UI",12,"bold")).pack(anchor="w",padx=15,pady=(14,5))
        self.type_chip=tk.Label(self.right,text="—",bg=pal["LIGHT_BLUE"],fg=pal["LIGHT_BLUE_TXT"],font=("Segoe UI",8,"bold"),padx=8,pady=4);self.type_chip.pack(anchor="w",padx=15,pady=(2,8))
        self.detail_product=tk.Label(self.right,text="Selecione um cartaz",bg=pal["CARD"],fg=pal["TEXT"],font=("Segoe UI",12,"bold"),wraplength=285,justify="left");self.detail_product.pack(anchor="w",padx=15)
        self.detail_values=tk.Label(self.right,text="",bg=pal["ROW_ALT"],fg=pal["TEXT"],font=("Segoe UI",9),justify="left",anchor="w",wraplength=285,padx=10,pady=10);self.detail_values.pack(fill="x",padx=15,pady=(10,7))
        tk.Label(self.right,text="Nível de atenção",bg=pal["CARD"],fg=pal["MUTED"],font=("Segoe UI",8,"bold")).pack(anchor="w",padx=15,pady=(3,2))
        self.attention_chip=tk.Label(self.right,text="OK",bg=pal["GREEN"],fg=pal["GREEN_TXT"],font=("Segoe UI",8,"bold"),padx=8,pady=4);self.attention_chip.pack(anchor="w",padx=15)
        self.alert_explain=tk.Label(self.right,text="Nenhum problema detectado.",bg=pal["CARD"],fg=pal["MUTED"],font=("Segoe UI",8),justify="left",wraplength=285);self.alert_explain.pack(anchor="w",padx=15,pady=(5,8))
        self.preview_btn=tk.Button(self.right,text="PRÉVIA REAL",command=self.real_preview,bg=pal["BLUE"],fg="white",relief="flat",font=("Segoe UI",8,"bold"),pady=8);self.preview_btn.pack(fill="x",padx=15,pady=(4,4))
        tk.Button(self.right,text="GERAR ESTE CARTAZ",command=lambda:self.generate_this(False),bg=pal["GREEN"],fg=pal["GREEN_TXT"],relief="flat",font=("Segoe UI",8,"bold"),pady=7).pack(fill="x",padx=15,pady=3)
        tk.Button(self.right,text="IMPRIMIR ESTE CARTAZ",command=lambda:self.generate_this(True),bg=pal["ORANGE"],fg=pal["ORANGE_TXT"],relief="flat",font=("Segoe UI",8,"bold"),pady=7).pack(fill="x",padx=15,pady=3)
        tk.Button(self.right,text="RESTAURAR ORIGINAL",command=self.restore_original,bg=pal["LIGHT_BLUE"],fg=pal["LIGHT_BLUE_TXT"],relief="flat",font=("Segoe UI",8,"bold"),pady=7).pack(fill="x",padx=15,pady=(3,12))
        add_tooltip(self.attention_chip,"CRÍTICO: impede/ameaça a geração. ATENÇÃO: revisar. INFO: alteração manual. OK: sem problemas conhecidos.")
        add_tooltip(self.alert_explain,"Aqui o SR Studio explica o motivo do alerta e o que deve ser conferido.")
    def kind_text(self,j):
        if j.get("tipo")==3:return "CLUBE EXCLUSIVO"
        if j.get("tipo")==2:return "PROMO + CLUBE"
        return "1 PREÇO"
    def attention(self,j):
        if j.get("layout_status")=="ERRO" or j.get("status")=="CRÍTICO" or any(str(x).startswith("[CRÍTICO]") for x in j.get("issues",[])):return "CRÍTICO"
        issues=[x for x in j.get("issues",[]) if "Nome longo" not in x]
        if j.get("layout_status")=="REVISAR" or issues or j.get("status") not in {"OK",None}:return "ATENÇÃO"
        if j.get("manual_edit"):return "INFO"
        return "OK"
    def visible_jobs(self):
        q=norm(self.search_var.get());f=self.filter_var.get();camp=self.campaign_var.get();out=[]
        for j in self.jobs:
            if q and q not in norm(j.get("produto")) and q not in norm(j.get("campanha")):continue
            if camp!="TODAS" and j.get("campanha")!=camp:continue
            if f=="SOMENTE PROBLEMAS" and self.attention(j) not in {"CRÍTICO","ATENÇÃO"}:continue
            if f=="CLUBE EXCLUSIVO" and j.get("tipo")!=3:continue
            if f=="PROMO + CLUBE" and j.get("tipo")!=2:continue
            if f=="1 PREÇO" and j.get("tipo")!=1:continue
            if f=="COM LIMITE" and not str(j.get("limite","")).strip():continue
            if f=="SEM LIMITE" and str(j.get("limite","")).strip():continue
            if f=="EDITADOS" and not j.get("manual_edit"):continue
            if f=="TEXTO LONGO" and text_fit_indicator(j) in {"1 LINHA","2 LINHAS"}:continue
            out.append(j)
        return out
    def refresh_current(self):
        if self.view_mode.get()=="Galeria":self.refresh_gallery()
        else:self.refresh_tree()
    def refresh_tree(self):
        if not hasattr(self,"tree"):return
        selected=set(self.tree.selection());self.tree.delete(*self.tree.get_children())
        for j in self.visible_jobs():
            att=self.attention(j);tag="edited" if j.get("manual_edit") else "club" if j.get("tipo")==3 else "problem" if att in {"CRÍTICO","ATENÇÃO"} else ""
            self.tree.insert("","end",iid=str(j["id"]),tags=(tag,) if tag else (),values=("✓" if j.get("selected") else "—",self.kind_text(j),j.get("campanha",""),j.get("produto",""),j.get("promocao") or "—",j.get("clube") or "—",j.get("unidade_exibicao",""),j.get("limite") or "—",j.get("copies",1),att))
        restore=[x for x in selected if self.tree.exists(x)]
        if restore:self.tree.selection_set(restore)
        self.update_counter()
    def apply_density(self):
        try:self.master_app.style.configure("Treeview",rowheight=24 if self.density.get()=="Compacto" else 31,font=("Segoe UI",8 if self.density.get()=="Compacto" else 9))
        except Exception:pass
    def selected_jobs(self):
        ids=set(self.tree.selection()) if self.view_mode.get()=="Tabela" else set()
        return [j for j in self.jobs if str(j.get("id")) in ids]
    def selected_job(self):
        a=self.selected_jobs();return a[0] if a else None
    def set_visible(self,v):
        for j in self.visible_jobs():j["selected"]=bool(v)
        self.save_changes();self.refresh_current()
    def on_double_click(self,event):
        region=self.tree.identify("region",event.x,event.y);iid=self.tree.identify_row(event.y);col=self.tree.identify_column(event.x)
        if region!="cell" or not iid:return
        idx=int(col[1:])-1;columns=("sel","tipo","camp","prod","promo","clube","unit","limite","copies","atencao");key=columns[idx]
        if key=="sel":
            j=next((x for x in self.jobs if str(x["id"])==iid),None)
            if j:j["selected"]=not j.get("selected",True);self.save_changes();self.refresh_tree()
            return
        if key not in {"prod","promo","clube","unit","limite","copies"}:return
        j=next((x for x in self.jobs if str(x["id"])==iid),None)
        if not j:return
        if key=="promo" and j.get("tipo")==3:return
        if key=="clube" and j.get("tipo")==1:return
        bbox=self.tree.bbox(iid,col)
        if not bbox:return
        x,y,w,h=bbox;field={"prod":"produto","promo":"promocao","clube":"clube","unit":"unidade_exibicao","limite":"limite","copies":"copies"}[key]
        old=str(j.get(field,"") or "")
        var=tk.StringVar(value=old)
        if key=="unit":editor=ttk.Combobox(self.tree,textvariable=var,state="readonly",values=UNIT_OPTIONS)
        else:editor=tk.Entry(self.tree,textvariable=var,bg=self.palette["CARD"],fg=self.palette["TEXT"],insertbackground=self.palette["TEXT"],relief="solid",bd=1)
        editor.place(x=x,y=y,width=w,height=h);editor.focus_set()
        if isinstance(editor,tk.Entry):editor.select_range(0,"end")
        done={"v":False}
        def commit(_=None):
            if done["v"]:return
            done["v"]=True;new=var.get().strip()
            editor.destroy()
            if new==old:return
            self.apply_edit(j,field,new,record=True)
        editor.bind("<Return>",commit);editor.bind("<FocusOut>",commit);editor.bind("<Escape>",lambda e:editor.destroy())
    def apply_edit(self,j,field,value,record=True):
        old=j.get(field,"")
        if "_original" not in j:j["_original"]={"produto":j.get("produto"),"promocao":j.get("promocao"),"clube":j.get("clube"),"unidade_exibicao":j.get("unidade_exibicao"),"limite":j.get("limite"),"copies":j.get("copies",1)}
        if record:self.undo_stack.append((j["id"],field,old,value));self.redo_stack.clear()
        if field in {"promocao","clube"}:value=money_str(value)
        if field=="limite":value=str(value).upper()
        if field=="copies":
            try:value=max(1,min(99,int(str(value).strip())))
            except Exception:value=1
        if field=="produto":value=normalize_product_name(value)
        j[field]=value;j["manual_edit"]=True
        if field=="produto":
            j["produto_render"]=wrap_product_name(value)
            if norm(old)!=norm(value) and messagebox.askyesno("Memória de correções",f"Usar esta correção automaticamente nas próximas importações?\n\n{old}\n→ {value}",parent=self):
                learn_correction(old,value)
        j["layout_status"]="";self.save_changes();self.refresh_current();self.details()
        try:self.master_app.toast.show("Revisão salva automaticamente.","ok",1700)
        except Exception:pass
    def undo(self):
        if not self.undo_stack:return
        jid,field,old,new=self.undo_stack.pop();j=next((x for x in self.jobs if x["id"]==jid),None)
        if j:self.redo_stack.append((jid,field,old,new));j[field]=old;j["manual_edit"]=True;self.save_changes();self.refresh_current();self.details()
    def redo(self):
        if not self.redo_stack:return
        jid,field,old,new=self.redo_stack.pop();j=next((x for x in self.jobs if x["id"]==jid),None)
        if j:self.undo_stack.append((jid,field,old,new));j[field]=new;j["manual_edit"]=True;self.save_changes();self.refresh_current();self.details()
    def restore_original(self):
        j=self.selected_job()
        if not j or not j.get("_original"):messagebox.showinfo("Revisão","Este item não possui alterações manuais para restaurar.");return
        original=j["_original"].copy()
        for f,v in original.items():j[f]=v
        j["produto_render"]=wrap_product_name(j["produto"]);j["manual_edit"]=False;j["layout_status"]="";self.save_changes();self.refresh_current();self.details()
    def details(self):
        j=self.selected_job()
        if not j:return
        pal=self.palette;kind=self.kind_text(j);self.type_chip.config(text=kind,bg=pal["YELLOW"] if j.get("tipo")==3 else pal["PURPLE"] if j.get("tipo")==2 else pal["LIGHT_BLUE"],fg=pal["YELLOW_TXT"] if j.get("tipo")==3 else pal["PURPLE_TXT"] if j.get("tipo")==2 else pal["LIGHT_BLUE_TXT"])
        self.detail_product.config(text=j.get("produto",""))
        values=[]
        if j.get("promocao"):values.append(f"Promoção: R$ {j['promocao']}")
        if j.get("clube"):values.append(f"Clube: R$ {j['clube']}")
        values.append(f"Unidade: {j.get('unidade_exibicao','')}")
        if j.get("validade"):values.append(f"Validade: {self.master_app.validity.get()} • {j['validade']}")
        if j.get("limite"):values.append(f"Limite: {j['limite']} por CPF")
        values.append(f"Cópias na impressão: {j.get('copies',1)}")
        if j.get("custo"):values.append(f"Custo: R$ {j['custo']}")
        if j.get("varejo"):values.append(f"Venda: R$ {j['varejo']}")
        if j.get("manual_edit") and j.get("_original"):
            changes=[]
            for f,label in [("produto","Produto"),("promocao","Promoção"),("clube","Clube"),("unidade_exibicao","Unidade"),("limite","Limite"),("copies","Cópias")]:
                a=str(j["_original"].get(f,"") or "");b=str(j.get(f,"") or "")
                if a!=b:changes.append(f"{label}: {a or '—'} → {b or '—'}")
            if changes:values.append("\nALTERAÇÃO MANUAL\n"+"\n".join(changes))
        self.detail_values.config(text="\n".join(values))
        att=self.attention(j);styles={"CRÍTICO":(pal["RED"],pal["RED_TXT"]),"ATENÇÃO":(pal["ORANGE"],pal["ORANGE_TXT"]),"INFO":(pal["LIGHT_BLUE"],pal["LIGHT_BLUE_TXT"]),"OK":(pal["GREEN"],pal["GREEN_TXT"])};bg,fg=styles[att];self.attention_chip.config(text=att,bg=bg,fg=fg)
        issues=list(j.get("issues",[]));
        if j.get("layout_detail"):issues.append(j["layout_detail"])
        if j.get("manual_edit"):issues.append("Item editado manualmente. O valor do cartaz pode ser diferente da planilha original.")
        self.alert_explain.config(text="\n".join("• "+x for x in issues) if issues else "Nenhum problema detectado.")
    def popup_menu(self,event):
        iid=self.tree.identify_row(event.y)
        if iid:
            self.tree.selection_set(iid);self.details()
        m=tk.Menu(self,tearoff=0)
        m.add_command(label="Prévia real",command=self.real_preview);m.add_command(label="Gerar este cartaz",command=lambda:self.generate_this(False));m.add_command(label="Imprimir este cartaz",command=lambda:self.generate_this(True));m.add_separator()
        m.add_command(label="Incluir / excluir",command=self.toggle_selected);m.add_command(label="Restaurar original",command=self.restore_original)
        try:m.tk_popup(event.x_root,event.y_root)
        finally:m.grab_release()
    def toggle_selected(self,event=None):
        js=self.selected_jobs()
        if not js:return
        v=not js[0].get("selected",True)
        for j in js:j["selected"]=v
        self.save_changes();self.refresh_tree()
    def toggle_view(self):
        if self.view_mode.get()=="Tabela":
            self.view_mode.set("Galeria");self.table_holder.pack_forget();self.gallery_holder.pack(fill="both",expand=True);self.refresh_gallery()
        else:
            self.view_mode.set("Tabela");self.gallery_holder.pack_forget();self.table_holder.pack(fill="both",expand=True);self.refresh_tree()
    def thumb_path(self,j):
        sig=cache_key(j.get("produto"),j.get("promocao"),j.get("clube"),j.get("limite"),j.get("unidade_exibicao"),j.get("validade"),self.master_app.validity.get(),
                      file_signature(CLUB_MODEL if j.get("tipo")==3 and not j.get("limite") else CLUB_MODEL_LIMIT if j.get("tipo")==3 else MODEL1 if j.get("tipo")==1 and not j.get("limite") else MODEL1_LIMIT if j.get("tipo")==1 else MODEL2 if not j.get("limite") else MODEL2_LIMIT))
        return self.thumb_dir/(sig+".png")
    def refresh_gallery(self):
        self._gallery_token+=1;token=self._gallery_token;self._gallery_images=[]
        for w in self.gallery_holder.winfo_children():w.destroy()
        canvas=tk.Canvas(self.gallery_holder,bg=self.palette["CARD"],highlightthickness=0);sb=ttk.Scrollbar(self.gallery_holder,orient="vertical",command=canvas.yview);canvas.configure(yscrollcommand=sb.set);canvas.pack(side="left",fill="both",expand=True);sb.pack(side="right",fill="y")
        inner=tk.Frame(canvas,bg=self.palette["CARD"]);win=canvas.create_window((0,0),window=inner,anchor="nw");inner.bind("<Configure>",lambda e:canvas.configure(scrollregion=canvas.bbox("all")));canvas.bind("<Configure>",lambda e:canvas.itemconfigure(win,width=e.width))
        jobs=self.visible_jobs()[:30]
        if not jobs:
            tk.Label(inner,text="Nenhum cartaz neste filtro.",bg=self.palette["CARD"],fg=self.palette["MUTED"],font=("Segoe UI",10)).pack(pady=40);return
        cols=3
        for i,j in enumerate(jobs):
            card=tk.Frame(inner,bg=self.palette["ROW_ALT"],highlightbackground=self.palette["LINE"],highlightthickness=1,width=210,height=300);card.grid(row=i//cols,column=i%cols,sticky="nsew",padx=7,pady=7);card.grid_propagate(False);inner.grid_columnconfigure(i%cols,weight=1)
            img_label=tk.Label(card,text="Gerando miniatura...",bg=self.palette["ROW_ALT"],fg=self.palette["MUTED"],font=("Segoe UI",8),height=11);img_label.pack(fill="both",expand=True,padx=8,pady=(8,3))
            tk.Label(card,text=self.kind_text(j),bg=self.palette["ROW_ALT"],fg=self.palette["BLUE2"],font=("Segoe UI",7,"bold")).pack(anchor="w",padx=8)
            tk.Label(card,text=j["produto"],bg=self.palette["ROW_ALT"],fg=self.palette["TEXT"],font=("Segoe UI",8,"bold"),wraplength=190,justify="left").pack(anchor="w",padx=8,pady=(2,8))
            img_label.bind("<Button-1>",lambda e,j=j:self._select_from_gallery(j))
            self.load_thumb_async(j,img_label,token)
    def _select_from_gallery(self,j):
        self.view_mode.set("Tabela");self.gallery_holder.pack_forget();self.table_holder.pack(fill="both",expand=True);self.refresh_tree();
        if self.tree.exists(str(j["id"])):self.tree.selection_set(str(j["id"]));self.tree.see(str(j["id"]));self.details()
    def load_thumb_async(self,j,label,token):
        path=self.thumb_path(j)
        def show(p):
            if token!=self._gallery_token or not label.winfo_exists():return
            try:
                img=tk.PhotoImage(file=str(p));factor=max(1,math.ceil(max(img.width()/180,img.height()/205)));img=img.subsample(factor,factor);self._gallery_images.append(img);label.config(image=img,text="",height=0)
            except Exception:label.config(text="Miniatura indisponível")
        if path.exists():self.after(0,lambda:show(path));return
        self._thumb_queue.put((j,label,token,path,show));self._ensure_thumb_worker()
    def _ensure_thumb_worker(self):
        if self._thumb_worker_running:return
        self._thumb_worker_running=True
        def worker():
            while True:
                try:j,label,token,path,show=self._thumb_queue.get_nowait()
                except queue.Empty:break
                try:
                    raw=generate_preview(j,self.master_app.validity.get());shutil.copy2(raw,path);self.after(0,lambda p=path,show=show:show(p))
                except Exception:self.after(0,lambda label=label:label.config(text="Prévia disponível no Windows") if label.winfo_exists() else None)
            self._thumb_worker_running=False
            if not self._thumb_queue.empty():self.after(0,self._ensure_thumb_worker)
        threading.Thread(target=worker,daemon=True).start()
    def real_preview(self):
        j=self.selected_job()
        if not j:messagebox.showinfo("SR Studio","Selecione um cartaz.");return
        self.preview_btn.config(state="disabled",text="GERANDO PRÉVIA...")
        def worker():
            try:
                png=generate_preview(j,self.master_app.validity.get());self.after(0,lambda:self.show_preview_popup(png,j))
            except Exception as e:
                msg=str(e)
                self.after(0,lambda msg=msg:messagebox.showerror("Prévia real",msg))
            finally:self.after(0,lambda:self.preview_btn.config(state="normal",text="PRÉVIA REAL"))
        threading.Thread(target=worker,daemon=True).start()
    def show_preview_popup(self,png,j):
        w=tk.Toplevel(self);w.title("Prévia real - "+j["produto"]);w.configure(bg=self.palette["APP_BG"])
        img=tk.PhotoImage(file=str(png));factor=max(1,math.ceil(max(img.width()/650,img.height()/720)));img=img.subsample(factor,factor);w._img=img
        tk.Label(w,image=img,bg=self.palette["APP_BG"]).pack(padx=12,pady=12);center_toplevel(w,self.master_app,max(420,img.width()+24),max(480,img.height()+24))
    def generate_this(self,do_print=False):
        j=self.selected_job()
        if not j:messagebox.showinfo("Revisão","Selecione um cartaz.");return
        if do_print:out=LOCAL_DATA/"impressao_cartaz_unico.pdf"
        else:
            val=filedialog.asksaveasfilename(title="Salvar cartaz",defaultextension=".pdf",filetypes=[("PDF","*.pdf")],initialfile=safe_name(j["produto"])+".pdf")
            if not val:return
            out=Path(val)
        def worker():
            try:
                result=generate_pdf([j],out,self.master_app.validity.get(),None,threading.Event())
                if not result.get("output_created"):raise RuntimeError("O PowerPoint não gerou o cartaz.")
                if do_print:print_pdf(out);self.after(0,lambda:messagebox.showinfo("Impressão","Cartaz enviado para a impressora padrão do Windows."))
                else:self.after(0,lambda:messagebox.showinfo("Cartaz",f"PDF salvo.\n\n{out}"))
            except Exception as e:
                msg=str(e)
                self.after(0,lambda msg=msg:messagebox.showerror("Cartaz",msg))
        threading.Thread(target=worker,daemon=True).start()
    def validate_powerpoint_layout(self):
        jobs=[j for j in self.jobs if j.get("selected")]
        if not jobs:messagebox.showinfo("SR Studio","Nenhum cartaz selecionado.");return
        self.validate_btn.config(state="disabled",text="VERIFICANDO...")
        self.master_app.busy=True;started=time.time();self.master_app.status_text.set("Verificando layout no PowerPoint...")
        def prog(a,b,t):self.master_app.after(0,lambda:self.master_app.status_text.set(f"{t} • {time.time()-started:.1f}s"))
        def worker():
            try:
                result=validate_layout(jobs,self.master_app.validity.get(),prog)
                def finish():
                    for j in jobs:
                        r=result.get(str(j.get("id")))
                        if r:j["layout_status"]=r.get("status","");j["layout_detail"]=r.get("detail","");j["layout_font"]=r.get("product_font",0)
                    self.save_changes();self.refresh_current();self.details();messagebox.showinfo("Verificação","Verificação de layout concluída.")
                self.after(0,finish)
            except Exception as e:
                msg=str(e)
                self.after(0,lambda msg=msg:messagebox.showerror("Verificação de layout",msg))
            finally:self.after(0,lambda:(self.validate_btn.config(state="normal",text="VERIFICAR LAYOUT"),setattr(self.master_app,"busy",False)))
        threading.Thread(target=worker,daemon=True).start()
    def update_counter(self):
        selected=sum(bool(j.get("selected")) for j in self.jobs);club=sum(j.get("selected") and j.get("tipo")==3 for j in self.jobs);edited=sum(bool(j.get("manual_edit")) for j in self.jobs)
        self.counter_var.set(f"{selected} selecionados • {club} Clube Exclusivo • {edited} editados")
    def save_changes(self):
        self.on_change();self.update_counter()
    def close_review(self):
        self.save_changes();self.destroy()



class StartupSplash:
    """Tela de abertura leve com progresso real das verificações iniciais."""
    STEPS = [
        ("AMBIENTE", "Preparando ambiente..."),
        ("MODELOS", "Carregando modelos PowerPoint..."),
        ("DADOS", "Preparando bancos e histórico..."),
        ("IMPRESSORAS", "Carregando impressoras..."),
        ("CONFIGURAÇÕES", "Carregando preferências..."),
        ("POWERPOINT", "Verificando Microsoft PowerPoint..."),
        ("INTERFACE", "Preparando interface..."),
    ]

    def __init__(self, settings):
        self.settings = settings or {}
        self.mode = self.settings.get("startup_animation", "Sempre")
        self.anim_mode = self.settings.get("animations", "Normal")
        self.reduced = self.anim_mode == "Reduzidas"
        self.duration_mode = self.settings.get("startup_duration", "Normal")
        self.minimum_duration = {"Rápida":2.2,"Normal":5.2,"Estendida":7.2}.get(self.duration_mode,5.2)
        if self.reduced:
            self.minimum_duration = min(self.minimum_duration,2.4)
        self.root = tk.Tk()
        self.root.title(f"SR Studio {APP_DISPLAY_VERSION}")
        self.root.overrideredirect(True)
        self.root.configure(bg="#071F45")
        try: self.root.attributes("-topmost", True)
        except Exception: pass
        try: self.root.attributes("-alpha", 0.0 if self.mode != "Desativada" else 1.0)
        except Exception: pass

        self.width = 600
        self.height = 485
        sw = self.root.winfo_screenwidth()
        sh = self.root.winfo_screenheight()
        x = max(0, (sw-self.width)//2)
        y = max(0, (sh-self.height)//2)
        self.root.geometry(f"{self.width}x{self.height}+{x}+{y}")

        self.started = time.time()
        self.finished = False
        self.visible = self.mode == "Sempre"
        self.logo = None
        self.stage_var = tk.StringVar(value="Inicializando SR Studio...")
        self.percent_var = tk.StringVar(value="0%")
        self.step_labels = []
        self._pulse_step = 0

        self._build()
        if self.mode == "Desativada":
            self.root.withdraw()
        elif self.mode == "Somente se demorar":
            self.root.withdraw()
            self.root.after(650, self._show_if_still_loading)
        else:
            self.root.deiconify()
            self._fade_in()

    def _build(self):
        # Splash 4.0.4: mesma linguagem visual usada na tela de atualização.
        BG = "#0E1828"
        CARD = "#111D2E"
        LINE = "#263750"
        BLUE = "#2F6FED"
        BLUE2 = "#66A0FF"
        TEXT = "#F5F8FC"
        MUTED = "#9DB0C9"
        TRACK = "#1D2B40"

        self.root.configure(bg=BG)
        self.width = 600
        self.height = 430
        try:
            sw = self.root.winfo_screenwidth(); sh = self.root.winfo_screenheight()
            x = max(0, (sw-self.width)//2); y = max(0, (sh-self.height)//2)
            self.root.geometry(f"{self.width}x{self.height}+{x}+{y}")
        except Exception:
            pass

        card = tk.Frame(self.root, bg=CARD, highlightbackground=LINE, highlightthickness=1)
        card.place(relx=.5, rely=.5, anchor="center", width=520, height=350)

        # Logo oficial do SR Studio.
        self.logo = _brand_photo(self.root, 78)
        if self.logo is not None:
            badge = tk.Label(card, image=self.logo, bg=CARD, bd=0, highlightthickness=0)
        else:
            badge = tk.Label(card, text="SR", bg=BLUE, fg="white",
                             font=("Segoe UI", 25, "bold"), width=4, height=1)
        badge.pack(pady=(22,8))

        tk.Label(card, text="SR STUDIO", bg=CARD, fg=TEXT,
                 font=("Segoe UI", 17, "bold")).pack()
        tk.Label(card, text="ENCARTES INTELLIGENCE", bg=CARD, fg=BLUE2,
                 font=("Segoe UI", 8, "bold")).pack(pady=(2,5))
        tk.Label(card, text=f"Versão {APP_DISPLAY_VERSION}", bg=CARD, fg=MUTED,
                 font=("Segoe UI", 8)).pack(pady=(0,14))

        self.stage_label = tk.Label(card, textvariable=self.stage_var,
                                    bg=CARD, fg=TEXT,
                                    font=("Segoe UI", 9, "bold"))
        self.stage_label.pack(pady=(0,7))

        bar_wrap = tk.Frame(card, bg=CARD)
        bar_wrap.pack(fill="x", padx=58)
        style = ttk.Style(self.root)
        try: style.theme_use("clam")
        except Exception: pass
        style.configure(
            "Splash.Horizontal.TProgressbar",
            troughcolor=TRACK, background=BLUE2,
            bordercolor=TRACK, lightcolor=BLUE2, darkcolor=BLUE2,
            borderwidth=0, thickness=8
        )
        self.progress = ttk.Progressbar(bar_wrap, maximum=100, mode="determinate",
                                        style="Splash.Horizontal.TProgressbar")
        self.progress.pack(side="left", fill="x", expand=True)
        tk.Label(bar_wrap, textvariable=self.percent_var, width=5,
                 bg=CARD, fg=MUTED, font=("Segoe UI", 8, "bold")).pack(side="right", padx=(8,0))

        # Pontos animados no mesmo estilo visual da tela de atualização.
        self._startup_dots = tk.StringVar(value="●  ○  ○")
        tk.Label(card, textvariable=self._startup_dots, bg=CARD, fg=BLUE2,
                 font=("Segoe UI Symbol", 11, "bold")).pack(pady=(10,4))
        tk.Label(card, text="Preparando o SR Studio para você...", bg=CARD, fg=MUTED,
                 font=("Segoe UI", 8)).pack()

        self.step_labels = []  # compatibilidade com _set_step
        footer = tk.Frame(card, bg=CARD)
        footer.pack(side="bottom", fill="x", padx=18, pady=(0,12))
        tk.Label(footer, text="Feito por Lucas", bg=CARD, fg="#7387A3",
                 font=("Segoe UI", 7)).pack(side="right")

        self._startup_anim_tick = 0
        self._animate_startup_badge()

    def _animate_startup_badge(self):
        if self.finished:
            return
        frames=("●  ○  ○","○  ●  ○","○  ○  ●","○  ●  ○")
        tick=int(getattr(self,"_startup_anim_tick",0)); self._startup_anim_tick=tick+1
        try:self._startup_dots.set(frames[tick%len(frames)])
        except Exception:pass
        try:self.root.after(320,self._animate_startup_badge)
        except Exception:pass

    def _show_if_still_loading(self):
        if self.finished or self.mode != "Somente se demorar":
            return
        self.visible = True
        self.root.deiconify()
        try: self.root.lift()
        except Exception: pass
        self._fade_in()

    def _fade_in(self, alpha=0.0):
        if not self.visible or self.reduced or self.finished:
            try: self.root.attributes("-alpha", 1.0)
            except Exception: pass
            return
        alpha = min(1.0, alpha + 0.10)
        try: self.root.attributes("-alpha", alpha)
        except Exception: return
        if alpha < 1.0:
            self.root.after(22, lambda:self._fade_in(alpha))

    def _fade_out(self, alpha=1.0):
        if self.reduced or not self.visible:
            self.root.destroy()
            return
        alpha = max(0.0, alpha - 0.12)
        try: self.root.attributes("-alpha", alpha)
        except Exception:
            self.root.destroy()
            return
        if alpha <= 0:
            self.root.destroy()
        else:
            self.root.after(18, lambda:self._fade_out(alpha))

    def _set_step(self, index, status="active"):
        for i,(icon,lbl) in enumerate(self.step_labels):
            if i < index:
                icon.config(text="✓", fg="#77D39A")
                lbl.config(fg="#DCE7F7", font=("Segoe UI",8))
            elif i == index:
                if status == "ok":
                    icon.config(text="✓", fg="#77D39A")
                    lbl.config(fg="#F7FAFF", font=("Segoe UI",8,"bold"))
                elif status == "warning":
                    icon.config(text="!", fg="#FFC96B")
                    lbl.config(fg="#F7FAFF", font=("Segoe UI",8,"bold"))
                else:
                    icon.config(text="●", fg="#5EA1FF")
                    lbl.config(fg="white", font=("Segoe UI",8,"bold"))
            elif i > index:
                icon.config(text="○", fg="#7895BC")
                lbl.config(fg="#AFC4E1", font=("Segoe UI",8))

    def update_step(self, index, text, percent, warning=False):
        if self.finished:
            return
        self.stage_var.set(text)
        self.progress["value"] = percent
        self.percent_var.set(f"{int(percent)}%")
        self._set_step(index, "warning" if warning else "active")
        try: self.root.update_idletasks()
        except Exception: pass

    def mark_step_done(self, index, warning=False):
        if self.finished:
            return
        self._set_step(index, "warning" if warning else "ok")

    def start_checks(self):
        def emit(index,text,percent,warning=False):
            try:
                self.root.after(0, lambda i=index,t=text,p=percent,w=warning:
                                self.update_step(i,t,p,w))
            except Exception:
                pass

        def mark(index,warning=False):
            try:
                self.root.after(0, lambda i=index,w=warning:self.mark_step_done(i,w))
            except Exception:
                pass

        def worker():
            global STARTUP_CACHE
            health={"powerpoint":False,"models":False,"memory":False,"backup":True}

            # 1) Ambiente
            emit(0,"Preparando ambiente...",6)
            try:
                LOCAL_DATA.mkdir(parents=True,exist_ok=True)
                (APP_DIR/"dados").mkdir(parents=True,exist_ok=True)
                (LOCAL_DATA/"thumb_cache").mkdir(parents=True,exist_ok=True)
                health["memory"]=True
                mark(0)
            except Exception:
                mark(0,True)

            # 2) Modelos + correção noTextEdit
            emit(1,"Carregando e preparando modelos...",20)
            model_warning=False
            try:
                ensure_all_models_unlocked()
                required=[MODEL1,MODEL2,MODEL1_LIMIT,MODEL2_LIMIT,CLUB_MODEL,CLUB_MODEL_LIMIT,ATACADO_MODEL,MODEL_SALE]
                model_warning=not all(Path(x).exists() for x in required)
                health["models"]=not model_warning
            except Exception:
                model_warning=True
            mark(1,model_warning)

            # 3) Dados pesados: deixa Banco, Montador, Atacado e Biblioteca prontos em memória.
            emit(2,"Pré-carregando bancos, catálogos, encartes e templates...",38)
            data_warning=False
            try:
                hist=list(self.startup_cache.get("history_entries") or load_json(HISTORY_FILE,[]))
                STARTUP_CACHE["history_entries"]=hist
                STARTUP_CACHE["history_count"]=len(hist)
                STARTUP_CACHE["atacado_reports"]=[dict(x) for x in reports_history()]
                STARTUP_CACHE["product_catalog_count"]=preload_product_catalog(force=True)
                STARTUP_CACHE["catalog_counts"]=catalog_counts()
                STARTUP_CACHE["sale_catalog_count"]=preload_sale_catalog(force=True)
                STARTUP_CACHE["builder_catalog_count"]=preload_builder_catalog(force=True)
                STARTUP_CACHE["encarte_campaign_count"]=preload_encarte_data(force=True)
                STARTUP_CACHE["atacado_catalog_count"]=preload_atacado_catalog(force=True)
                STARTUP_CACHE["library_cache_count"]=preload_library_cache(force=True)
                try:
                    from PromotionLibrary import library_counts as _library_counts
                    STARTUP_CACHE["library_counts"]=_library_counts()
                except Exception:
                    STARTUP_CACHE["library_counts"]={}
                health["memory"]=True
            except Exception:
                data_warning=True
            mark(2,data_warning)

            # 4) Impressoras: esta era a principal causa das travadas ao abrir Manual/Config.
            emit(3,"Carregando impressora padrão e impressoras instaladas...",56)
            printer_warning=False
            try:
                STARTUP_CACHE["default_printer"]=default_printer_name(refresh=True)
                STARTUP_CACHE["printers"]=list_printers(refresh=True)
            except Exception:
                printer_warning=True
            mark(3,printer_warning)

            # 5) Configurações + filas/centros: pré-carrega tudo que não exige abrir PowerPoint.
            emit(4,"Carregando configurações, fila e centros do sistema...",69)
            settings_warning=False
            try:
                STARTUP_CACHE["print_profiles"]=load_print_profiles(refresh=True)
                STARTUP_CACHE["corrections_count"]=len(corrections(refresh=True))
                STARTUP_CACHE["queue"]=(queue_load() or {}).get("tasks",[])
                STARTUP_CACHE["reprints"]=reprint_items()
                STARTUP_CACHE["ciss_last_import"]=ciss_last_import_info() or {}
                STARTUP_CACHE["ui_settings"]=load_json(UI_SETTINGS_FILE,{})
                STARTUP_CACHE["update_history"]=load_json(UPDATE_HISTORY_FILE,[])
                STARTUP_CACHE["model_signatures"]={str(x.name):file_signature(x) for x in [MODEL1,MODEL2,MODEL1_LIMIT,MODEL2_LIMIT,CLUB_MODEL,CLUB_MODEL_LIMIT,ATACADO_MODEL,MODEL_SALE] if Path(x).exists()}
                STARTUP_CACHE["output_root"]=str(default_output_root(load_json(UI_SETTINGS_FILE,{})))
            except Exception:
                settings_warning=True
            mark(4,settings_warning)

            # 6) PowerPoint
            emit(5,"Verificando Microsoft PowerPoint...",82)
            ppt_warning=False
            if os.name=="nt":
                try:
                    cmd=[find_powershell(),"-NoProfile","-Command","if([type]::GetTypeFromProgID('PowerPoint.Application')){'OK'}"]
                    r=subprocess.run(cmd,capture_output=True,text=True,timeout=6,creationflags=getattr(subprocess,"CREATE_NO_WINDOW",0))
                    ppt_warning="OK" not in (r.stdout or "")
                    health["powerpoint"]=not ppt_warning
                except Exception:
                    ppt_warning=True
            else:
                health["powerpoint"]=False
            mark(5,ppt_warning)

            # 7) Interface
            emit(6,"Finalizando dados da interface...",94)
            STARTUP_CACHE["health"]=health
            STARTUP_CACHE["preloaded_at"]=datetime.now().isoformat(timespec="seconds")
            time.sleep(0.08 if self.reduced else 0.18)
            mark(6)
            try:self.root.after(0,self.complete)
            except Exception:pass

        threading.Thread(target=worker,daemon=True).start()

    def complete(self):
        if self.finished:
            return
        self.progress["value"]=100
        self.percent_var.set("100%")
        self.stage_var.set("Tudo pronto  ✓")
        self._set_step(len(self.STEPS)-1,"ok")
        self.finished=True

        # Se estava no modo "somente se demorar" e nunca chegou a aparecer,
        # não acrescenta atraso artificial.
        if not self.visible:
            self.root.destroy()
            return

        # Splash 2.0: mantém a abertura tempo suficiente para uma transição
        # elegante enquanto o pré-carregamento real acontece em segundo plano.
        elapsed=time.time()-self.started
        wait=max(650,int(max(0,self.minimum_duration-elapsed)*1000))
        self.root.after(wait,self._fade_out)

    def run(self):
        self.start_checks()
        self.root.mainloop()


def preload_startup_headless():
    """Pré-carrega os componentes caros mesmo quando a Splash está desativada."""
    global STARTUP_CACHE
    health={"powerpoint":False,"models":False,"memory":False,"backup":True}
    try:
        LOCAL_DATA.mkdir(parents=True,exist_ok=True);(APP_DIR/"dados").mkdir(parents=True,exist_ok=True)
        ensure_all_models_unlocked();health["models"]=all(Path(x).exists() for x in [MODEL1,MODEL2,MODEL1_LIMIT,MODEL2_LIMIT,CLUB_MODEL,CLUB_MODEL_LIMIT,ATACADO_MODEL,MODEL_SALE])
    except Exception:pass
    try:
        hist=load_json(HISTORY_FILE,[]);STARTUP_CACHE["history_entries"]=hist;STARTUP_CACHE["history_count"]=len(hist)
        STARTUP_CACHE["atacado_reports"]=[dict(x) for x in reports_history()]
        STARTUP_CACHE["product_catalog_count"]=preload_product_catalog(force=True);STARTUP_CACHE["catalog_counts"]=catalog_counts()
        STARTUP_CACHE["sale_catalog_count"]=preload_sale_catalog(force=True);STARTUP_CACHE["builder_catalog_count"]=preload_builder_catalog(force=True)
        STARTUP_CACHE["atacado_catalog_count"]=preload_atacado_catalog(force=True);STARTUP_CACHE["library_cache_count"]=preload_library_cache(force=True);health["memory"]=True
        try:
            from PromotionLibrary import library_counts as _library_counts
            STARTUP_CACHE["library_counts"]=_library_counts()
        except Exception:pass
    except Exception:pass
    try:
        STARTUP_CACHE["default_printer"]=default_printer_name(refresh=True);STARTUP_CACHE["printers"]=list_printers(refresh=True)
    except Exception:pass
    try:
        STARTUP_CACHE["print_profiles"]=load_print_profiles(refresh=True);STARTUP_CACHE["corrections_count"]=len(corrections(refresh=True))
        STARTUP_CACHE["queue"]=(queue_load() or {}).get("tasks",[]);STARTUP_CACHE["reprints"]=reprint_items();STARTUP_CACHE["ciss_last_import"]=ciss_last_import_info() or {}
        STARTUP_CACHE["ui_settings"]=load_json(UI_SETTINGS_FILE,{});STARTUP_CACHE["update_history"]=load_json(UPDATE_HISTORY_FILE,[])
        STARTUP_CACHE["model_signatures"]={str(x.name):file_signature(x) for x in [MODEL1,MODEL2,MODEL1_LIMIT,MODEL2_LIMIT,CLUB_MODEL,CLUB_MODEL_LIMIT,ATACADO_MODEL,MODEL_SALE] if Path(x).exists()}
        STARTUP_CACHE["output_root"]=str(default_output_root(load_json(UI_SETTINGS_FILE,{})))
    except Exception:pass
    if os.name=="nt":
        try:
            cmd=[find_powershell(),"-NoProfile","-Command","if([type]::GetTypeFromProgID('PowerPoint.Application')){'OK'}"]
            r=subprocess.run(cmd,capture_output=True,text=True,timeout=6,creationflags=getattr(subprocess,"CREATE_NO_WINDOW",0));health["powerpoint"]="OK" in (r.stdout or "")
        except Exception:pass
    STARTUP_CACHE["health"]=health;STARTUP_CACHE["preloaded_at"]=datetime.now().isoformat(timespec="seconds")


def run_startup_splash():
    """Executa a abertura e usa esse tempo para preparar as telas mais pesadas."""
    settings=load_json(UI_SETTINGS_FILE,{})
    mode=settings.get("startup_animation","Sempre")
    if mode=="Desativada":
        preload_startup_headless()
        return
    splash=StartupSplash(settings)
    splash.run()


class App(APP_TK_BASE):
    def __init__(self):
        super().__init__()
        self.title(f"SR Studio {APP_DISPLAY_VERSION}")
        try:
            if BRAND_ICON.exists(): self.iconbitmap(str(BRAND_ICON))
        except Exception:
            pass
        try:
            self.window_brand_icon=_brand_photo(self,64)
            if self.window_brand_icon is not None: self.iconphoto(True,self.window_brand_icon)
        except Exception:
            self.window_brand_icon=None
        self.ui_settings = load_json(UI_SETTINGS_FILE, {})
        self.theme_mode = tk.StringVar(value=self.ui_settings.get("theme", "Automático"))
        self.palette = choose_palette(self.theme_mode.get())
        self.geometry(self.ui_settings.get("geometry", "1320x820"))
        self.minsize(940, 620)
        self.configure(bg=self.palette["APP_BG"])
        self.animations_mode = tk.StringVar(value=self.ui_settings.get("animations", "Normal"))
        self.startup_animation_mode = tk.StringVar(value=self.ui_settings.get("startup_animation", "Sempre"))
        self.startup_duration_mode = tk.StringVar(value=self.ui_settings.get("startup_duration", "Normal"))
        self.scale_mode = tk.StringVar(value=self.ui_settings.get("scale", "100%"))
        apply_scaling(self, self.scale_mode.get())
        install_centered_messageboxes()
        install_parented_filedialogs()
        self.toast = ToastManager(self)
        self.previous_crash = RUNNING_FLAG.exists()
        try: RUNNING_FLAG.write_text(datetime.now().isoformat(),encoding="utf-8")
        except Exception: pass
        self.busy = False
        self.sidebar_collapsed = bool(self.ui_settings.get("sidebar_collapsed", False))
        self.last_update_text = tk.StringVar(value="Sem importações nesta sessão")
        self.startup_cache = dict(STARTUP_CACHE)

        self.file_path = tk.StringVar()
        self.validity = tk.StringVar(value="VÁLIDO DE")
        self.output_mode = tk.StringVar(value="PDF ÚNICO")
        self.status_text = tk.StringVar(value="Pronto para começar")
        self.analysis = None
        # Os modelos já são preparados na etapa de abertura (Splash).
        self.cancel_event = threading.Event()
        self.logo_img = None
        self.logo_source_img = None
        self.last_failed_jobs = []
        self.last_error_log = None
        self.last_generation_base = None

        self.style = ttk.Style(self)
        try: self.style.theme_use("clam")
        except Exception: pass
        # Barras de carregamento 2.0: finas, consistentes e sem borda pesada.
        for sty,color in [
            ("Horizontal.TProgressbar",self.palette["BLUE2"]),
            ("SR.Horizontal.TProgressbar",self.palette["BLUE2"]),
            ("SR.Loading.Horizontal.TProgressbar",self.palette["BLUE2"]),
            ("SR.Success.Horizontal.TProgressbar",self.palette["GREEN_TXT"]),
            ("SR.Warning.Horizontal.TProgressbar",self.palette["ORANGE_TXT"]),
        ]:
            self.style.configure(sty,troughcolor=self.palette["ROW_ALT"],background=color,
                                 bordercolor=self.palette["ROW_ALT"],lightcolor=color,darkcolor=color,
                                 borderwidth=0,thickness=9)
        self.style.configure("Treeview",background=CARD,fieldbackground=CARD,foreground=TEXT,rowheight=29,font=("Segoe UI",9),borderwidth=0)
        self.style.configure("Treeview.Heading",font=("Segoe UI",8,"bold"),padding=(8,7),background=self.palette["ROW_ALT"],foreground=TEXT,relief="flat")
        self.style.map("Treeview",background=[("selected",SELECT_BG)],foreground=[("selected",TEXT)])
        self.style.configure("TCombobox",padding=5,arrowsize=13)
        self.style.configure("TScrollbar",background=self.palette["ROW_ALT"],troughcolor=self.palette["APP_BG"],borderwidth=0,arrowcolor=self.palette["MUTED"])

        self.build_layout()
        # Sempre inicia maximizado. Repetimos após o idle/120 ms para garantir
        # o estado mesmo em máquinas onde o Windows demora a materializar a janela.
        self.after_idle(self._open_maximized)
        self.after(120, self._open_maximized)
        try: preload_sria_data()
        except Exception: pass
        self.protocol("WM_DELETE_WINDOW", self.on_close_app)
        self.bind("<Configure>", self._responsive_shell)
        self.show_home()
        self.install_shortcuts()
        self.after(300,lambda:cleanup_temp(3))
        self.after(700,self.offer_crash_recovery)
        self.after(1100,self.offer_queue_recovery)

    def _open_maximized(self):
        """Abre o SR Studio maximizado, preservando a barra de tarefas do sistema."""
        try:
            self.state("zoomed")
            return
        except Exception:
            pass
        try:
            self.attributes("-zoomed", True)
            return
        except Exception:
            pass
        # Fallback para ambientes sem suporte ao estado zoomed.
        try:
            self.update_idletasks()
            self.geometry(f"{self.winfo_screenwidth()}x{self.winfo_screenheight()}+0+0")
        except Exception:
            pass

    def build_layout(self):
        pal=self.palette
        self.sidebar = tk.Frame(self, bg=pal["SIDEBAR"], width=72 if self.sidebar_collapsed else 196)
        self.sidebar.pack(side="left", fill="y")
        self.sidebar.pack_propagate(False)

        brand=tk.Frame(self.sidebar,bg=pal["SIDEBAR"],height=112)
        brand.pack(fill="x"); brand.pack_propagate(False)
        # Logo oficial do projeto: mesma imagem usada na abertura.
        self.logo_source_img=None
        self.logo_img=_brand_photo(self, 58)
        if self.logo_img is not None:
            self.logo_label=tk.Label(brand,image=self.logo_img,bg=pal["SIDEBAR"],bd=0,highlightthickness=0)
        else:
            self.logo_label=tk.Label(brand,text="SR",bg=pal["BLUE"],fg="white",
                                     font=("Segoe UI",22,"bold"),width=3,height=1,bd=0,relief="flat")
        self.logo_label.pack(pady=(12,4))
        self.brand_text=tk.Label(brand,text=f"STUDIO {APP_VERSION.rsplit('.',1)[0]}",bg=pal["SIDEBAR"],fg="white",font=("Segoe UI",10,"bold"))
        self.brand_text.pack()

        self.nav_holder=tk.Frame(self.sidebar,bg=pal["SIDEBAR"])
        self.nav_holder.pack(fill="x",pady=(2,0))
        self.nav_buttons={}
        self.nav_defs=[
            ("home","⌂","Início"),("sria","✦","SR IA"),("builder","◆","Montador"),("encartes","▥","Encartes Studio"),("promo","▣","Promoções"),("promo_list","▤","Lista de Promoções"),("products","◫","Banco de Produtos"),
            ("atacado","▦","Atacado"),("manual","＋","Geração Manual"),("queue","≡","Fila"),
            ("reprint","↻","Reimpressão"),("historico","◷","Histórico"),("modelos","▧","Modelos"),
            ("config","⚙","Configurações"),
        ]
        for key,icon,label in self.nav_defs:
            text=icon if self.sidebar_collapsed else f"{icon}  {label}"
            b=tk.Button(self.nav_holder,text=text,anchor="center" if self.sidebar_collapsed else "w",
                        bg=pal["SIDEBAR"],fg="#DCE7F7",activebackground=pal["SIDEBAR_HOVER"],activeforeground="white",
                        relief="flat",bd=0,font=("Segoe UI",9,"bold"),padx=16,pady=8,
                        command=lambda k=key:self.navigate(k))
            b.pack(fill="x",padx=7,pady=1); self.nav_buttons[key]=b
            add_tooltip(b,label)

        foot=tk.Frame(self.sidebar,bg=pal["SIDEBAR"]); foot.pack(side="bottom",fill="x",padx=10,pady=12)
        self.footer_credit=tk.Label(foot,text="" if self.sidebar_collapsed else "Feito por Lucas",bg=pal["SIDEBAR"],fg="#829BC1",font=("Segoe UI",7))
        self.footer_credit.pack(pady=(0,3))
        self.collapse_btn=tk.Button(foot,text="»" if self.sidebar_collapsed else "«",command=self.toggle_sidebar,
                                    bg=pal["SIDEBAR"],fg="#AFC4E1",activebackground=pal["SIDEBAR_HOVER"],
                                    relief="flat",bd=0,font=("Segoe UI",14,"bold"),pady=6)
        self.collapse_btn.pack(fill="x")
        add_tooltip(self.collapse_btn,"Recolher/expandir menu lateral")

        self.main=tk.Frame(self,bg=pal["APP_BG"]); self.main.pack(side="left",fill="both",expand=True)
        self.topbar=tk.Frame(self.main,bg=pal["TOPBAR"],height=58,highlightbackground=pal["LINE"],highlightthickness=1)
        self.topbar.pack(fill="x"); self.topbar.pack_propagate(False)
        self.page_title=tk.Label(self.topbar,text="Início",bg=pal["TOPBAR"],fg=pal["TEXT"],font=("Segoe UI",15,"bold"))
        self.page_title.pack(side="left",padx=(22,14))
        self.global_search_var=tk.StringVar()
        search_wrap=tk.Frame(self.topbar,bg=pal["ROW_ALT"],highlightbackground=pal["LINE"],highlightthickness=1);search_wrap.pack(side="left",fill="x",expand=True,padx=(0,12),pady=10)
        tk.Label(search_wrap,text="⌕",bg=pal["ROW_ALT"],fg=pal["MUTED"],font=("Segoe UI Symbol",12)).pack(side="left",padx=(9,3))
        self.global_search_entry=tk.Entry(search_wrap,textvariable=self.global_search_var,bg=pal["ROW_ALT"],fg=pal["TEXT"],insertbackground=pal["TEXT"],relief="flat",font=("Segoe UI",9))
        self.global_search_entry.pack(side="left",fill="x",expand=True,ipady=4,padx=(0,7))
        self.global_search_entry.insert(0,"")
        self.global_search_entry.bind("<Return>",lambda e:self.open_global_search())
        add_tooltip(self.global_search_entry,"Busca global: produto, código, campanha ou PDF já gerado. Atalho Ctrl+F.")

        self.health_frame=tk.Frame(self.topbar,bg=pal["TOPBAR"]); self.health_frame.pack(side="right",padx=(8,16))
        self.health_labels={}
        for key,label in [("powerpoint","PowerPoint"),("models","Modelos"),("memory","Banco"),("backup","Backup")]:
            w=tk.Label(self.health_frame,text="●",bg=pal["TOPBAR"],fg=pal["MUTED"],font=("Segoe UI",10,"bold"),padx=2)
            w.pack(side="left",padx=2); self.health_labels[key]=w
            add_tooltip(w,label)
        self.version_label=tk.Label(self.topbar,text=f"v{APP_DISPLAY_VERSION}",bg=pal["TOPBAR"],fg=pal["MUTED"],font=("Segoe UI",8,"bold"))
        self.version_label.pack(side="right",padx=(4,7))

        self.content=tk.Frame(self.main,bg=pal["APP_BG"]); self.content.pack(fill="both",expand=True)
        cached_health=self.startup_cache.get("health",{}) if isinstance(getattr(self,"startup_cache",{}),dict) else {}
        if cached_health:
            self.after(80,lambda h=dict(cached_health):self._apply_health(h))
            # Atualiza novamente só depois que a interface já estiver totalmente pronta.
            self.after(10000,self.refresh_health_async)
        else:
            self.after(250,self.refresh_health_async)

    def toggle_sidebar(self):
        self.sidebar_collapsed=not self.sidebar_collapsed
        width=72 if self.sidebar_collapsed else 196
        self.sidebar.config(width=width)
        for key,icon,label in self.nav_defs:
            self.nav_buttons[key].config(text=icon if self.sidebar_collapsed else f"{icon}  {label}",
                                         anchor="center" if self.sidebar_collapsed else "w")
        self.brand_text.config(text=APP_VERSION.rsplit(".",1)[0] if self.sidebar_collapsed else f"STUDIO {APP_VERSION.rsplit('.',1)[0]}")
        try:
            # Mantém o mesmo emblema azul da tela de atualização em ambos os modos.
            self.logo_label.config(text="SR",image="",font=("Segoe UI",18 if self.sidebar_collapsed else 22,"bold"),
                                   width=3,height=1,bg=self.palette["BLUE"],fg="white")
        except Exception:pass
        self.footer_credit.config(text="" if self.sidebar_collapsed else "Feito por Lucas")
        self.collapse_btn.config(text="»" if self.sidebar_collapsed else "«")

    def refresh_health_async(self):
        def worker():
            status={"powerpoint":False,"models":False,"memory":False,"backup":False}
            try:
                missing=[x for x in [MODEL1,MODEL2,MODEL1_LIMIT,MODEL2_LIMIT,CLUB_MODEL,CLUB_MODEL_LIMIT,ATACADO_MODEL,MODEL_SALE] if not Path(x).exists()]
                status["models"]=not missing
            except Exception: pass
            try:
                data_dir=APP_DIR/"dados"
                data_dir.mkdir(exist_ok=True)
                probe=data_dir/".health"; probe.write_text("ok",encoding="utf-8"); probe.unlink()
                status["memory"]=True
                backups=list((data_dir/"backups").glob("*.db")) if (data_dir/"backups").exists() else []
                status["backup"]=bool(backups) or True
            except Exception: pass
            if os.name=="nt":
                try:
                    cmd=[find_powershell(),"-NoProfile","-Command","if([type]::GetTypeFromProgID('PowerPoint.Application')){'OK'}"]
                    r=subprocess.run(cmd,capture_output=True,text=True,timeout=7,creationflags=getattr(subprocess,"CREATE_NO_WINDOW",0))
                    status["powerpoint"]="OK" in (r.stdout or "")
                except Exception: pass
            self.after(0,lambda:self._apply_health(status))
        threading.Thread(target=worker,daemon=True).start()

    def _apply_health(self,status):
        pal=self.palette
        names={"powerpoint":"PowerPoint","models":"Modelos","memory":"Banco de dados","backup":"Backup"}
        for k,v in status.items():
            if k in self.health_labels:
                self.health_labels[k].config(text="●",fg=pal["GREEN_TXT"] if v else pal["ORANGE_TXT"])

    def navigate(self, key):
        if getattr(self,"busy",False):
            messagebox.showinfo("SR Studio","Aguarde a tarefa atual terminar ou cancele antes de trocar de tela.")
            return
        for k,b in self.nav_buttons.items():
            b.config(bg=SIDEBAR_HOVER if k==key else SIDEBAR,
                     fg="white" if k==key else "#DCE7F7")
        if key == "sria": self.show_sria()
        elif key == "builder": self.show_promotion_builder()
        elif key == "encartes": self.show_encartes()
        elif key == "promo": self.show_promotions()
        elif key == "promo_list": self.show_promotion_library()
        elif key == "products": self.show_product_catalog()
        elif key == "atacado": self.show_atacado()
        elif key == "manual": self.show_manual()
        elif key == "queue": self.show_queue()
        elif key == "reprint": self.show_reprint_center()
        elif key == "historico": self.show_history()
        elif key == "modelos": self.show_models()
        elif key == "config": self.show_config()
        elif key == "home": self.show_home()
        else: self.show_placeholder(key)

    def install_shortcuts(self):
        self.bind_all("<Control-f>",lambda e:(self.global_search_entry.focus_set(),self.global_search_entry.select_range(0,"end")))
        self.bind_all("<Control-i>",lambda e:self.shortcut_import())
        self.bind_all("<Control-g>",lambda e:self.shortcut_generate())
        self.bind_all("<Control-p>",lambda e:self.shortcut_print())
        self.bind_all("<Control-Shift-L>",lambda e:self.shortcut_clear())
        self.bind_all("<Escape>",lambda e:self.shortcut_cancel())

        # Rolagem global com a rodinha do mouse.
        # Treeviews/Listboxes recebem um binding de classe, então todas as telas
        # atuais e futuras ganham scroll sem precisar configurar uma por uma.
        try:self.bind_class("Treeview","<MouseWheel>",self._tree_mousewheel)
        except Exception:pass
        try:self.bind_class("Listbox","<MouseWheel>",self._listbox_mousewheel)
        except Exception:pass
        try:self.bind_class("TCombobox","<MouseWheel>",self._combobox_mousewheel)
        except Exception:pass
        try:self.bind_class("Treeview","<Button-4>",lambda e:self._unix_scroll(e,-1))
        except Exception:pass
        try:self.bind_class("Treeview","<Button-5>",lambda e:self._unix_scroll(e,1))
        except Exception:pass
        try:self.bind_class("Listbox","<Button-4>",lambda e:self._unix_scroll(e,-1))
        except Exception:pass
        try:self.bind_class("Listbox","<Button-5>",lambda e:self._unix_scroll(e,1))
        except Exception:pass

        # Para páginas construídas dentro de Canvas com barra vertical, o evento
        # pode nascer em qualquer filho (Label, Frame, Button, Entry...). Este
        # binding procura o Canvas rolável ancestral e move a página inteira.
        self.bind_all("<MouseWheel>",self._page_mousewheel,add="+")
        self.bind_all("<Button-4>",lambda e:self._page_mousewheel(e,-1),add="+")
        self.bind_all("<Button-5>",lambda e:self._page_mousewheel(e,1),add="+")

    @staticmethod
    def _wheel_units(event, fixed_direction=None):
        if fixed_direction is not None:
            return int(fixed_direction) * 3
        try:delta=int(getattr(event,"delta",0) or 0)
        except Exception:delta=0
        if delta==0:return 0
        # Windows normalmente entrega +/-120 por passo; touchpads podem usar
        # valores menores, por isso garantimos ao menos 1 passo.
        amount=max(1,abs(delta)//120)
        direction=-1 if delta>0 else 1
        return direction * amount * 3

    def _tree_mousewheel(self,event):
        # Ctrl+rodinha fica livre para eventuais recursos de zoom do Windows.
        try:
            if int(getattr(event,"state",0) or 0) & 0x0004:return None
            units=self._wheel_units(event)
            if units:
                event.widget.yview_scroll(units,"units")
                return "break"
        except Exception:pass
        return None

    def _listbox_mousewheel(self,event):
        try:
            if int(getattr(event,"state",0) or 0) & 0x0004:return None
            units=self._wheel_units(event)
            if units:
                event.widget.yview_scroll(units,"units")
                return "break"
        except Exception:pass
        return None

    def _combobox_mousewheel(self,event):
        # Evita trocar acidentalmente a opção do Combobox ao tentar rolar uma
        # página (ex.: Configurações). Se estiver em página rolável, rola a tela.
        try:
            canvas=self._scrollable_canvas_for(event.widget)
            if canvas is not None:
                units=self._wheel_units(event)
                if units:canvas.yview_scroll(units,"units")
            return "break"
        except Exception:return "break"

    def _unix_scroll(self,event,direction):
        try:
            event.widget.yview_scroll(self._wheel_units(event,direction),"units")
            return "break"
        except Exception:return None

    def _scrollable_canvas_for(self,widget):
        current=widget
        for _ in range(20):
            if current is None:break
            try:
                if isinstance(current,tk.Canvas):
                    ycmd=str(current.cget("yscrollcommand") or "").strip()
                    if ycmd:return current
            except Exception:pass
            current=getattr(current,"master",None)
        return None

    def _page_mousewheel(self,event,fixed_direction=None):
        try:
            if int(getattr(event,"state",0) or 0) & 0x0004:return None
            # Treeview/Listbox já foram tratados pelo binding de classe.
            if isinstance(event.widget,(ttk.Treeview,tk.Listbox)):return None
            canvas=self._scrollable_canvas_for(event.widget)
            if canvas is None:return None
            units=self._wheel_units(event,fixed_direction)
            if not units:return None
            before=canvas.yview()
            canvas.yview_scroll(units,"units")
            after=canvas.yview()
            if before!=after:return "break"
        except Exception:pass
        return None

    def shortcut_import(self):
        title=self.page_title.cget("text")
        if title=="Promoções" and hasattr(self,"pick_file"):self.pick_file()
        elif title=="Atacado" and getattr(self,"atacado_panel",None):self.atacado_panel.pick()
        else:self.navigate("promo");self.after(100,self.pick_file)
    def shortcut_generate(self):
        title=self.page_title.cget("text")
        if title=="Promoções":self.generate("save")
        elif title=="Atacado" and getattr(self,"atacado_panel",None):self.atacado_panel.generate("save")
    def shortcut_print(self):
        title=self.page_title.cget("text")
        if title=="Promoções":self.generate("print")
        elif title=="Atacado" and getattr(self,"atacado_panel",None):self.atacado_panel.generate("print")
    def shortcut_clear(self):
        title=self.page_title.cget("text")
        if title=="Promoções" and hasattr(self,"clear_promotion_generation"):self.clear_promotion_generation()
        elif title=="Atacado" and getattr(self,"atacado_panel",None):self.atacado_panel.clear_generation()
    def shortcut_cancel(self):
        if not getattr(self,"busy",False):return
        if self.page_title.cget("text")=="Promoções":self.cancel_generation()
        elif self.page_title.cget("text")=="Atacado" and getattr(self,"atacado_panel",None):self.atacado_panel.cancel()

    def print_document(self,path,kind="promo",copies_override=None):
        return print_with_profile(path,kind,copies_override)

    def open_global_search(self):
        query=self.global_search_var.get().strip()
        if not query:
            self.global_search_entry.focus_set();return
        pal=self.palette;w=tk.Toplevel(self);w.title("Busca Global - SR Studio");w.configure(bg=pal["APP_BG"]);center_toplevel(w,self,1080,650)
        top=tk.Frame(w,bg=pal["CARD"],highlightbackground=pal["LINE"],highlightthickness=1);top.pack(fill="x")
        qv=tk.StringVar(value=query);entry=tk.Entry(top,textvariable=qv,bg=pal["ROW_ALT"],fg=pal["TEXT"],insertbackground=pal["TEXT"],relief="flat",font=("Segoe UI",10));entry.pack(side="left",fill="x",expand=True,padx=16,pady=14,ipady=6)
        body=tk.Frame(w,bg=pal["CARD"],highlightbackground=pal["LINE"],highlightthickness=1);body.pack(fill="both",expand=True,padx=16,pady=16)
        cols=("tipo","data","codigo","produto","detalhe");tree=ttk.Treeview(body,columns=cols,show="headings")
        for c,t,ww in [("tipo","Origem",110),("data","Data",130),("codigo","Código",120),("produto","Produto/Arquivo",320),("detalhe","Detalhe",330)]:tree.heading(c,text=t);tree.column(c,width=ww,anchor="w")
        sb=ttk.Scrollbar(body,orient="vertical",command=tree.yview);tree.configure(yscrollcommand=sb.set);tree.pack(side="left",fill="both",expand=True);sb.pack(side="right",fill="y")
        records={}
        def refresh(*_):
            tree.delete(*tree.get_children());records.clear();q=qv.get().strip()
            if not q:return
            idx=0
            for x in search_product_history(q,120):
                idx+=1;iid=f"p{idx}";records[iid]=("product",x)
                price=x.get("clube") or x.get("promocao") or x.get("varejo") or ""
                tree.insert("","end",iid=iid,values=(x.get("origem"),x.get("gerado_em","").replace("T"," ")[:16],x.get("codigo",""),x.get("produto",""),f"{x.get('campanha','')} • R$ {price}" if price else x.get("campanha","")))
            nq=norm(q)
            for x in reprint_items():
                if nq not in norm(x.get("title")) and nq not in norm(Path(x.get("file","")).name):continue
                idx+=1;iid=f"r{idx}";records[iid]=("file",x)
                tree.insert("","end",iid=iid,values=("PDF",x.get("date",""),"",Path(x.get("file","")).name,x.get("title",x.get("kind",""))))
        def open_selected(_=None):
            sel=tree.selection()
            if not sel:return
            kind,x=records.get(sel[0],(None,None))
            if kind=="file" and Path(x.get("file","")).exists():
                try:os.startfile(str(Path(x["file"])))
                except Exception:pass
            elif kind=="product":self.show_product_history_dialog(x.get("codigo",""),x.get("produto",""))
        tk.Button(top,text="BUSCAR",command=refresh,bg=pal["BLUE"],fg="white",relief="flat",font=("Segoe UI",8,"bold"),padx=12,pady=7).pack(side="right",padx=(0,16))
        entry.bind("<Return>",refresh);tree.bind("<Double-1>",open_selected);refresh();entry.focus_set();entry.select_range(0,"end")

    def show_product_history_dialog(self,codigo,produto):
        rows=product_history(codigo,produto,100);pal=self.palette;w=tk.Toplevel(self);w.title(f"Histórico do Produto - {produto}");w.configure(bg=pal["APP_BG"]);center_toplevel(w,self,1080,720)
        tk.Label(w,text=produto,bg=pal["APP_BG"],fg=pal["TEXT"],font=("Segoe UI",16,"bold")).pack(anchor="w",padx=18,pady=(16,2))
        current_ciss=ciss_current_product_snapshot(codigo,produto)
        ciss_bits=[]
        if current_ciss.get("codigo_ciss"):ciss_bits.append(f"CISS {current_ciss.get('codigo_ciss')}")
        if current_ciss.get("custo_reposicao"):ciss_bits.append(f"Custo reposição R$ {current_ciss.get('custo_reposicao')}")
        if current_ciss.get("preco_varejo"):ciss_bits.append(f"Varejo R$ {current_ciss.get('preco_varejo')}")
        ciss_extra=(" • "+" • ".join(ciss_bits)) if ciss_bits else ""
        tk.Label(w,text=f"EAN/Código: {codigo or 'não informado'}{ciss_extra} • {len(rows)} registro(s) promocionais",bg=pal["APP_BG"],fg=pal["MUTED"],font=("Segoe UI",9)).pack(anchor="w",padx=18,pady=(0,8))

        def _num(v):
            try:
                s=str(v or "").replace("R$","").replace(" ","")
                if "," in s and "." in s:s=s.replace(".","").replace(",",".")
                elif "," in s:s=s.replace(",",".")
                return float(s)
            except Exception:return None
        values=[]
        for x in rows:
            v=_num(x.get("clube"))
            if v is None:v=_num(x.get("promocao"))
            if v is not None:values.append(v)
        stats=tk.Frame(w,bg=pal["APP_BG"]);stats.pack(fill="x",padx=18,pady=(0,8))
        if values:
            for title,val in (("Último preço",values[0]),("Menor",min(values)),("Maior",max(values)),("Média",sum(values)/len(values))):
                c=tk.Frame(stats,bg=pal["CARD"],highlightbackground=pal["LINE"],highlightthickness=1);c.pack(side="left",fill="x",expand=True,padx=(0,6))
                tk.Label(c,text=f"R$ {val:.2f}".replace(".",","),bg=pal["CARD"],fg=pal["TEXT"],font=("Segoe UI",12,"bold")).pack(anchor="w",padx=10,pady=(7,0))
                tk.Label(c,text=title,bg=pal["CARD"],fg=pal["MUTED"],font=("Segoe UI",7,"bold")).pack(anchor="w",padx=10,pady=(0,7))

        chart=tk.Canvas(w,height=190,bg=pal["CARD"],highlightbackground=pal["LINE"],highlightthickness=1);chart.pack(fill="x",padx=18,pady=(0,10))
        def draw_chart(_=None):
            chart.delete("all");cw=max(200,chart.winfo_width());ch=max(150,chart.winfo_height());pad=28
            series=[]
            for x in reversed(rows[-30:]):
                v=_num(x.get("clube"));label="Clube"
                if v is None:v=_num(x.get("promocao"));label="Promo"
                if v is not None:series.append((v,str(x.get("gerado_em","") or "")[:10],label))
            if len(series)<2:
                chart.create_text(cw/2,ch/2,text="Histórico insuficiente para desenhar o gráfico.",fill=pal["MUTED"],font=("Segoe UI",9));return
            vals=[x[0] for x in series];lo=min(vals);hi=max(vals);span=max(.01,hi-lo)
            chart.create_line(pad,ch-pad,cw-pad,ch-pad,fill=pal["LINE"]);chart.create_line(pad,pad,pad,ch-pad,fill=pal["LINE"])
            pts=[]
            for i,(v,dt,label) in enumerate(series):
                x=pad+(cw-2*pad)*(i/(len(series)-1));y=ch-pad-(ch-2*pad)*((v-lo)/span);pts.extend((x,y))
                chart.create_oval(x-3,y-3,x+3,y+3,fill=pal["BLUE"],outline="")
            chart.create_line(*pts,fill=pal["BLUE"],width=2)
            chart.create_text(pad+2,pad-8,text=f"R$ {hi:.2f}".replace(".",","),fill=pal["MUTED"],anchor="w",font=("Segoe UI",7))
            chart.create_text(pad+2,ch-pad+10,text=f"R$ {lo:.2f}".replace(".",","),fill=pal["MUTED"],anchor="w",font=("Segoe UI",7))
            chart.create_text(cw-pad,ch-pad+10,text=series[-1][1],fill=pal["MUTED"],anchor="e",font=("Segoe UI",7))
        chart.bind("<Configure>",draw_chart);w.after(80,draw_chart)

        cols=("data","origem","camp","promo","clube","custo","venda");tree=ttk.Treeview(w,columns=cols,show="headings")
        for c,tit,ww in [("data","Data",135),("origem","Origem",100),("camp","Campanha",250),("promo","Promo",75),("clube","Clube",75),("custo","Custo",75),("venda","Venda",75)]:tree.heading(c,text=tit);tree.column(c,width=ww,anchor="w")
        for x in rows:tree.insert("","end",values=(x.get("gerado_em","").replace("T"," ")[:16],x.get("origem",""),x.get("campanha",""),x.get("promocao",""),x.get("clube",""),x.get("custo",""),x.get("varejo","")))
        tree.pack(fill="both",expand=True,padx=18,pady=(0,18))

    def show_reprint_center(self):
        self.clear_content();self.page_title.config(text="Reimpressão");pal=self.palette
        for k,b in self.nav_buttons.items():b.config(bg=pal["SIDEBAR_HOVER"] if k=="reprint" else pal["SIDEBAR"],fg="white" if k=="reprint" else "#DCE7F7")
        frame=tk.Frame(self.content,bg=pal["APP_BG"]);frame.pack(fill="both",expand=True,padx=26,pady=20)
        tk.Label(frame,text="Reimpressão",bg=pal["APP_BG"],fg=pal["TEXT"],font=("Segoe UI",20,"bold")).pack(anchor="w",pady=(0,10))
        card=tk.Frame(frame,bg=pal["CARD"],highlightbackground=pal["LINE"],highlightthickness=1);card.pack(fill="both",expand=True)
        cols=("data","tipo","arquivo","cartazes");tree=ttk.Treeview(card,columns=cols,show="headings",selectmode="browse")
        for c,t,ww in [("data","Data",140),("tipo","Tipo",120),("arquivo","Arquivo",520),("cartazes","Cartazes",90)]:tree.heading(c,text=t);tree.column(c,width=ww,anchor="w")
        records={}
        for i,x in enumerate(reprint_items()):
            iid=str(i);records[iid]=x;tree.insert("","end",iid=iid,values=(x.get("date",""),x.get("kind",""),Path(x.get("file","")).name,x.get("count",0)))
        tree.pack(fill="both",expand=True,padx=12,pady=12)
        bar=tk.Frame(card,bg=pal["CARD"]);bar.pack(fill="x",padx=12,pady=(0,12))
        def selected():
            sel=tree.selection();return records.get(sel[0]) if sel else None
        def open_file():
            x=selected();p=Path(x.get("file","")) if x else None
            if p and p.exists():os.startfile(str(p))
            else:messagebox.showwarning("Reimpressão","O arquivo original não foi encontrado.")
        def open_folder():
            x=selected();p=Path(x.get("file","")) if x else None
            if p and p.exists():os.startfile(str(p.parent))
        def print_again():
            x=selected();p=Path(x.get("file","")) if x else None
            if not p or not p.exists():messagebox.showwarning("Reimpressão","O arquivo não foi encontrado.");return
            kind="atacado" if str(x.get("kind","")).lower().startswith("atacado") else "manual" if str(x.get("kind","")).lower().startswith("manual") else "promo"
            self.print_document(p,kind);self.toast.show("PDF enviado para impressão.","ok")
        for text,cmd,bg,fg in [("ABRIR PDF",open_file,pal["BLUE"],"white"),("ABRIR PASTA",open_folder,pal["LIGHT_BLUE"],pal["LIGHT_BLUE_TXT"]),("IMPRIMIR NOVAMENTE",print_again,pal["GREEN"],pal["GREEN_TXT"])]:tk.Button(bar,text=text,command=cmd,bg=bg,fg=fg,relief="flat",font=("Segoe UI",8,"bold"),padx=12,pady=7).pack(side="left",padx=(0,6))

    def enqueue_promo_jobs(self,jobs,validity_label=None,mode=None):
        jobs=[dict(j) for j in jobs if j.get("selected",True)]
        if not jobs:return None
        mode=mode or self.output_mode.get();first=jobs[0];folder=dated_output_dir("Promoções",self.ui_settings)
        if mode=="PDF ÚNICO":target=unique_path(folder/smart_pdf_name("Promoções",first.get("campanha","PROMOCOES"),first.get("validade","")))
        else:target=folder/("LOTE_"+datetime.now().strftime("%Y%m%d_%H%M%S"));target.mkdir(parents=True,exist_ok=True)
        task=queue_add({"kind":"promo","title":f"Promoções • {len(jobs)} cartazes","jobs":json.loads(json.dumps(jobs,ensure_ascii=False,default=str)),"target":str(target),"mode":mode,"validity":validity_label or self.validity.get(),"action":"save"})
        self.toast.show(f"{len(jobs)} cartazes adicionados à fila.","ok");return task

    def enqueue_atacado(self,posters,report_id=None,date_text=""):
        ps=[dict(p) for p in posters]
        if not ps:return None
        folder=dated_output_dir("Atacado",self.ui_settings);target=unique_path(folder/smart_pdf_name("Atacado","ATACADO",date_text))
        task=queue_add({"kind":"atacado","title":f"Atacado • {len(ps)} cartazes","posters":json.loads(json.dumps(ps,ensure_ascii=False,default=str)),"target":str(target),"report_id":report_id,"action":"save"})
        self.toast.show(f"{len(ps)} cartazes do Atacado adicionados à fila.","ok");return task

    def offer_queue_recovery(self):
        pending=queue_pending()
        if not pending:return
        # PROCESSANDO de uma execução encerrada vira INTERROMPIDA.
        for t in pending:
            if t.get("status")=="PROCESSANDO":queue_update(t["id"],status="INTERROMPIDA")
        if messagebox.askyesno("Fila recuperável",f"Há {len(pending)} geração(ões) pendente(s) na fila.\n\nDeseja abrir a Fila para continuar?"):
            self.navigate("queue")

    def show_queue(self):
        self.clear_content();self.page_title.config(text="Fila");pal=self.palette
        for k,b in self.nav_buttons.items():b.config(bg=pal["SIDEBAR_HOVER"] if k=="queue" else pal["SIDEBAR"],fg="white" if k=="queue" else "#DCE7F7")
        frame=tk.Frame(self.content,bg=pal["APP_BG"]);frame.pack(fill="both",expand=True,padx=26,pady=20)
        tk.Label(frame,text="Fila",bg=pal["APP_BG"],fg=pal["TEXT"],font=("Segoe UI",20,"bold")).pack(anchor="w",pady=(0,10))
        card=tk.Frame(frame,bg=pal["CARD"],highlightbackground=pal["LINE"],highlightthickness=1);card.pack(fill="both",expand=True)
        self.queue_tree=ttk.Treeview(card,columns=("status","tipo","titulo","checkpoint","saida"),show="headings",selectmode="browse")
        for c,t,ww in [("status","Status",110),("tipo","Tipo",100),("titulo","Lote",260),("checkpoint","Progresso",100),("saida","Saída",420)]:self.queue_tree.heading(c,text=t);self.queue_tree.column(c,width=ww,anchor="w")
        self.queue_tree.pack(fill="both",expand=True,padx=12,pady=12)
        self.queue_status=tk.StringVar(value="Pronto.");tk.Label(card,textvariable=self.queue_status,bg=pal["CARD"],fg=pal["MUTED"],font=("Segoe UI",8)).pack(anchor="w",padx=12)
        self.queue_progress=ttk.Progressbar(card,maximum=100,style="SR.Horizontal.TProgressbar");self.queue_progress.pack(fill="x",padx=12,pady=(5,8))
        bar=tk.Frame(card,bg=pal["CARD"]);bar.pack(fill="x",padx=12,pady=(0,12))
        tk.Button(bar,text="PROCESSAR FILA",command=self.process_queue,bg=pal["BLUE"],fg="white",relief="flat",font=("Segoe UI",8,"bold"),padx=12,pady=7).pack(side="left")
        tk.Button(bar,text="REMOVER",command=self.remove_queue_selected,bg=pal["RED"],fg=pal["RED_TXT"],relief="flat",font=("Segoe UI",8,"bold"),padx=12,pady=7).pack(side="left",padx=6)
        tk.Button(bar,text="LIMPAR CONCLUÍDOS",command=lambda:(queue_clear_done(),self.refresh_queue_tree()),bg=pal["LIGHT_BLUE"],fg=pal["LIGHT_BLUE_TXT"],relief="flat",font=("Segoe UI",8,"bold"),padx=12,pady=7).pack(side="left")
        self.refresh_queue_tree()

    def refresh_queue_tree(self):
        if not hasattr(self,"queue_tree"):return
        self.queue_tree.delete(*self.queue_tree.get_children())
        for t in queue_load().get("tasks",[]):
            total=len(t.get("jobs") or t.get("posters") or []);cp=int(t.get("checkpoint") or 0)
            self.queue_tree.insert("","end",iid=t.get("id"),values=(t.get("status"),t.get("kind"),t.get("title"),f"{min(cp,total)}/{total}",t.get("target","")))
    def remove_queue_selected(self):
        sel=self.queue_tree.selection() if hasattr(self,"queue_tree") else ()
        if not sel:return
        queue_remove(sel[0]);self.refresh_queue_tree()

    def process_queue(self):
        if self.busy:return
        tasks=[t for t in queue_load().get("tasks",[]) if t.get("status") in {"PENDENTE","INTERROMPIDA","PROCESSANDO"}]
        if not tasks:messagebox.showinfo("Fila","Não há gerações pendentes.");return
        self.busy=True
        def ui_progress(task,cur,total,text):
            def apply():
                if hasattr(self,"queue_status"):self.queue_status.set(f"{task.get('title')} • {text}")
                if hasattr(self,"queue_progress"):self.queue_progress["value"]=(cur/total*100 if total else 0)
                self.refresh_queue_tree()
            self.after(0,apply)
        def worker():
            try:
                for task in tasks:
                    tid=task["id"];kind=task.get("kind");queue_update(tid,status="PROCESSANDO",error="")
                    work=LOCAL_DATA/"queue_work"/str(tid);work.mkdir(parents=True,exist_ok=True)
                    checkpoint=max(0,int(task.get("checkpoint") or 0));interrupted=False
                    if kind=="promo":
                        jobs=smart_queue_jobs(task.get("jobs",[]));total=len(jobs);pages=[]
                        # Se um checkpoint aponta para página que não existe, volta ao primeiro arquivo ausente.
                        for idx,j in enumerate(jobs):
                            page=work/f"{idx:05d}.pdf"
                            if idx<checkpoint and page.exists():pages.append(page);continue
                            if idx<checkpoint and not page.exists():checkpoint=idx;queue_update(tid,checkpoint=idx)
                            ui_progress(task,idx,total,f"Gerando cartaz {idx+1} de {total}")
                            r=generate_pdf([j],page,task.get("validity","VÁLIDO DE"),lambda a,b,t,idx=idx:ui_progress(task,idx+a,total,t),threading.Event())
                            if r.get("failed_jobs") or not r.get("output_created"):
                                msg=(r.get("failed_jobs") or [{}])[0].get("generation_error","Falha ao gerar cartaz")
                                queue_update(tid,status="INTERROMPIDA",checkpoint=idx,error=msg);interrupted=True;break
                            pages.append(page);checkpoint=idx+1;queue_update(tid,checkpoint=checkpoint)
                        if interrupted:continue
                        mode=task.get("mode","PDF ÚNICO");target=Path(task.get("target"));outs=[]
                        if mode=="PDF ÚNICO":
                            target.parent.mkdir(parents=True,exist_ok=True);merge_pdfs(pages,target);outs=[target]
                        else:
                            target.mkdir(parents=True,exist_ok=True);groups={}
                            for j,page in zip(jobs,pages):groups.setdefault(j.get("campanha","PROMOCAO"),[]).append(page)
                            for camp,pp in groups.items():
                                out=unique_path(target/(safe_name(camp.replace("!!",""))+".pdf"));merge_pdfs(pp,out);outs.append(out)
                        queue_update(tid,status="CONCLUÍDA",checkpoint=total,completed_at=datetime.now().isoformat(timespec="seconds"))
                        record_product_jobs(jobs,"Promoções",";".join(map(str,outs)));record_reprint("Promoções",outs,total,task.get("title",""))
                    elif kind=="atacado":
                        ps=list(task.get("posters",[]));total=len(ps);pages=[];success_keys=[]
                        for idx,poster in enumerate(ps):
                            page=work/f"{idx:05d}.pdf"
                            if idx<checkpoint and page.exists():pages.append(page);success_keys.append(poster.get("cartaz_chave"));continue
                            if idx<checkpoint and not page.exists():checkpoint=idx;queue_update(tid,checkpoint=idx)
                            ui_progress(task,idx,total,f"Gerando cartaz {idx+1} de {total}")
                            r=atacado_run_engine([poster],page,lambda a,b,t,idx=idx:ui_progress(task,idx+a,total,t),threading.Event())
                            if r.get("failed") or not page.exists():
                                msg=(r.get("failed") or [{}])[0].get("message","Falha ao gerar cartaz")
                                queue_update(tid,status="INTERROMPIDA",checkpoint=idx,error=msg);interrupted=True;break
                            pages.append(page);success_keys += r.get("success_keys") or [poster.get("cartaz_chave")];checkpoint=idx+1;queue_update(tid,checkpoint=checkpoint)
                        if interrupted:continue
                        target=Path(task.get("target"));target.parent.mkdir(parents=True,exist_ok=True);merge_pdfs(pages,target)
                        mark_posters_generated(task.get("report_id"),success_keys);queue_update(tid,status="CONCLUÍDA",checkpoint=total,completed_at=datetime.now().isoformat(timespec="seconds"))
                        record_product_jobs(ps,"Atacado",str(target));record_reprint("Atacado",[target],total,task.get("title",""))
                self.after(0,self._queue_finished)
            except Exception as e:
                msg=str(e);self.after(0,lambda msg=msg:self._queue_failed(msg))
        threading.Thread(target=worker,daemon=True).start()
    def _queue_finished(self):
        self.busy=False
        if hasattr(self,"queue_status"):self.queue_status.set("✓ Fila processada.");self.queue_progress["value"]=100;self.refresh_queue_tree()
        messagebox.showinfo("Fila","Processamento da fila concluído.")
    def _queue_failed(self,msg):
        self.busy=False
        if hasattr(self,"queue_status"):self.queue_status.set(msg);self.refresh_queue_tree()
        messagebox.showerror("Fila",msg)

    def _responsive_shell(self, event=None):
        if event is not None and event.widget is not self:
            return
        try:
            w=self.winfo_width()
            if w < 1020 and not self.sidebar_collapsed:
                self.toggle_sidebar()
        except Exception:
            pass

    def on_close_app(self):
        if getattr(self,"busy",False):
            messagebox.showinfo("SR Studio","Há uma tarefa em andamento. Cancele e aguarde finalizar antes de fechar o programa.")
            return
        promo_pending=bool(self.analysis and any(j.get("selected",True) for j in self.analysis.get("jobs",[])))
        atacado_panel=getattr(self,"atacado_panel",None);atacado_pending=bool(atacado_panel and getattr(atacado_panel,"analysis",None) and atacado_panel.posters_to_generate())
        qpending=queue_pending()
        if promo_pending or atacado_pending or qpending:
            items=[]
            if promo_pending:items.append("Promoções carregadas")
            if atacado_pending:items.append("Atacado ainda não concluído")
            if qpending:items.append(f"{len(qpending)} lote(s) pendente(s) na fila")
            choice=SRDialog(self,"Geração pendente","Existe trabalho realmente não concluído:\n\n• "+"\n• ".join(items)+"\n\nO que deseja fazer?","question",("CANCELAR","DESCARTAR","SALVAR PARA CONTINUAR")).result
            if choice=="CANCELAR":return
            if choice=="DESCARTAR":
                try:self.clear_promotion_generation(automatic=True)
                except Exception:pass
                try:
                    if atacado_pending:atacado_panel.clear_generation(automatic=True)
                except Exception:pass
                for t in qpending:
                    try:queue_remove(t.get("id"))
                    except Exception:pass
            else:
                if promo_pending:self.save_session()
                if atacado_pending:
                    try:self.enqueue_atacado(atacado_panel.posters_to_generate(),atacado_panel.analysis.get("report_id"),atacado_panel.analysis.get("parsed",{}).get("data_relatorio",""))
                    except Exception:pass
        try:
            state={
                "geometry": self.geometry(),
                "animations": self.animations_mode.get(),
                "startup_animation": self.startup_animation_mode.get(),
                "theme": self.theme_mode.get(),
                "scale": self.scale_mode.get(),
                "sidebar_collapsed": self.sidebar_collapsed,
                "output_folder": str(self.ui_settings.get("output_folder", "")),
            }
            save_json(UI_SETTINGS_FILE,state)
        except Exception:
            pass
        try:
            if RUNNING_FLAG.exists(): RUNNING_FLAG.unlink()
        except Exception: pass
        self.destroy()

    def offer_crash_recovery(self):
        if not self.previous_crash:
            return
        sess=load_json(SESSION_FILE,{})
        p=sess.get("file")
        if p and Path(p).exists():
            if messagebox.askyesno("Recuperação de sessão","O SR Studio detectou que a execução anterior não foi encerrada normalmente.\n\nDeseja recuperar a última revisão de Promoções?"):
                self.navigate("promo")
                self.import_file(p,restore_session=True)
        else:
            self.toast.show("Execução anterior foi interrompida, mas não havia revisão para recuperar.","warning")

    def open_club_review(self):
        if not self.analysis or not any(j.get("tipo")==3 for j in self.analysis.get("jobs",[])):
            self.navigate("promo")
            self.toast.show("Importe uma planilha com ofertas do Clube SR para revisar o Clube Exclusivo.","warning")
            return
        self.review_filter_override="CLUBE EXCLUSIVO"
        self.navigate("promo")
        self.after(120,self.open_review)

    def _sync_module_palettes(self):
        """Faz os módulos auxiliares seguirem Claro/Escuro/Automático da janela principal."""
        pal=self.palette
        mapping={
            "APP_BG":"APP_BG","CARD":"CARD","TEXT":"TEXT","MUTED":"MUTED","LINE":"LINE",
            "BLUE":"BLUE","BLUE2":"BLUE2","GREEN":"GREEN","GREEN_TXT":"GREEN_TXT",
            "ORANGE":"ORANGE","ORANGE_TXT":"ORANGE_TXT","LIGHT_BLUE":"LIGHT_BLUE",
            "LIGHT_BLUE_TXT":"LIGHT_BLUE_TXT","RED":"RED","RED_TXT":"RED_TXT",
            "YELLOW":"YELLOW","YELLOW_TXT":"YELLOW_TXT","PURPLE":"PURPLE",
            "PURPLE_TXT":"PURPLE_TXT","ROW_ALT":"ROW_ALT","SELECT_BG":"SELECT"
        }
        for mod in (_atacado_mod,_manual_mod):
            try: mod.PAL=pal; mod._PAL=pal
            except Exception: pass
            for name,key in mapping.items():
                try:setattr(mod,name,pal[key])
                except Exception:pass

    def show_promotion_builder(self):
        self.clear_content()
        self.page_title.config(text="Montador de Promoções")
        pal=self.palette
        for k,b in self.nav_buttons.items():
            b.config(bg=pal["SIDEBAR_HOVER"] if k=="builder" else pal["SIDEBAR"],
                     fg="white" if k=="builder" else "#DCE7F7")
        self.promotion_builder_panel=PromotionBuilderPanel(self.content,self,pal)
        self.promotion_builder_panel.pack(fill="both",expand=True)

    def load_builder_jobs(self,jobs,campaign,validity,campaign_id=None):
        """Envia uma campanha criada no Montador para o fluxo oficial de revisão/geração."""
        if not jobs:return
        one=sum(1 for j in jobs if j.get("tipo")==1)
        two=sum(1 for j in jobs if j.get("tipo")==2)
        club=sum(1 for j in jobs if j.get("tipo")==3)
        commercial=validate_promo_jobs(jobs)
        warnings=[x.get("message","") for x in commercial if x.get("severity")!="error"]
        errors=[x.get("message","") for x in commercial if x.get("severity")=="error"]
        analysis={
            "jobs":jobs,
            "campaigns":[{"name":campaign,"total":len(jobs),"one":one,"two":two,"club":club,"validity":validity}],
            "errors":errors,"warnings":warnings,"skips":[],"total":len(jobs),"one":one,"two":two,"club":club,
            "sample":jobs[0] if jobs else None,"commercial_issues":commercial,
        }
        self.navigate("promo")
        self._builder_source=True
        self._builder_campaign_id=campaign_id
        self.file_path.set(f"MONTADOR - {campaign}")
        self.file_label.config(text=f"Montador • {campaign}",fg=self.palette["TEXT"])
        self.finish_import(analysis)
        self.status_text.set("Campanha do Montador carregada. Revise os cartazes antes de gerar.")
        self.after(120,self.open_review)

    def show_sria(self):
        self._sync_module_palettes()
        self.clear_content()
        self.page_title.config(text="SR IA")
        for k,b in self.nav_buttons.items():
            b.config(bg=self.palette["SIDEBAR_HOVER"] if k=="sria" else self.palette["SIDEBAR"],
                     fg="white" if k=="sria" else "#DCE7F7")
        self.sria_panel=SRIAPanel(self.content,self,self.palette)
        self.sria_panel.pack(fill="both",expand=True)

    def show_encartes(self):
        self.clear_content()
        self.page_title.config(text="Encartes Studio")
        pal=self.palette
        for k,b in self.nav_buttons.items():
            b.config(bg=pal["SIDEBAR_HOVER"] if k=="encartes" else pal["SIDEBAR"],
                     fg="white" if k=="encartes" else "#DCE7F7")
        self.encarte_panel=EncartePanel(self.content,self,pal)
        self.encarte_panel.pack(fill="both",expand=True)

    def show_promotion_library(self):
        self.clear_content()
        self.page_title.config(text="Lista de Promoções")
        pal=self.palette
        for k,b in self.nav_buttons.items():
            b.config(bg=pal["SIDEBAR_HOVER"] if k=="promo_list" else pal["SIDEBAR"],
                     fg="white" if k=="promo_list" else "#DCE7F7")
        self.promotion_library_panel=PromotionLibraryPanel(self.content,self,analyze_workbook,pal)
        self.promotion_library_panel.pack(fill="both",expand=True)

    def show_product_catalog(self):
        self.clear_content()
        self.page_title.config(text="Banco de Produtos")
        pal=self.palette
        for k,b in self.nav_buttons.items():
            b.config(bg=pal["SIDEBAR_HOVER"] if k=="products" else pal["SIDEBAR"],
                     fg="white" if k=="products" else "#DCE7F7")
        self.product_organizer_panel=ProductOrganizerPanel(self.content,self,pal)
        self.product_organizer_panel.pack(fill="both",expand=True)

    def show_manual(self):
        self._sync_module_palettes()
        self.clear_content()
        self.page_title.config(text="Geração Manual")
        for k,b in self.nav_buttons.items():
            b.config(bg=self.palette["SIDEBAR_HOVER"] if k=="manual" else self.palette["SIDEBAR"],
                     fg="white" if k=="manual" else "#DCE7F7")
        panel=ManualPanel(self.content,self,generate_pdf,generate_preview)
        self.manual_panel=panel
        panel.pack(fill="both",expand=True)

    def show_atacado(self):
        self._sync_module_palettes()
        self.clear_content()
        self.page_title.config(text="Atacado")
        for k,b in self.nav_buttons.items():
            b.config(bg=SIDEBAR_HOVER if k=="atacado" else SIDEBAR,
                     fg="white" if k=="atacado" else "#DCE7F7")
        panel=AtacadoPanel(
            self.content,
            self,
            reduced_animations=self.animations_mode.get()=="Reduzidas"
        )
        self.atacado_panel=panel
        panel.pack(fill="both",expand=True)
        if enable_drop(panel,panel.import_report,{".pdf"}):
            try:panel.file_label.config(text="Arraste o PDF aqui ou clique em IMPORTAR PDF")
            except Exception:pass

    def show_config(self):
        self.clear_content();self.page_title.config(text="Configurações")
        for k,b in self.nav_buttons.items():b.config(bg=self.palette["SIDEBAR_HOVER"] if k=="config" else self.palette["SIDEBAR"],fg="white" if k=="config" else "#DCE7F7")
        pal=self.palette
        canvas=tk.Canvas(self.content,bg=pal["APP_BG"],highlightthickness=0);sb=ttk.Scrollbar(self.content,orient="vertical",command=canvas.yview);canvas.configure(yscrollcommand=sb.set);canvas.pack(side="left",fill="both",expand=True);sb.pack(side="right",fill="y")
        frame=tk.Frame(canvas,bg=pal["APP_BG"]);win=canvas.create_window((0,0),window=frame,anchor="nw");frame.bind("<Configure>",lambda e:canvas.configure(scrollregion=canvas.bbox("all")));canvas.bind("<Configure>",lambda e:canvas.itemconfigure(win,width=e.width));frame.configure(padx=28,pady=22)
        tk.Label(frame,text="Configurações",bg=pal["APP_BG"],fg=pal["TEXT"],font=("Segoe UI",20,"bold")).pack(anchor="w",pady=(0,10))

        appearance=tk.Frame(frame,bg=pal["CARD"],highlightbackground=pal["LINE"],highlightthickness=1);appearance.pack(fill="x")
        tk.Label(appearance,text="Aparência",bg=pal["CARD"],fg=pal["TEXT"],font=("Segoe UI",12,"bold")).pack(anchor="w",padx=18,pady=(15,3))
        ainfo=tk.Label(appearance,text="ⓘ",bg=pal["CARD"],fg=pal["BLUE2"],font=("Segoe UI",9,"bold"));ainfo.place(relx=1.0,x=-18,y=14,anchor="ne");add_tooltip(ainfo,"Tema, escala e animações da interface. O modo Automático acompanha o Windows.")
        row=tk.Frame(appearance,bg=pal["CARD"]);row.pack(fill="x",padx=18,pady=(12,15))
        tk.Label(row,text="Tema",bg=pal["CARD"],fg=pal["TEXT"],font=("Segoe UI",8,"bold")).grid(row=0,column=0,sticky="w")
        theme=ttk.Combobox(row,textvariable=self.theme_mode,state="readonly",values=["Automático","Claro","Escuro"],width=18);theme.grid(row=1,column=0,sticky="w",pady=(4,0),padx=(0,6))
        tk.Button(row,text="APLICAR TEMA",command=self.apply_theme_setting,bg=pal["LIGHT_BLUE"],fg=pal["LIGHT_BLUE_TXT"],relief="flat",font=("Segoe UI",7,"bold"),padx=8,pady=5).grid(row=2,column=0,sticky="w",pady=(6,0),padx=(0,20))
        tk.Label(row,text="Escala da interface",bg=pal["CARD"],fg=pal["TEXT"],font=("Segoe UI",8,"bold")).grid(row=0,column=1,sticky="w")
        scale=ttk.Combobox(row,textvariable=self.scale_mode,state="readonly",values=["90%","100%","110%","125%"],width=14);scale.grid(row=1,column=1,sticky="w",pady=(4,0),padx=(0,20))
        tk.Label(row,text="Animações",bg=pal["CARD"],fg=pal["TEXT"],font=("Segoe UI",8,"bold")).grid(row=0,column=2,sticky="w")
        anim=ttk.Combobox(row,textvariable=self.animations_mode,state="readonly",values=["Normal","Reduzidas"],width=14);anim.grid(row=1,column=2,sticky="w",pady=(4,0),padx=(0,20))
        tk.Label(row,text="Animação de abertura",bg=pal["CARD"],fg=pal["TEXT"],font=("Segoe UI",8,"bold")).grid(row=0,column=3,sticky="w")
        startup_anim=ttk.Combobox(row,textvariable=self.startup_animation_mode,state="readonly",
                                 values=["Sempre","Somente se demorar","Desativada"],width=19)
        startup_anim.grid(row=1,column=3,sticky="w",pady=(4,0))
        tk.Label(row,text="Duração da abertura",bg=pal["CARD"],fg=pal["TEXT"],font=("Segoe UI",8,"bold")).grid(row=2,column=2,sticky="w",pady=(10,0))
        startup_duration=ttk.Combobox(row,textvariable=self.startup_duration_mode,state="readonly",values=["Rápida","Normal","Estendida"],width=14)
        startup_duration.grid(row=3,column=2,sticky="w",pady=(4,0),padx=(0,20))
        add_tooltip(scale,"90% mostra mais informações em telas menores; 125% aumenta a leitura em monitores de alta resolução.")
        add_tooltip(startup_anim,"Sempre mostra a abertura. 'Somente se demorar' só aparece quando a inicialização passar de aproximadamente meio segundo.")
        add_tooltip(startup_duration,"Rápida ≈ 2 s • Normal ≈ 5 s • Estendida ≈ 7 s. O tempo é usado para pré-carregar o sistema.")
        def save_ui(*_):
            apply_scaling(self,self.scale_mode.get())
            data=load_json(UI_SETTINGS_FILE,{})
            data.update({
                "animations":self.animations_mode.get(),
                "startup_animation":self.startup_animation_mode.get(),
                "startup_duration":self.startup_duration_mode.get(),
                "theme":self.theme_mode.get(),
                "scale":self.scale_mode.get(),
                "geometry":self.geometry(),
                "sidebar_collapsed":self.sidebar_collapsed,
                "output_folder":getattr(self,"output_folder_var",tk.StringVar(value=str(self.ui_settings.get("output_folder", "")))).get()
            })
            save_json(UI_SETTINGS_FILE,data)
        self.scale_mode.trace_add("write",save_ui)
        self.animations_mode.trace_add("write",save_ui)
        self.startup_animation_mode.trace_add("write",save_ui)
        self.startup_duration_mode.trace_add("write",save_ui)

        production=tk.Frame(frame,bg=pal["CARD"],highlightbackground=pal["LINE"],highlightthickness=1);production.pack(fill="x",pady=(12,0))
        tk.Label(production,text="Saída e impressão",bg=pal["CARD"],fg=pal["TEXT"],font=("Segoe UI",12,"bold")).pack(anchor="w",padx=18,pady=(15,3))
        pinfo=tk.Label(production,text="ⓘ",bg=pal["CARD"],fg=pal["BLUE2"],font=("Segoe UI",9,"bold"));pinfo.place(relx=1.0,x=-18,y=14,anchor="ne");add_tooltip(pinfo,"Pasta padrão e perfis de impressão para Promoções, Atacado e Manual.")
        outrow=tk.Frame(production,bg=pal["CARD"]);outrow.pack(fill="x",padx=18,pady=(10,8))
        self.output_folder_var=tk.StringVar(value=str(self.ui_settings.get("output_folder") or default_output_root(self.ui_settings)))
        tk.Label(outrow,text="Pasta padrão dos cartazes",bg=pal["CARD"],fg=pal["TEXT"],font=("Segoe UI",8,"bold")).pack(anchor="w")
        outsub=tk.Frame(outrow,bg=pal["CARD"]);outsub.pack(fill="x",pady=(4,0))
        tk.Entry(outsub,textvariable=self.output_folder_var,bg=pal["ROW_ALT"],fg=pal["TEXT"],insertbackground=pal["TEXT"],relief="flat").pack(side="left",fill="x",expand=True,ipady=6)
        def choose_output_folder():
            folder=filedialog.askdirectory(title="Pasta padrão dos cartazes",initialdir=self.output_folder_var.get() or str(Path.home()))
            if folder:self.output_folder_var.set(folder);self.save_production_settings()
        tk.Button(outsub,text="ESCOLHER",command=choose_output_folder,bg=pal["LIGHT_BLUE"],fg=pal["LIGHT_BLUE_TXT"],relief="flat",font=("Segoe UI",8,"bold"),padx=10,pady=6).pack(side="right",padx=(8,0))

        profiles=(self.startup_cache.get("print_profiles") or load_print_profiles())
        cached_printers=self.startup_cache.get("printers") if isinstance(getattr(self,"startup_cache",{}),dict) else None
        printers=["Impressora padrão do Windows"]+list(cached_printers if isinstance(cached_printers,list) else list_printers())
        # Remove nomes repetidos sem consultar novamente o Windows.
        printers=list(dict.fromkeys(printers));self.print_profile_vars={}
        pgrid=tk.Frame(production,bg=pal["CARD"]);pgrid.pack(fill="x",padx=18,pady=(3,8))
        for i,(key,title) in enumerate([("promo","Promoções"),("atacado","Atacado"),("manual","Manual")]):
            pgrid.grid_columnconfigure(i,weight=1)
            cell=tk.Frame(pgrid,bg=pal["ROW_ALT"],highlightbackground=pal["LINE"],highlightthickness=1);cell.grid(row=0,column=i,sticky="ew",padx=(0 if i==0 else 4,0 if i==2 else 4))
            tk.Label(cell,text=title,bg=pal["ROW_ALT"],fg=pal["TEXT"],font=("Segoe UI",8,"bold")).pack(anchor="w",padx=10,pady=(9,2))
            saved=profiles.get(key,{})
            pv=tk.StringVar(value=saved.get("printer") or "Impressora padrão do Windows");cv=tk.IntVar(value=max(1,int(saved.get("copies") or 1)));self.print_profile_vars[key]=(pv,cv)
            ttk.Combobox(cell,textvariable=pv,state="readonly",values=printers,width=28).pack(fill="x",padx=10,pady=3)
            cr=tk.Frame(cell,bg=pal["ROW_ALT"]);cr.pack(fill="x",padx=10,pady=(3,9));tk.Label(cr,text="Cópias padrão",bg=pal["ROW_ALT"],fg=pal["MUTED"],font=("Segoe UI",7)).pack(side="left");tk.Spinbox(cr,from_=1,to=20,textvariable=cv,width=5).pack(side="right")
        tk.Button(production,text="SALVAR",command=self.save_production_settings,bg=pal["BLUE"],fg="white",relief="flat",font=("Segoe UI",8,"bold"),padx=12,pady=7).pack(anchor="e",padx=18,pady=(0,14))

        # SR IA / OpenAI: chave protegida localmente e permissões de leitura.
        build_sria_settings_card(frame,self,pal)

        learned=tk.Frame(frame,bg=pal["CARD"],highlightbackground=pal["LINE"],highlightthickness=1);learned.pack(fill="x",pady=(12,0))
        tk.Label(learned,text="Correções aprendidas",bg=pal["CARD"],fg=pal["TEXT"],font=("Segoe UI",12,"bold")).pack(anchor="w",padx=18,pady=(15,3))
        learned_count=self.startup_cache.get("corrections_count") if isinstance(getattr(self,"startup_cache",{}),dict) else None
        if learned_count is None:learned_count=len(corrections())
        tk.Label(learned,text=f"{learned_count} memorizada(s)",bg=pal["CARD"],fg=pal["MUTED"],font=("Segoe UI",8)).pack(anchor="w",padx=18)
        tk.Button(learned,text="GERENCIAR",command=self.manage_learned_corrections,bg=pal["LIGHT_BLUE"],fg=pal["LIGHT_BLUE_TXT"],relief="flat",font=("Segoe UI",8,"bold"),padx=12,pady=7).pack(anchor="w",padx=18,pady=(9,14))

        diag=tk.Frame(frame,bg=pal["CARD"],highlightbackground=pal["LINE"],highlightthickness=1);diag.pack(fill="x",pady=(12,0))
        tk.Label(diag,text="Diagnóstico do sistema",bg=pal["CARD"],fg=pal["TEXT"],font=("Segoe UI",12,"bold")).pack(anchor="w",padx=18,pady=(15,3))
        dinfo=tk.Label(diag,text="ⓘ",bg=pal["CARD"],fg=pal["BLUE2"],font=("Segoe UI",9,"bold"));dinfo.place(relx=1.0,x=-18,y=14,anchor="ne");add_tooltip(dinfo,"Verifica PowerPoint, modelos, memória e pasta temporária.")
        grid=tk.Frame(diag,bg=pal["CARD"]);grid.pack(fill="x",padx=18,pady=12)
        self.diag_labels={}
        for i,(key,title) in enumerate([("powerpoint","PowerPoint"),("models","Modelos"),("memory","Memória"),("temp","Pasta temporária")]):
            grid.grid_columnconfigure(i,weight=1)
            c=tk.Frame(grid,bg=pal["ROW_ALT"],highlightbackground=pal["LINE"],highlightthickness=1);c.grid(row=0,column=i,sticky="ew",padx=(0 if i==0 else 4,0 if i==3 else 4))
            tk.Label(c,text=title,bg=pal["ROW_ALT"],fg=pal["TEXT"],font=("Segoe UI",8,"bold")).pack(pady=(9,2))
            l=tk.Label(c,text="Aguardando teste",bg=pal["ROW_ALT"],fg=pal["MUTED"],font=("Segoe UI",8));l.pack(pady=(0,9));self.diag_labels[key]=l
        buttons=tk.Frame(diag,bg=pal["CARD"]);buttons.pack(fill="x",padx=18,pady=(0,15))
        tk.Button(buttons,text="DIAGNÓSTICO",command=self.run_diagnostic,bg=pal["BLUE"],fg="white",relief="flat",font=("Segoe UI",8,"bold"),padx=12,pady=7).pack(side="left")
        tk.Button(buttons,text="CORRIGIR",command=self.auto_correct_environment,bg=pal["ORANGE"],fg=pal["ORANGE_TXT"],relief="flat",font=("Segoe UI",8,"bold"),padx=12,pady=7).pack(side="left",padx=6)
        tk.Button(buttons,text="MODELOS",command=lambda:self.navigate("modelos"),bg=pal["LIGHT_BLUE"],fg=pal["LIGHT_BLUE_TXT"],relief="flat",font=("Segoe UI",8,"bold"),padx=12,pady=7).pack(side="left")

        # Banco de dados CISSPoder — fica imediatamente acima da atualização de versão.
        ciss=tk.Frame(frame,bg=pal["CARD"],highlightbackground=pal["LINE"],highlightthickness=1);ciss.pack(fill="x",pady=(12,0))
        chead=tk.Frame(ciss,bg=pal["CARD"]);chead.pack(fill="x",padx=18,pady=(15,3))
        tk.Label(chead,text="Banco de Dados CISSPoder",bg=pal["CARD"],fg=pal["TEXT"],font=("Segoe UI",12,"bold")).pack(side="left")
        tk.Label(chead,text="RELATÓRIO 208",bg=pal["LIGHT_BLUE"],fg=pal["LIGHT_BLUE_TXT"],font=("Segoe UI",7,"bold"),padx=8,pady=3).pack(side="right")
        cinfo=tk.Label(chead,text="ⓘ",bg=pal["CARD"],fg=pal["BLUE2"],font=("Segoe UI",9,"bold"));cinfo.pack(side="right",padx=(0,8));add_tooltip(cinfo,"Relatório 208: atualiza custo de reposição, varejo e atacado. Código CISS e EAN permanecem separados.")
        last_ciss=ciss_last_import_info()
        if last_ciss:
            dt=str(last_ciss.get("report_datetime") or last_ciss.get("imported_at") or "").replace("T"," ")[:16]
            ciss_last_text=(f"Última atualização: {dt or '—'} • {int(last_ciss.get('total_products') or 0):,} produtos • "
                            f"{int(last_ciss.get('changed_products') or 0):,} alterados").replace(",",".")
        else:
            ciss_last_text="Banco CISSPoder ainda não atualizado neste computador."
        self.ciss_sync_last_label=tk.Label(ciss,text=ciss_last_text,bg=pal["CARD"],fg=pal["TEXT"],font=("Segoe UI",8,"bold"));self.ciss_sync_last_label.pack(anchor="w",padx=18,pady=(9,2))
        self.ciss_sync_status_var=tk.StringVar(value="Selecione um PDF do relatório 208 para atualizar o Banco de Produtos.")
        tk.Label(ciss,textvariable=self.ciss_sync_status_var,bg=pal["CARD"],fg=pal["MUTED"],font=("Segoe UI",8),anchor="w",justify="left").pack(fill="x",padx=18,pady=(0,5))
        self.ciss_sync_progress_var=tk.DoubleVar(value=0)
        self.ciss_sync_progress=ttk.Progressbar(ciss,variable=self.ciss_sync_progress_var,maximum=100,style="SR.Horizontal.TProgressbar")
        self.ciss_sync_progress.pack(fill="x",padx=18,pady=(0,8))
        cb=tk.Frame(ciss,bg=pal["CARD"]);cb.pack(fill="x",padx=18,pady=(0,15))
        self.ciss_sync_button=tk.Button(cb,text="ATUALIZAR BANCO",command=self.import_ciss_database,bg=pal["BLUE"],fg="white",relief="flat",font=("Segoe UI",8,"bold"),padx=14,pady=8);self.ciss_sync_button.pack(side="left")
        add_tooltip(self.ciss_sync_button,"Importa o relatório 208. A leitura e atualização acontecem em segundo plano, com progresso nesta própria tela.")
        

        updates=tk.Frame(frame,bg=pal["CARD"],highlightbackground=pal["LINE"],highlightthickness=1);updates.pack(fill="x",pady=(12,0))
        uhead=tk.Frame(updates,bg=pal["CARD"]);uhead.pack(fill="x",padx=18,pady=(15,3))
        tk.Label(uhead,text="Atualizações do SR Studio",bg=pal["CARD"],fg=pal["TEXT"],font=("Segoe UI",12,"bold")).pack(side="left")
        tk.Label(uhead,text=f"Versão instalada: {APP_DISPLAY_VERSION}",bg=pal["CARD"],fg=pal["BLUE2"],font=("Segoe UI",8,"bold")).pack(side="right")
        uinfo=tk.Label(uhead,text="ⓘ",bg=pal["CARD"],fg=pal["BLUE2"],font=("Segoe UI",9,"bold"));uinfo.pack(side="right",padx=(0,8));add_tooltip(uinfo,"Pacotes .srupdate são validados e aplicados com backup. Bancos e relatórios do usuário são preservados.")
        uhist=load_json(UPDATE_HISTORY_FILE,[])
        last=(uhist[-1] if isinstance(uhist,list) and uhist else None)
        last_text=(f"Última atualização: {last.get('from','')} → {last.get('to','')} • {last.get('date','')}" if last else "Nenhuma atualização importada ainda.")
        self.update_status_label=tk.Label(updates,text=last_text,bg=pal["CARD"],fg=pal["MUTED"],font=("Segoe UI",8));self.update_status_label.pack(anchor="w",padx=18,pady=(8,8))
        ub=tk.Frame(updates,bg=pal["CARD"]);ub.pack(fill="x",padx=18,pady=(0,15))
        btn_up=tk.Button(ub,text="IMPORTAR .SRUPDATE",command=self.import_update_package,bg=pal["BLUE"],fg="white",relief="flat",font=("Segoe UI",8,"bold"),padx=14,pady=8);btn_up.pack(side="left")
        add_tooltip(btn_up,"Selecione um pacote .srupdate fornecido para o SR Studio. Um backup é criado antes de substituir qualquer arquivo.")
        # O diagnóstico pesado não roda mais no clique de Configurações; a saúde já foi verificada na abertura.
        cached_health=self.startup_cache.get("health",{}) if isinstance(getattr(self,"startup_cache",{}),dict) else {}
        if cached_health:
            self.after(50,lambda h=dict(cached_health):self.apply_diagnostic({"powerpoint":h.get("powerpoint",False),"models":h.get("models",False),"memory":h.get("memory",False),"temp":True}))

    def _set_ciss_sync_progress(self,value,text):
        try:
            if hasattr(self,"ciss_sync_progress_var"):self.ciss_sync_progress_var.set(max(0,min(100,float(value))))
            if hasattr(self,"ciss_sync_status_var"):self.ciss_sync_status_var.set(str(text or ""))
            self.update_idletasks()
        except Exception:
            pass

    def import_ciss_database(self):
        if getattr(self,"busy",False):
            self.toast.show("Aguarde a tarefa atual terminar antes de atualizar o banco CISSPoder.","warn")
            return
        pdf=filedialog.askopenfilename(
            title="Selecionar relatório 208 do CISSPoder",
            filetypes=[("Relatório CISSPoder em PDF","*.pdf")],
            parent=self
        )
        if not pdf:return
        self.busy=True
        try:self.ciss_sync_button.config(state="disabled",text="ATUALIZANDO BANCO...")
        except Exception:pass
        self._set_ciss_sync_progress(1,"Validando relatório 208 do CISSPoder...")
        started=time.time()
        def progress(v,t):
            self.after(0,lambda v=v,t=t:self._set_ciss_sync_progress(v,t))
        def worker():
            try:
                result=import_ciss_report_208(pdf,progress)
                elapsed=time.time()-started
                self.after(0,lambda r=result,e=elapsed:self._finish_ciss_database_update(r,e))
            except Exception as exc:
                msg=str(exc)
                self.after(0,lambda msg=msg:self._fail_ciss_database_update(msg))
        threading.Thread(target=worker,daemon=True).start()

    def _finish_ciss_database_update(self,result,elapsed=0):
        self.busy=False
        total=int(result.get("total_products") or 0);changed=int(result.get("changed_products") or 0)
        costs=int(result.get("changed_cost") or 0);retail=int(result.get("changed_retail") or 0)
        linked=int(result.get("exact_links") or 0);auto=int(result.get("auto_links") or 0);review=int(result.get("review_duplicates") or 0);only=int(result.get("ciss_only") or 0)
        dt=str(result.get("report_datetime") or "").replace("T"," ")[:16]
        summary=(f"Concluído em {elapsed:.1f}s • {total:,} produtos • {changed:,} alterados • "
                 f"{costs:,} custos • {retail:,} varejos").replace(",",".")
        self._set_ciss_sync_progress(100,summary)
        try:
            self.ciss_sync_last_label.config(text=(f"Última atualização: {dt or 'agora'} • {total:,} produtos • "
                                                   f"{linked+auto:,} vinculados • {review:,} revisar • {only:,} somente CISS").replace(",","."))
            self.ciss_sync_button.config(state="normal",text="ATUALIZAR BANCO")
        except Exception:pass
        try:
            invalidate_catalog_cache();invalidate_builder_catalog_cache()
            self.startup_cache["product_catalog_count"]=preload_product_catalog(force=True)
            self.startup_cache["catalog_counts"]=catalog_counts()
            self.startup_cache["sale_catalog_count"]=preload_sale_catalog(force=True)
            self.startup_cache["builder_catalog_count"]=preload_builder_catalog(force=True)
            self.startup_cache["library_cache_count"]=preload_library_cache(force=True)
            self.startup_cache["ciss_last_import"]=ciss_last_import_info() or {}
        except Exception:pass
        self.toast.show(f"Banco CISSPoder atualizado: {total:,} produtos.".replace(",","."),"ok")

    def _fail_ciss_database_update(self,msg):
        self.busy=False
        self._set_ciss_sync_progress(0,"Falha ao atualizar banco CISSPoder: "+str(msg))
        try:self.ciss_sync_button.config(state="normal",text="ATUALIZAR BANCO")
        except Exception:pass
        self.toast.show("Não foi possível atualizar o banco CISSPoder.","error")

    def apply_theme_setting(self):
        data=load_json(UI_SETTINGS_FILE,{})
        data.update({"theme":self.theme_mode.get(),"geometry":self.geometry(),"animations":self.animations_mode.get(),"startup_animation":self.startup_animation_mode.get(),"startup_duration":self.startup_duration_mode.get(),"scale":self.scale_mode.get(),"sidebar_collapsed":self.sidebar_collapsed,"output_folder":getattr(self,"output_folder_var",tk.StringVar(value="")).get()})
        save_json(UI_SETTINGS_FILE,data)
        if messagebox.askyesno("Aplicar tema",f"Tema '{self.theme_mode.get()}' salvo.\n\nReiniciar o SR Studio agora para aplicar em toda a interface?"):
            self.restart_application()

    def restart_application(self):
        try:
            if RUNNING_FLAG.exists():RUNNING_FLAG.unlink()
        except Exception:pass
        try:
            if getattr(sys,"frozen",False) or os.environ.get("SR_STUDIO_EXE")=="1":
                cmd=[sys.executable]
            else:
                cmd=[sys.executable,str(APP_DIR/"SR_Studio_Gerador.py")]
            subprocess.Popen(cmd,cwd=str(APP_DIR),creationflags=getattr(subprocess,"CREATE_NEW_PROCESS_GROUP",0))
            self.destroy()
        except Exception as e:messagebox.showerror("Reiniciar",str(e))

    def save_production_settings(self):
        data=load_json(UI_SETTINGS_FILE,{})
        folder=getattr(self,"output_folder_var",tk.StringVar(value="")).get().strip()
        if folder:
            try:Path(folder).mkdir(parents=True,exist_ok=True)
            except Exception as e:messagebox.showerror("Pasta de saída",f"Não foi possível usar esta pasta.\n\n{e}");return
        data.update({"output_folder":folder,"theme":self.theme_mode.get(),"geometry":self.geometry(),"animations":self.animations_mode.get(),"startup_animation":self.startup_animation_mode.get(),"startup_duration":self.startup_duration_mode.get(),"scale":self.scale_mode.get(),"sidebar_collapsed":self.sidebar_collapsed})
        save_json(UI_SETTINGS_FILE,data);self.ui_settings=data
        profiles=load_print_profiles()
        for key,(pv,cv) in getattr(self,"print_profile_vars",{}).items():
            printer=pv.get().strip();profiles[key]={"printer":"" if printer=="Impressora padrão do Windows" else printer,"copies":max(1,int(cv.get() or 1))}
        save_print_profiles(profiles);self.toast.show("Configurações de saída e impressão salvas.","ok")

    def manage_learned_corrections(self):
        pal=self.palette;w=tk.Toplevel(self);w.title("Correções aprendidas");w.configure(bg=pal["APP_BG"]);center_toplevel(w,self,800,520)
        card=tk.Frame(w,bg=pal["CARD"],highlightbackground=pal["LINE"],highlightthickness=1);card.pack(fill="both",expand=True,padx=18,pady=18)
        tree=ttk.Treeview(card,columns=("orig","dest"),show="headings",selectmode="browse");tree.heading("orig",text="Chave original");tree.heading("dest",text="Nome aplicado");tree.column("orig",width=330);tree.column("dest",width=390);tree.pack(fill="both",expand=True,padx=10,pady=10)
        def refresh():
            tree.delete(*tree.get_children())
            for i,(k,v) in enumerate(sorted(corrections().items())):tree.insert("","end",iid=str(i),values=(k,v),tags=(k,))
        def remove():
            sel=tree.selection()
            if not sel:return
            key=tree.item(sel[0],"values")[0]
            if messagebox.askyesno("Correções",f"Remover a correção aprendida para:\n{key}?"):
                from SRStudio21 import remove_correction
                remove_correction(key);refresh()
        tk.Button(card,text="REMOVER CORREÇÃO SELECIONADA",command=remove,bg=pal["RED"],fg=pal["RED_TXT"],relief="flat",font=("Segoe UI",8,"bold"),padx=12,pady=7).pack(anchor="e",padx=10,pady=(0,10));refresh()

    def _show_update_overlay(self, target_version=""):
        """Overlay interno animado durante instalação de .srupdate."""
        try:
            old=getattr(self,"_update_overlay",None)
            if old and old.winfo_exists(): old.destroy()
        except Exception: pass
        pal=self.palette
        ov=tk.Frame(self,bg="#0E1828",bd=0,highlightthickness=0)
        ov.place(x=0,y=0,relwidth=1,relheight=1);ov.lift()
        card=tk.Frame(ov,bg=pal["CARD"],highlightbackground=pal["LINE"],highlightthickness=1)
        card.place(relx=.5,rely=.48,anchor="center",width=520,height=250)
        logo=tk.Label(card,text="SR",bg=pal["BLUE"],fg="white",font=("Segoe UI",22,"bold"),width=3,height=1)
        logo.pack(pady=(28,10))
        tk.Label(card,text="ATUALIZANDO SR STUDIO",bg=pal["CARD"],fg=pal["TEXT"],font=("Segoe UI",14,"bold")).pack()
        subtitle=(f"Preparando a versão {target_version}" if target_version else "Aplicando atualização")
        self._update_overlay_text=tk.StringVar(value=subtitle)
        tk.Label(card,textvariable=self._update_overlay_text,bg=pal["CARD"],fg=pal["MUTED"],font=("Segoe UI",9)).pack(pady=(5,12))
        bar=ttk.Progressbar(card,mode="indeterminate",length=360)
        bar.pack();bar.start(9)
        self._update_overlay_bar=bar
        self._update_overlay_dots=tk.StringVar(value="●  ○  ○")
        tk.Label(card,textvariable=self._update_overlay_dots,bg=pal["CARD"],fg=pal["BLUE2"],font=("Segoe UI Symbol",11,"bold")).pack(pady=(10,4))
        tk.Label(card,text="Não feche o programa durante esta etapa.",bg=pal["CARD"],fg=pal["MUTED"],font=("Segoe UI",8)).pack()
        self._update_overlay=ov;self._update_anim_tick=0;self._animate_update_overlay()
        try:self.update_idletasks()
        except Exception:pass

    def _animate_update_overlay(self):
        ov=getattr(self,"_update_overlay",None)
        try:
            if not ov or not ov.winfo_exists():return
        except Exception:return
        frames=("●  ○  ○","○  ●  ○","○  ○  ●","○  ●  ○")
        steps=("Validando arquivos da atualização...","Criando backup de segurança...","Aplicando novos componentes...","Atualizando o motor e a interface...","Finalizando e verificando arquivos...")
        tick=int(getattr(self,"_update_anim_tick",0));self._update_anim_tick=tick+1
        try:self._update_overlay_dots.set(frames[tick%len(frames)])
        except Exception:pass
        try:
            if tick%5==0:self._update_overlay_text.set(steps[min(tick//5,len(steps)-1)])
        except Exception:pass
        try:ov.after(320,self._animate_update_overlay)
        except Exception:pass

    def _hide_update_overlay(self):
        try:
            bar=getattr(self,"_update_overlay_bar",None)
            if bar:bar.stop()
        except Exception:pass
        try:
            ov=getattr(self,"_update_overlay",None)
            if ov and ov.winfo_exists():ov.destroy()
        except Exception:pass
        self._update_overlay=None

    def import_update_package(self):
        if getattr(self,"busy",False):
            messagebox.showinfo("Atualização","Aguarde a tarefa atual terminar antes de atualizar o SR Studio.")
            return
        package=filedialog.askopenfilename(title="Importar atualização do SR Studio",filetypes=[("Atualização SR Studio","*.srupdate"),("Arquivo ZIP","*.zip")])
        if not package:return
        try:
            manifest=inspect_update(package,APP_VERSION)
        except Exception as e:
            messagebox.showerror("Atualização inválida",str(e));return
        notes=str(manifest.get("notes","")).strip()
        msg=(f"Atualizar SR Studio {APP_VERSION} para {manifest.get('to_version')}?\n\n"
             f"Arquivos do programa serão atualizados e um backup será criado automaticamente.")
        if notes:msg+=f"\n\nNovidades/correções:\n{notes}"
        if not messagebox.askyesno("Confirmar atualização",msg):return
        self.busy=True
        if hasattr(self,"update_status_label"):self.update_status_label.config(text="Aplicando atualização...")
        self._show_update_overlay(str(manifest.get("to_version") or ""))
        def worker():
            try:
                result,backup=apply_update(package,APP_DIR,APP_VERSION,UPDATE_HISTORY_FILE)
                self.after(0,lambda:self._finish_update(result,backup))
            except Exception as e:
                msg=str(e)
                self.after(0,lambda msg=msg:self._update_error(msg))
        threading.Thread(target=worker,daemon=True).start()

    def _update_error(self,e):
        self.busy=False
        self._hide_update_overlay()
        if hasattr(self,"update_status_label"):self.update_status_label.config(text="Falha ao aplicar atualização.")
        messagebox.showerror("Atualização",str(e))

    def _finish_update(self,manifest,backup):
        self.busy=False
        self._hide_update_overlay()
        target=manifest.get("to_version","")
        if hasattr(self,"update_status_label"):self.update_status_label.config(text=f"Atualização {target} aplicada. Reinicie para ativar.")
        restart=messagebox.askyesno("Atualização concluída",f"SR Studio atualizado para {target}.\n\nBackup criado em:\n{backup}\n\nDeseja reiniciar agora?")
        if restart:self.restart_after_update()

    def restart_after_update(self):
        try:
            if RUNNING_FLAG.exists():RUNNING_FLAG.unlink()
        except Exception:pass
        try:
            exe_env=str(os.environ.get("SR_STUDIO_LAUNCHER_PATH","")).strip()
            exe_path=Path(exe_env) if exe_env else (APP_DIR/"SR Studio.exe")
            bat_path=APP_DIR/"INICIAR_SR_STUDIO.bat"

            if os.name=="nt" and exe_path.exists():
                os.startfile(str(exe_path))
            elif os.name=="nt" and bat_path.exists():
                os.startfile(str(bat_path))
            else:
                subprocess.Popen(
                    [sys.executable,str(APP_DIR/"SR_Studio_Gerador.py")],
                    cwd=str(APP_DIR),
                    creationflags=getattr(subprocess,"CREATE_NEW_PROCESS_GROUP",0)
                )
            self.destroy()
        except Exception as e:
            messagebox.showerror("Reiniciar",f"A atualização foi aplicada, mas o reinício automático falhou.\n\nFeche e abra o SR Studio manualmente.\n\n{e}")

    def run_diagnostic(self):
        if not hasattr(self,"diag_labels"):return
        for l in self.diag_labels.values():l.config(text="Verificando...",fg=self.palette["MUTED"])
        def worker():
            res={"powerpoint":False,"models":False,"memory":False,"temp":False}
            try:res["models"]=all(Path(x).exists() for x in [MODEL1,MODEL2,MODEL1_LIMIT,MODEL2_LIMIT,CLUB_MODEL,CLUB_MODEL_LIMIT,ATACADO_MODEL,MODEL_SALE])
            except Exception:pass
            try:
                test=LOCAL_DATA/"diag_write.tmp";test.write_text("ok",encoding="utf-8");test.unlink();res["memory"]=True
            except Exception:pass
            try:
                td=Path(tempfile.gettempdir())/"srstudio_diag.tmp";td.write_text("ok",encoding="utf-8");td.unlink();res["temp"]=True
            except Exception:pass
            if os.name=="nt":
                try:
                    cmd=[find_powershell(),"-NoProfile","-Command","if([type]::GetTypeFromProgID('PowerPoint.Application')){'OK'}"]
                    r=subprocess.run(cmd,capture_output=True,text=True,timeout=8,creationflags=getattr(subprocess,"CREATE_NO_WINDOW",0));res["powerpoint"]="OK" in (r.stdout or "")
                except Exception:pass
            self.after(0,lambda:self.apply_diagnostic(res))
        threading.Thread(target=worker,daemon=True).start()
    def apply_diagnostic(self,res):
        for key,val in res.items():
            if key in getattr(self,"diag_labels",{}):self.diag_labels[key].config(text="✓ OK" if val else "! ATENÇÃO",fg=self.palette["GREEN_TXT"] if val else self.palette["ORANGE_TXT"])
        self._apply_health({"powerpoint":res.get("powerpoint"),"models":res.get("models"),"memory":res.get("memory"),"backup":True})
    def auto_correct_environment(self):
        changed=ensure_all_models_unlocked()
        try:
            (APP_DIR/"dados").mkdir(parents=True,exist_ok=True);(LOCAL_DATA/"preview").mkdir(parents=True,exist_ok=True);(LOCAL_DATA/"thumb_cache").mkdir(parents=True,exist_ok=True)
            self.toast.show(f"Correção automática concluída • {changed} trava(s) removida(s).","ok")
        except Exception as e:messagebox.showerror("Correção automática",str(e));return
        self.run_diagnostic();self.refresh_health_async()

    def clear_content(self):
        for w in self.content.winfo_children():
            w.destroy()

    def _open_pending_items(self, signature, promo_attention, atacado_pending, atacado_alerts):
        """Marca as pendências atuais como vistas e abre a área correspondente."""
        try:
            save_json(PENDING_STATE_FILE,{
                "seen_signature": signature,
                "seen_at": datetime.now().isoformat(timespec="seconds"),
            })
        except Exception:
            pass
        # O aviso some no próximo retorno ao Dashboard, mas a pendência real continua
        # disponível dentro do módulo até ser efetivamente resolvida.
        if promo_attention:
            self.navigate("promo")
        elif atacado_pending or atacado_alerts:
            self.navigate("atacado")
        else:
            self.toast.show("Não há novas pendências para revisar.","ok")

    def refresh_home_if_visible(self):
        """Atualiza o Dashboard imediatamente quando uma pendência é concluída."""
        try:
            if self.page_title.cget("text")=="Início" and not getattr(self,"busy",False):
                self.show_home()
        except Exception:
            pass

    def show_home(self):
        self.clear_content();self.page_title.config(text="Início")
        for k,b in self.nav_buttons.items():
            b.config(bg=self.palette["SIDEBAR_HOVER"] if k=="home" else self.palette["SIDEBAR"],fg="white" if k=="home" else "#DCE7F7")
        pal=self.palette
        canvas=tk.Canvas(self.content,bg=pal["APP_BG"],highlightthickness=0)
        sb=ttk.Scrollbar(self.content,orient="vertical",command=canvas.yview);canvas.configure(yscrollcommand=sb.set)
        canvas.pack(side="left",fill="both",expand=True);sb.pack(side="right",fill="y")
        frame=tk.Frame(canvas,bg=pal["APP_BG"]);win=canvas.create_window((0,0),window=frame,anchor="nw")
        frame.bind("<Configure>",lambda e:canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>",lambda e:canvas.itemconfigure(win,width=e.width))
        frame.configure(padx=26,pady=20)

        # Cabeçalho curto: o Dashboard é uma central de ação, não uma tela de relatório.
        head=tk.Frame(frame,bg=pal["APP_BG"]);head.pack(fill="x",pady=(0,14))
        left_head=tk.Frame(head,bg=pal["APP_BG"]);left_head.pack(side="left",fill="x",expand=True)
        tk.Label(left_head,text="SR Studio",bg=pal["APP_BG"],fg=pal["TEXT"],font=("Segoe UI",25,"bold")).pack(anchor="w")
        tk.Label(left_head,text="Central de produção de cartazes",bg=pal["APP_BG"],fg=pal["MUTED"],font=("Segoe UI",9)).pack(anchor="w",pady=(1,0))
        ready=tk.Label(head,text="●  SR STUDIO PRONTO",bg=pal["GREEN"],fg=pal["GREEN_TXT"],font=("Segoe UI",8,"bold"),padx=12,pady=6)
        ready.pack(side="right",anchor="n",pady=4);add_tooltip(ready,"PowerPoint, modelos, banco e preferências são verificados na abertura.")

        hist=load_json(HISTORY_FILE,[]);today=datetime.now().strftime("%d/%m/%Y")
        today_count=sum(int(x.get("cartazes",0)) for x in hist if x.get("data","").startswith(today))
        promo_attention=(len(self.analysis.get("errors",[]))+len(self.analysis.get("warnings",[]))) if self.analysis else 0
        try:
            rh=reports_history();latest_at=rh[0] if rh else None
            atacado_pending=(int(latest_at["novos"] or 0)+int(latest_at["alterados"] or 0)) if latest_at else 0
            atacado_alerts=int(latest_at["alertas"] or 0) if latest_at else 0
        except Exception:
            latest_at=None;atacado_pending=0;atacado_alerts=0
        try:
            cc=catalog_counts();product_count=int(cc.get("unique",0));review_count=int(cc.get("review",0))
        except Exception:
            product_count=0;review_count=0
        raw_pending=promo_attention+atacado_pending+atacado_alerts+review_count
        pending_signature=cache_key(json.dumps({
            "promo_file": self.file_path.get() if self.analysis else "",
            "promo_errors": list((self.analysis or {}).get("errors",[])),
            "promo_warnings": list((self.analysis or {}).get("warnings",[])),
            "atacado_id": int(latest_at["id"]) if latest_at and "id" in latest_at.keys() else 0,
            "atacado_new": atacado_pending,"atacado_alerts": atacado_alerts,"catalog_review":review_count,
        },ensure_ascii=False,sort_keys=True))
        seen_state=load_json(PENDING_STATE_FILE,{})
        pending_seen=(raw_pending>0 and seen_state.get("seen_signature")==pending_signature)
        pending=0 if pending_seen else raw_pending

        # Ações principais: poucas palavras, cartões grandes e clicáveis.
        actions=tk.Frame(frame,bg=pal["APP_BG"]);actions.pack(fill="x",pady=(0,14))
        for i in range(6):actions.grid_columnconfigure(i,weight=1,uniform="homeaction")
        def action_card(col,icon,title,caption,cmd,bg,fg,badge=""):
            c=tk.Frame(actions,bg=bg,highlightbackground=pal["LINE"],highlightthickness=1,cursor="hand2",height=126)
            c.grid(row=0,column=col,sticky="nsew",padx=(0 if col==0 else 4,0 if col==5 else 4));c.grid_propagate(False)
            top=tk.Frame(c,bg=bg);top.pack(fill="x",padx=12,pady=(12,4))
            tk.Label(top,text=icon,bg=bg,fg=fg,font=("Segoe UI Symbol",17,"bold")).pack(side="left")
            if badge:
                tk.Label(top,text=badge,bg=pal["CARD"],fg=fg,font=("Segoe UI",7,"bold"),padx=6,pady=2).pack(side="right")
            tk.Label(c,text=title,bg=bg,fg=fg,font=("Segoe UI",9,"bold")).pack(anchor="w",padx=12,pady=(1,0))
            tk.Label(c,text=caption,bg=bg,fg=fg,font=("Segoe UI",7),wraplength=160,justify="left").pack(anchor="w",padx=12,pady=(3,9))
            for w in [c]+c.winfo_children()+top.winfo_children():
                w.bind("<Button-1>",lambda e:cmd())
            return c
        action_card(0,"◆","Montador","Criar sem Excel",lambda:self.navigate("builder"),pal["LIGHT_BLUE"],pal["LIGHT_BLUE_TXT"])
        action_card(1,"▥","Encartes Studio","Criar e editar encartes",lambda:self.navigate("encartes"),pal["CARD"],pal["BLUE2"])
        action_card(2,"▣","Promoções","Importar planilha",lambda:self.navigate("promo"),pal["PURPLE"],pal["PURPLE_TXT"])
        action_card(3,"▦","Atacado","Novos e alterações",lambda:self.navigate("atacado"),pal["ORANGE"],pal["ORANGE_TXT"],str(atacado_pending) if atacado_pending else "")
        action_card(4,"＋","Manual","Um cartaz rápido",lambda:self.navigate("manual"),pal["GREEN"],pal["GREEN_TXT"])
        action_card(5,"◫","Banco","Produtos e imagens",lambda:self.navigate("products"),pal["CARD"],pal["TEXT"],f"{product_count:,}".replace(",",".") if product_count else "")

        ai_banner=tk.Frame(frame,bg=pal["LIGHT_BLUE"],highlightbackground=pal["LINE"],highlightthickness=1,cursor="hand2")
        ai_banner.pack(fill="x",pady=(0,14))
        tk.Label(ai_banner,text="✦  SR IA",bg=pal["LIGHT_BLUE"],fg=pal["LIGHT_BLUE_TXT"],font=("Segoe UI",10,"bold")).pack(side="left",padx=(14,8),pady=9)
        tk.Label(ai_banner,text="OpenAI focada no SR Studio • Guardião de créditos ativo • Beta 5 somente leitura",bg=pal["LIGHT_BLUE"],fg=pal["LIGHT_BLUE_TXT"],font=("Segoe UI",8)).pack(side="left",pady=9)
        tk.Button(ai_banner,text="ABRIR SR IA →",command=lambda:self.navigate("sria"),bg=pal["BLUE"],fg="white",relief="flat",font=("Segoe UI",8,"bold"),padx=11,pady=6).pack(side="right",padx=10,pady=6)

        # Faixa compacta com o que importa hoje.
        strip=tk.Frame(frame,bg=pal["CARD"],highlightbackground=pal["LINE"],highlightthickness=1);strip.pack(fill="x",pady=(0,14))
        metrics=[("HOJE",today_count,"cartazes"),("ATACADO",atacado_pending,"pendentes"),("BANCO",product_count,"produtos"),("ATENÇÃO",pending,"itens")]
        for i,(title,value,caption) in enumerate(metrics):
            box=tk.Frame(strip,bg=pal["CARD"]);box.pack(side="left",fill="x",expand=True,padx=(14 if i==0 else 8,14 if i==len(metrics)-1 else 8),pady=10)
            tk.Label(box,text=title,bg=pal["CARD"],fg=pal["MUTED"],font=("Segoe UI",7,"bold")).pack(anchor="w")
            line=tk.Frame(box,bg=pal["CARD"]);line.pack(fill="x")
            tk.Label(line,text=str(value),bg=pal["CARD"],fg=pal["RED_TXT"] if title=="ATENÇÃO" and value else pal["TEXT"],font=("Segoe UI",15,"bold")).pack(side="left")
            tk.Label(line,text=caption,bg=pal["CARD"],fg=pal["MUTED"],font=("Segoe UI",8)).pack(side="left",padx=(6,0),pady=(5,0))
        if pending:
            pending_btn=tk.Button(strip,text="REVISAR →",command=lambda:self._open_pending_items(pending_signature,promo_attention,atacado_pending,atacado_alerts),bg=pal["RED"],fg=pal["RED_TXT"],relief="flat",font=("Segoe UI",8,"bold"),padx=12,pady=7)
            pending_btn.pack(side="right",padx=12)

        lower=tk.Frame(frame,bg=pal["APP_BG"]);lower.pack(fill="x")
        lower.grid_columnconfigure(0,weight=3);lower.grid_columnconfigure(1,weight=2)
        recent=tk.Frame(lower,bg=pal["CARD"],highlightbackground=pal["LINE"],highlightthickness=1);recent.grid(row=0,column=0,sticky="nsew",padx=(0,7))
        rhd=tk.Frame(recent,bg=pal["CARD"]);rhd.pack(fill="x",padx=15,pady=(12,6))
        tk.Label(rhd,text="Atividade recente",bg=pal["CARD"],fg=pal["TEXT"],font=("Segoe UI",10,"bold")).pack(side="left")
        tk.Button(rhd,text="HISTÓRICO",command=lambda:self.navigate("historico"),bg=pal["ROW_ALT"],fg=pal["MUTED"],relief="flat",font=("Segoe UI",7,"bold"),padx=9,pady=4).pack(side="right")
        if hist:
            for h in reversed(hist[-3:]):
                row=tk.Frame(recent,bg=pal["ROW_ALT"]);row.pack(fill="x",padx=15,pady=3)
                tk.Label(row,text=str(h.get("data","")).split(" ")[0],bg=pal["ROW_ALT"],fg=pal["BLUE2"],font=("Segoe UI",8,"bold")).pack(side="left",padx=8,pady=7)
                tk.Label(row,text=f"{h.get('cartazes',0)} cartazes • {h.get('campanhas','')}",bg=pal["ROW_ALT"],fg=pal["TEXT"],font=("Segoe UI",8),anchor="w").pack(side="left",fill="x",expand=True)
        else:
            tk.Label(recent,text="Ainda não há gerações registradas.",bg=pal["CARD"],fg=pal["MUTED"],font=("Segoe UI",9)).pack(anchor="w",padx=15,pady=(8,18))

        continue_card=tk.Frame(lower,bg=pal["CARD"],highlightbackground=pal["LINE"],highlightthickness=1);continue_card.grid(row=0,column=1,sticky="nsew",padx=(7,0))
        tk.Label(continue_card,text="Continuar",bg=pal["CARD"],fg=pal["TEXT"],font=("Segoe UI",10,"bold")).pack(anchor="w",padx=15,pady=(12,8))
        if self.analysis and self.file_path.get():
            tk.Label(continue_card,text=Path(self.file_path.get()).name,bg=pal["CARD"],fg=pal["TEXT"],font=("Segoe UI",9,"bold"),wraplength=330,justify="left").pack(anchor="w",padx=15)
            tk.Label(continue_card,text=f"{len(self.analysis.get('jobs',[]))} produtos carregados",bg=pal["CARD"],fg=pal["MUTED"],font=("Segoe UI",8)).pack(anchor="w",padx=15,pady=(2,8))
            tk.Button(continue_card,text="ABRIR PROMOÇÃO →",command=lambda:self.navigate("promo"),bg=pal["BLUE"],fg="white",relief="flat",font=("Segoe UI",8,"bold"),padx=12,pady=7).pack(anchor="w",padx=15,pady=(0,14))
        elif latest_at:
            tk.Label(continue_card,text="Atacado",bg=pal["CARD"],fg=pal["TEXT"],font=("Segoe UI",9,"bold")).pack(anchor="w",padx=15)
            tk.Label(continue_card,text=f"{atacado_pending} itens aguardando ação",bg=pal["CARD"],fg=pal["MUTED"],font=("Segoe UI",8)).pack(anchor="w",padx=15,pady=(2,8))
            tk.Button(continue_card,text="ABRIR ATACADO →",command=lambda:self.navigate("atacado"),bg=pal["BLUE"],fg="white",relief="flat",font=("Segoe UI",8,"bold"),padx=12,pady=7).pack(anchor="w",padx=15,pady=(0,14))
        else:
            tk.Label(continue_card,text="Comece pelo Montador, Promoções ou Geração Manual.",bg=pal["CARD"],fg=pal["MUTED"],font=("Segoe UI",8),wraplength=320,justify="left").pack(anchor="w",padx=15,pady=(0,14))

    def show_placeholder(self, key):
        self.clear_content()
        names = {"atacado":"Atacado","config":"Configurações"}
        self.page_title.config(text=names.get(key,key.title()))
        frame = tk.Frame(self.content, bg=APP_BG)
        frame.pack(fill="both", expand=True, padx=30, pady=28)
        card = tk.Frame(frame, bg=CARD, highlightbackground=LINE, highlightthickness=1)
        card.pack(fill="both", expand=True)
        tk.Label(card, text=names.get(key,key.title()), bg=CARD, fg=TEXT,
                 font=("Segoe UI", 22, "bold")).pack(pady=(90,10))
        tk.Label(card, text="Área preparada para próxima etapa.", bg=CARD, fg=MUTED,
                 font=("Segoe UI", 11)).pack()

    def make_stat(self,parent,col,title,value,bg,fg):
        f=tk.Frame(parent,bg=bg,highlightbackground=self.palette["LINE"],highlightthickness=1)
        f.grid(row=0,column=col,sticky="ew",padx=(0 if col==0 else 3,0 if col==5 else 3))
        row=tk.Frame(f,bg=bg);row.pack(fill="x",padx=9,pady=7)
        val=tk.Label(row,text=value,bg=bg,fg=fg,font=("Segoe UI",13,"bold"))
        val.pack(side="left")
        tk.Label(row,text=title,bg=bg,fg=fg,font=("Segoe UI",7,"bold")).pack(side="left",padx=(6,0),pady=(3,0))
        return val

    def show_promotions(self):
        self.clear_content()
        self.page_title.config(text="Promoções")
        for k,b in self.nav_buttons.items():
            b.config(bg=SIDEBAR_HOVER if k=="promo" else SIDEBAR,
                     fg="white" if k=="promo" else "#DCE7F7")

        promo_canvas=tk.Canvas(self.content,bg=APP_BG,highlightthickness=0)
        promo_scroll=ttk.Scrollbar(self.content,orient="vertical",command=promo_canvas.yview)
        promo_canvas.configure(yscrollcommand=promo_scroll.set)
        promo_canvas.pack(side="left",fill="both",expand=True)
        promo_scroll.pack(side="right",fill="y")
        outer=tk.Frame(promo_canvas,bg=APP_BG)
        promo_win=promo_canvas.create_window((0,0),window=outer,anchor="nw")
        outer.bind("<Configure>",lambda e:promo_canvas.configure(scrollregion=promo_canvas.bbox("all")))
        promo_canvas.bind("<Configure>",lambda e:promo_canvas.itemconfigure(promo_win,width=e.width))
        outer.configure(padx=28,pady=22)

        tk.Label(outer,text="Promoções",bg=APP_BG,fg=TEXT,
                 font=("Segoe UI",20,"bold")).pack(anchor="w")
        tk.Label(
            outer,
            text="Importar, revisar e gerar.",
            bg=APP_BG,fg=MUTED,font=("Segoe UI",10)
        ).pack(anchor="w",pady=(3,12))

        body=tk.Frame(outer,bg=APP_BG)
        body.pack(fill="both",expand=True)
        body.grid_columnconfigure(0,weight=3)
        body.grid_columnconfigure(1,weight=2)
        body.grid_rowconfigure(0,weight=1)

        left=tk.Frame(body,bg=APP_BG)
        left.grid(row=0,column=0,sticky="nsew",padx=(0,14))
        right=tk.Frame(body,bg=APP_BG)
        right.grid(row=0,column=1,sticky="nsew")

        def _promo_reflow(event):
            if event.width < 820:
                left.grid_configure(row=0,column=0,padx=0)
                right.grid_configure(row=1,column=0,padx=0,pady=(12,0))
                body.grid_columnconfigure(1,weight=0)
            else:
                left.grid_configure(row=0,column=0,padx=(0,14),pady=0)
                right.grid_configure(row=0,column=1,padx=0,pady=0)
                body.grid_columnconfigure(1,weight=2)
        body.bind("<Configure>",_promo_reflow)

        upload=tk.Frame(left,bg=CARD,highlightbackground=LINE,highlightthickness=1)
        upload.pack(fill="x")
        inner=tk.Frame(upload,bg=CARD)
        inner.pack(fill="x",padx=22,pady=16)
        tk.Label(inner,text="Planilha",bg=CARD,fg=TEXT,
                 font=("Segoe UI",12,"bold")).pack(anchor="w")

        fr=tk.Frame(inner,bg=ROW_ALT,highlightbackground=LINE,highlightthickness=1)
        fr.pack(fill="x",pady=(10,0))
        self.file_label=tk.Label(fr,text="Nenhuma planilha selecionada",bg=ROW_ALT,
                                 fg=MUTED,anchor="w",font=("Segoe UI",9))
        self.file_label.pack(side="left",fill="x",expand=True,padx=12)
        self.import_btn=tk.Button(fr,text="Selecionar .XLSX",command=self.pick_file,
                                  bg=BLUE,fg="white",relief="flat",
                                  font=("Segoe UI",9,"bold"),padx=14,pady=8)
        self.import_btn.pack(side="right",padx=8,pady=8)
        self.builder_open_btn=tk.Button(fr,text="CRIAR SEM EXCEL",command=lambda:self.navigate("builder"),
                                        bg=GREEN,fg=GREEN_TXT,relief="flat",
                                        font=("Segoe UI",8,"bold"),padx=12,pady=8)
        self.builder_open_btn.pack(side="right",padx=(4,0),pady=8)
        add_tooltip(self.builder_open_btn,"Abra o Montador de Promoções para criar uma campanha direto no SR Studio.")

        promo_clear_row=tk.Frame(inner,bg=CARD)
        promo_clear_row.pack(fill="x",pady=(8,0))
        self.clear_promo_btn=tk.Button(
            promo_clear_row,text="LIMPAR GERAÇÃO",command=self.clear_promotion_generation,
            bg=RED,fg=RED_TXT,activebackground=RED,activeforeground=RED_TXT,
            relief="flat",bd=0,font=("Segoe UI",8,"bold"),padx=14,pady=7,
            state="disabled"
        )
        self.clear_promo_btn.pack(side="right")
        add_tooltip(
            self.clear_promo_btn,
            "Remove a planilha/importação atual e apaga o auto-save desta geração. "
            "O histórico e os PDFs já gerados permanecem."
        )
        if enable_drop(fr,self.import_file,{".xlsx"}):
            self.file_label.config(text="Arraste a planilha .XLSX aqui ou clique em Selecionar")

        session = load_json(SESSION_FILE, {})
        if session.get("file") and Path(session["file"]).exists():
            tk.Button(inner,text="RECUPERAR ÚLTIMA SESSÃO",command=self.recover_session,
                      bg=LIGHT_BLUE,fg=LIGHT_BLUE_TXT,relief="flat",
                      font=("Segoe UI",8,"bold"),padx=10,pady=6).pack(anchor="e",pady=(7,0))

        self.loading_box=tk.Frame(left,bg=CARD,highlightbackground=LINE,highlightthickness=1)
        self.loading_box.pack(fill="x",pady=(8,0))
        self.loading_bar=ttk.Progressbar(self.loading_box,mode="indeterminate",style="SR.Loading.Horizontal.TProgressbar")
        self.loading_label=tk.Label(self.loading_box,text="",bg=CARD,fg=TEXT,font=("Segoe UI",9,"bold"))
        self.loading_elapsed=tk.StringVar(value="")
        self.loading_elapsed_label=tk.Label(self.loading_box,textvariable=self.loading_elapsed,bg=CARD,fg=BLUE_2,font=("Segoe UI",8,"bold"))
        self.import_started=None; self.import_loading_active=False

        stats=tk.Frame(left,bg=APP_BG)
        stats.pack(fill="x",pady=12)
        for i in range(6): stats.grid_columnconfigure(i,weight=1)
        self.stat_total=self.make_stat(stats,0,"Total","0",GREEN,GREEN_TXT)
        self.stat_one=self.make_stat(stats,1,"1 preço","0",LIGHT_BLUE,LIGHT_BLUE_TXT)
        self.stat_two=self.make_stat(stats,2,"Promo + Clube","0",PURPLE,PURPLE_TXT)
        self.stat_club=self.make_stat(stats,3,"Clube exclusivo","0",YELLOW,YELLOW_TXT)
        self.stat_error=self.make_stat(stats,4,"Erros","0",RED,RED_TXT)
        self.stat_warn=self.make_stat(stats,5,"Atenções","0",ORANGE,ORANGE_TXT)

        camp_card=tk.Frame(left,bg=CARD,highlightbackground=LINE,highlightthickness=1)
        camp_card.pack(fill="both",expand=True)
        ch=tk.Frame(camp_card,bg=CARD)
        ch.pack(fill="x",padx=20,pady=(14,8))
        tk.Label(ch,text="Campanhas detectadas",bg=CARD,fg=TEXT,
                 font=("Segoe UI",12,"bold")).pack(side="left")
        self.campaign_count=tk.Label(ch,text="0 campanhas",bg=CARD,fg=MUTED,
                                     font=("Segoe UI",9))
        self.campaign_count.pack(side="right")
        self.campaign_list=tk.Frame(camp_card,bg=CARD)
        self.campaign_list.pack(fill="both",expand=True,padx=18,pady=(0,10))
        self.render_campaigns()

        # Right panel
        panel=tk.Frame(right,bg=CARD,highlightbackground=LINE,highlightthickness=1)
        panel.pack(fill="both",expand=True)
        tk.Label(panel,text="Gerar cartazes",bg=CARD,fg=TEXT,
                 font=("Segoe UI",12,"bold")).pack(anchor="w",padx=18,pady=(16,6))
        tip=tk.Label(panel,text="ⓘ",bg=CARD,fg=BLUE_2,font=("Segoe UI",10,"bold"))
        tip.place(relx=1.0,x=-26,y=16,anchor="ne");add_tooltip(tip,"Revise os produtos antes de gerar. Atenções e erros aparecem na revisão.")

        self.summary_box=tk.Label(panel,text="Aguardando planilha",bg=ROW_ALT,fg=MUTED,
                                  font=("Segoe UI",10,"bold"),wraplength=390,justify="left",
                                  padx=12,pady=12)
        self.summary_box.pack(fill="x",padx=18,pady=(14,8))

        self.review_btn=tk.Button(panel,text="REVISAR CARTAZES",command=self.open_review,
                                  bg=LIGHT_BLUE,fg=LIGHT_BLUE_TXT,relief="flat",
                                  font=("Segoe UI",10,"bold"),pady=10,state="disabled")
        self.review_btn.pack(fill="x",padx=18,pady=(4,12))

        self.promo_options_visible=False
        self.promo_options_btn=tk.Button(panel,text="OPÇÕES  ▾",command=self.toggle_promo_options,bg=ROW_ALT,fg=MUTED,relief="flat",font=("Segoe UI",8,"bold"),pady=6)
        self.promo_options_btn.pack(fill="x",padx=18,pady=(0,4))
        self.promo_options_frame=tk.Frame(panel,bg=CARD)
        options=self.promo_options_frame
        tk.Label(options,text="Validade",bg=CARD,fg=MUTED,font=("Segoe UI",8,"bold")).pack(anchor="w")
        ttk.Combobox(options,textvariable=self.validity,values=VALIDITY_OPTIONS,state="readonly").pack(fill="x",pady=(4,8))
        tk.Label(options,text="Saída",bg=CARD,fg=MUTED,font=("Segoe UI",8,"bold")).pack(anchor="w")
        ttk.Combobox(options,textvariable=self.output_mode,values=OUTPUT_OPTIONS,state="readonly").pack(fill="x",pady=(4,8))

        self.gen_progress=ttk.Progressbar(panel,style="SR.Horizontal.TProgressbar",maximum=100)
        self.gen_progress.pack(fill="x",padx=18,pady=(10,4))
        self.gen_spinner=ttk.Progressbar(panel,mode="indeterminate",style="SR.Loading.Horizontal.TProgressbar")
        self.gen_status=tk.Label(panel,textvariable=self.status_text,bg=CARD,fg=MUTED,
                                 font=("Segoe UI",8),wraplength=390,justify="left")
        self.gen_status.pack(fill="x",padx=18,pady=(2,2))
        self.gen_elapsed=tk.StringVar(value="")
        tk.Label(panel,textvariable=self.gen_elapsed,bg=CARD,fg=BLUE_2,font=("Segoe UI",8,"bold")).pack(fill="x",padx=18,pady=(0,4))

        self.generate_btn=tk.Button(panel,text="SALVAR PDF",command=lambda:self.generate("save"),
                                    bg=BLUE,fg="white",relief="flat",
                                    font=("Segoe UI",10,"bold"),pady=10,state="disabled")
        self.generate_btn.pack(fill="x",padx=18,pady=(8,5))
        printrow=tk.Frame(panel,bg=CARD);printrow.pack(fill="x",padx=18,pady=(0,5))
        self.promo_print_btn=tk.Button(printrow,text="IMPRIMIR DIRETO",command=lambda:self.generate("print"),bg=GREEN,fg=GREEN_TXT,relief="flat",font=("Segoe UI",8,"bold"),pady=7,state="disabled")
        self.promo_print_btn.pack(side="left",fill="x",expand=True,padx=(0,3))
        self.promo_both_btn=tk.Button(printrow,text="SALVAR + IMPRIMIR",command=lambda:self.generate("both"),bg=ORANGE,fg=ORANGE_TXT,relief="flat",font=("Segoe UI",8,"bold"),pady=7,state="disabled")
        self.promo_both_btn.pack(side="left",fill="x",expand=True,padx=(3,0))
        self.promo_queue_btn=tk.Button(panel,text="ADICIONAR À FILA",command=self.add_current_promo_to_queue,bg=PURPLE,fg=PURPLE_TXT,relief="flat",font=("Segoe UI",8,"bold"),pady=7,state="disabled")
        self.promo_queue_btn.pack(fill="x",padx=18,pady=(0,5))

        self.cancel_btn=tk.Button(panel,text="CANCELAR GERAÇÃO",command=self.cancel_generation,
                                  bg=RED,fg=RED_TXT,relief="flat",
                                  font=("Segoe UI",9,"bold"),pady=8,state="disabled")
        self.cancel_btn.pack(fill="x",padx=18,pady=(0,14))

    def toggle_promo_options(self):
        self.promo_options_visible=not getattr(self,"promo_options_visible",False)
        if self.promo_options_visible:
            self.promo_options_frame.pack(fill="x",padx=18,pady=(2,6),before=self.gen_progress)
            self.promo_options_btn.config(text="OPÇÕES  ▴")
        else:
            self.promo_options_frame.pack_forget();self.promo_options_btn.config(text="OPÇÕES  ▾")

    def render_campaigns(self):
        if not hasattr(self,"campaign_list"): return
        for w in self.campaign_list.winfo_children(): w.destroy()
        if not self.analysis:
            empty=tk.Frame(self.campaign_list,bg=ROW_ALT,highlightbackground=LINE,highlightthickness=1)
            empty.pack(fill="x",pady=10)
            tk.Label(empty,text="▣",bg=ROW_ALT,fg=BLUE_2,font=("Segoe UI",22,"bold")).pack(pady=(14,2))
            tk.Label(empty,text="Nenhuma campanha carregada",bg=ROW_ALT,fg=TEXT,font=("Segoe UI",10,"bold")).pack()
            tk.Label(empty,text="Importe uma planilha .XLSX para começar.",bg=ROW_ALT,fg=MUTED,font=("Segoe UI",8),wraplength=520).pack(padx=15,pady=(3,14))
            return
        self.campaign_count.config(text=f"{len(self.analysis['campaigns'])} campanha(s)")
        for c in self.analysis["campaigns"]:
            row=tk.Frame(self.campaign_list,bg=ROW_ALT,highlightbackground=LINE,highlightthickness=1)
            row.pack(fill="x",pady=3)
            tk.Label(row,text=c["name"],bg=ROW_ALT,fg=TEXT,
                     font=("Segoe UI",9,"bold")).pack(anchor="w",padx=10,pady=(7,1))
            detail=f"{c['total']} produtos • {c['one']} 1 preço • {c['two']} Promo+Clube • {c.get('club',0)} Clube Exclusivo"
            if c["validity"]: detail += f" • {c['validity']}"
            tk.Label(row,text=detail,bg=ROW_ALT,fg=MUTED,
                     font=("Segoe UI",8)).pack(anchor="w",padx=10,pady=(0,7))

    def start_import_animation(self):
        self.loading_label.pack(anchor="w",padx=14,pady=(10,4))
        self.loading_bar.pack(fill="x",padx=14)
        self.loading_elapsed_label.pack(anchor="w",padx=14,pady=(5,10))
        self.loading_bar.start(12)
        self.import_started=time.time(); self.import_loading_active=True; self._tick_import_elapsed()
        self.import_btn.config(state="disabled",text="IMPORTANDO...")
        self.status_text.set("Lendo planilha...")
        self._animate_import_message(0)

    def _tick_import_elapsed(self):
        if not getattr(self,"import_loading_active",False) or not self.import_started:return
        self.loading_elapsed.set(f"{time.time()-self.import_started:.1f}s decorridos")
        self.after(250 if self.animations_mode.get()=="Normal" else 650,self._tick_import_elapsed)

    def _animate_import_message(self, step):
        if str(self.import_btn["state"]) != "disabled":
            return
        msgs=["Lendo planilha...","Detectando campanhas...","Verificando preços...",
              "Conferindo unidades...","Validando datas...","Preparando revisão..."]
        icons=["◌","◔","◑","◕","●"]
        self.loading_label.config(text=icons[step % len(icons)] + "  " + msgs[step % len(msgs)])
        delay=700 if self.animations_mode.get()=="Reduzidas" else 260
        self.after(delay, lambda:self._animate_import_message(step+1))

    def stop_import_animation(self):
        self.import_loading_active=False
        self.loading_bar.stop()
        self.loading_bar.pack_forget()
        self.loading_label.pack_forget()
        self.loading_elapsed_label.pack_forget()
        self.import_btn.config(state="normal",text="Selecionar .XLSX")

    def clear_promotion_generation(self, automatic=False):
        """Limpa somente a geração atual de Promoções sem apagar histórico ou PDFs."""
        if getattr(self,"busy",False):
            if not automatic:
                messagebox.showinfo("Limpar geração","Aguarde a tarefa atual terminar.")
            return False

        if self.analysis and not automatic:
            if not messagebox.askyesno(
                "Limpar geração",
                "Remover a planilha e a revisão atuais?\n\n"
                "O histórico e os PDFs já gerados não serão apagados."
            ):
                return False

        self.analysis=None
        self._builder_source=False
        self._builder_campaign_id=None
        self.file_path.set("")
        self.last_failed_jobs=[]
        self.last_error_log=None
        self.last_generation_base=None
        try:
            if SESSION_FILE.exists():
                SESSION_FILE.unlink()
        except Exception:
            pass

        if hasattr(self,"file_label"):
            self.file_label.config(text="Nenhuma planilha selecionada",fg=MUTED)
        for attr in ("stat_total","stat_one","stat_two","stat_club","stat_error","stat_warn"):
            w=getattr(self,attr,None)
            if w is not None:
                try:w.config(text="0")
                except Exception:pass
        if hasattr(self,"campaign_count"):
            self.campaign_count.config(text="0 campanhas")
        if hasattr(self,"summary_box"):
            self.summary_box.config(
                text=("✓ Geração concluída • pronto para uma nova planilha"
                      if automatic else "Aguardando planilha"),
                fg=GREEN_TXT if automatic else MUTED
            )
        for attr in ("review_btn","generate_btn","promo_print_btn","promo_both_btn","promo_queue_btn"):
            w=getattr(self,attr,None)
            if w is not None:
                try:w.config(state="disabled")
                except Exception:pass
        if hasattr(self,"clear_promo_btn"):
            self.clear_promo_btn.config(state="disabled")
        if hasattr(self,"gen_progress"):
            self.gen_progress["value"]=0
        if hasattr(self,"gen_elapsed"):
            self.gen_elapsed.set("")
        if hasattr(self,"status_text"):
            self.status_text.set(
                "Geração concluída. Selecione uma nova planilha."
                if automatic else "Geração limpa. Selecione uma nova planilha."
            )
        try:self.render_campaigns()
        except Exception:pass
        return True

    def pick_file(self):
        p=filedialog.askopenfilename(title="Selecionar planilha",
                                     filetypes=[("Planilha Excel","*.xlsx")])
        if p: self.import_file(p)

    def recover_session(self):
        sess=load_json(SESSION_FILE,{})
        p=sess.get("file")
        if p and Path(p).exists():
            self.import_file(p, restore_session=True)

    def import_file(self,p,restore_session=False):
        if self.busy:
            messagebox.showinfo("SR Studio","Aguarde a tarefa atual terminar.")
            return
        self.busy=True
        self.file_path.set(p)
        self.file_label.config(text=Path(p).name,fg=TEXT)
        self.start_import_animation()

        def worker():
            try:
                analysis=analyze_workbook(p)
                if restore_session:
                    sess=load_json(SESSION_FILE,{})
                    mods=sess.get("mods",{})
                    for j in analysis["jobs"]:
                        key=f"{j['sheet']}|{j['linha']}"
                        old_key=f"{j['sheet']}|{j['linha']}|{j['produto']}"
                        m=mods.get(key) or mods.get(old_key)
                        if m:
                            j["selected"]=m.get("selected",True)
                            j["_original"]={"produto":j.get("produto"),"promocao":j.get("promocao"),"clube":j.get("clube"),"unidade_exibicao":j.get("unidade_exibicao"),"limite":j.get("limite",""),"copies":j.get("copies",1)}
                            if m.get("product"): j["produto"]=m["product"]; j["produto_render"]=wrap_product_name(j["produto"])
                            if "promo" in m: j["promocao"]=m.get("promo",j["promocao"])
                            if "club" in m: j["clube"]=m.get("club",j["clube"])
                            if m.get("unit") in UNIT_OPTIONS:j["unidade_exibicao"]=m["unit"]
                            if "limit" in m:j["limite"]=m.get("limit","")
                            if "copies" in m:
                                try:j["copies"]=max(1,int(m.get("copies") or 1))
                                except Exception:j["copies"]=1
                            j["manual_edit"]=bool(m.get("manual_edit"))
                self.after(0,lambda:self.finish_import(analysis))
            except Exception as e:
                msg=str(e)
                self.after(0,lambda msg=msg:self.import_error(msg))
        threading.Thread(target=worker,daemon=True).start()

    def finish_import(self,analysis):
        self.busy=False
        self.stop_import_animation()
        self.analysis=analysis
        self.stat_total.config(text=str(analysis["total"]))
        self.stat_one.config(text=str(analysis["one"]))
        self.stat_two.config(text=str(analysis["two"]))
        self.stat_club.config(text=str(analysis.get("club",0)))
        self.stat_error.config(text=str(len(analysis["errors"])))
        self.stat_warn.config(text=str(len(analysis["warnings"])))
        self.render_campaigns()
        selected=sum(1 for j in analysis["jobs"] if j["selected"])
        limited=sum(1 for j in analysis["jobs"] if j["selected"] and str(j.get("limite","")).strip())
        self.summary_box.config(
            text=f"{selected} cartazes selecionados • {analysis.get('club',0)} Clube Exclusivo • {limited} com limite\n"
                 f"{len(analysis['errors'])} erro(s) • {len(analysis['warnings'])} atenção(ões)\n"
                 f"{len(analysis['skips'])} bloco(s) ignorado(s)",
            fg=RED_TXT if analysis["errors"] else TEXT
        )
        self.review_btn.config(state="normal" if analysis["jobs"] else "disabled")
        self.generate_btn.config(state="normal" if analysis["jobs"] else "disabled")
        self.promo_print_btn.config(state="normal" if analysis["jobs"] else "disabled")
        self.promo_both_btn.config(state="normal" if analysis["jobs"] else "disabled")
        if hasattr(self,"clear_promo_btn"):self.clear_promo_btn.config(state="normal")
        if hasattr(self,"promo_queue_btn"):self.promo_queue_btn.config(state="normal" if analysis["jobs"] else "disabled")
        self.last_update_text.set(f"Promoções: {Path(self.file_path.get()).name}")
        self.status_text.set("Planilha analisada. Revise os cartazes antes de gerar.")
        self.save_session()
        if analysis["errors"]:
            messagebox.showwarning(
                "SR Studio - Correção necessária",
                "Foram encontrados erros que impedem a geração:\n\n" +
                "\n".join(analysis["errors"][:8])
            )

    def import_error(self,e):
        self.busy=False
        self.stop_import_animation()
        self.status_text.set(str(e))
        messagebox.showerror("SR Studio",str(e))

    def open_review(self):
        if self.analysis:
            ReviewWindow(self,self.analysis,self.review_changed)

    def review_changed(self):
        if not self.analysis: return
        selected=sum(1 for j in self.analysis["jobs"] if j["selected"])
        limited=sum(1 for j in self.analysis["jobs"] if j["selected"] and str(j.get("limite","")).strip())
        one=sum(1 for j in self.analysis["jobs"] if j["selected"] and j["tipo"]==1)
        two=sum(1 for j in self.analysis["jobs"] if j["selected"] and j["tipo"]==2)
        club=sum(1 for j in self.analysis["jobs"] if j["selected"] and j["tipo"]==3)
        self.stat_total.config(text=str(selected))
        self.stat_one.config(text=str(one))
        self.stat_two.config(text=str(two))
        self.stat_club.config(text=str(club))
        self.summary_box.config(
            text=f"{selected} cartazes serão gerados • {club} Clube Exclusivo • {limited} com limite\n"
                 f"{len(self.analysis['errors'])} erro(s) • {len(self.analysis['warnings'])} atenção(ões)"
        )
        self.save_session()

    def save_session(self):
        if not self.analysis or not self.file_path.get(): return
        if getattr(self,"_builder_source",False): return
        mods={}
        for j in self.analysis["jobs"]:
            key=f"{j['sheet']}|{j['linha']}"
            mods[key]={"selected":j.get("selected",True),"product":j.get("produto",""),"promo":j.get("promocao",""),"club":j.get("clube",""),"unit":j.get("unidade_exibicao","UN"),"limit":j.get("limite",""),"copies":j.get("copies",1),"manual_edit":bool(j.get("manual_edit"))}
        save_json(SESSION_FILE,{"file":self.file_path.get(),"mods":mods,"saved_at":datetime.now().isoformat(),"autosave":True})

    def cancel_generation(self):
        self.cancel_event.set()
        self.status_text.set("Cancelando após o cartaz atual...")

    def set_progress_threadsafe(self,current,total,text):
        def ui():
            self.gen_progress["value"]=0 if total==0 else current/total*100
            elapsed=time.time()-getattr(self,"generation_started",time.time())
            self.status_text.set(text)
            if hasattr(self,"gen_elapsed"): self.gen_elapsed.set(f"{elapsed:.1f}s decorridos • {current}/{total}" if total else f"{elapsed:.1f}s decorridos")
        self.after(0,ui)

    def add_current_promo_to_queue(self):
        if not self.analysis:return
        jobs=[j for j in self.analysis.get("jobs",[]) if j.get("selected",True)]
        if not jobs:messagebox.showwarning("Fila","Nenhum cartaz selecionado.");return
        issues=validate_promo_jobs(jobs)
        dlg=PreGenerationDialog(self,jobs,issues,on_correct=self.correct_verification_issue,palette=self.palette)
        if not dlg.show():return
        self.enqueue_promo_jobs(jobs,self.validity.get(),self.output_mode.get());self.navigate("queue")

    def correct_verification_issue(self,issue):
        jid=issue.get("job_id") if issue else None
        if jid is not None and self.analysis:
            ReviewWindow(self,self.analysis,self.review_changed,focus_job_id=jid)

    def generate(self,action="save"):
        if self.busy:
            messagebox.showinfo("SR Studio","Aguarde a tarefa atual terminar.")
            return
        if not self.analysis:
            return
        if self.analysis["errors"]:
            messagebox.showwarning(
                "SR Studio - Corrigir antes de gerar",
                "\n".join(self.analysis["errors"][:8])
            )
            return

        jobs=[j for j in self.analysis["jobs"] if j["selected"]]
        if not jobs:
            messagebox.showwarning("SR Studio","Nenhum produto está selecionado.")
            return

        # Painel de verificações: custo, preço fora do padrão, duplicidade, Promo/Clube e unidade.
        issues=validate_promo_jobs(jobs)
        dlg=PreGenerationDialog(self,jobs,issues,on_correct=self.correct_verification_issue,palette=self.palette)
        if not dlg.show():return

        mode=self.output_mode.get();base_dir=dated_output_dir("Promoções",self.ui_settings);first=jobs[0]
        suggested=smart_pdf_name("Promoções",first.get("campanha","PROMOCOES"),first.get("validade",""))
        if action=="print":
            mode="PDF ÚNICO"; target=LOCAL_DATA/"SR_STUDIO_PROMOCOES_IMPRESSAO.pdf"
        elif mode=="PDF ÚNICO":
            output=filedialog.asksaveasfilename(title="Salvar PDF final",defaultextension=".pdf",filetypes=[("PDF","*.pdf")],initialdir=str(base_dir),initialfile=suggested)
            if not output:return
            target=unique_path(Path(output)) if Path(output).exists() else Path(output)
        else:
            folder=filedialog.askdirectory(title="Escolher pasta para PDFs por campanha",initialdir=str(base_dir))
            if not folder:return
            target=Path(folder)

        self._run_generation(jobs,target,mode,is_retry=False,action=action)

    def _run_generation(self,jobs,target,mode,is_retry=False,action="save"):
        self.busy=True
        self.cancel_event.clear()
        self.generate_btn.config(state="disabled",text="GERANDO...")
        self.promo_print_btn.config(state="disabled"); self.promo_both_btn.config(state="disabled")
        self.cancel_btn.config(state="normal")
        self.review_btn.config(state="disabled")
        self.gen_progress["value"]=0
        self.gen_spinner.pack(fill="x",padx=18,pady=(0,4))
        self.gen_spinner.start(12)
        self.status_text.set("Preparando o PowerPoint...")
        started=time.time(); self.generation_started=started
        if hasattr(self,"gen_elapsed"): self.gen_elapsed.set("0.0s decorridos")
        self.last_generation_base=target
        ordered_jobs=smart_queue_jobs(jobs)

        def worker():
            try:
                outputs=[]
                failed=[]
                success=0

                if mode=="PDF ÚNICO":
                    result=generate_pdf(ordered_jobs,target,self.validity.get(),
                                        self.set_progress_threadsafe,self.cancel_event)
                    success += result["success_count"]
                    failed.extend(result["failed_jobs"])
                    if result["output_created"]:
                        outputs.append(target)
                else:
                    grouped={}
                    for j in ordered_jobs:
                        grouped.setdefault(j["campanha"],[]).append(j)
                    done=0
                    for camp,cjobs in grouped.items():
                        if self.cancel_event.is_set():
                            raise RuntimeError("Geração cancelada pelo usuário.")
                        out=target/(safe_name(camp.replace("!!",""))+".pdf")
                        result=generate_pdf(
                            cjobs,out,self.validity.get(),
                            lambda a,b,t,done=done,camp=camp:self.set_progress_threadsafe(
                                done+a,len(ordered_jobs),f"{camp.replace('!!','')} • {t}"
                            ),
                            self.cancel_event
                        )
                        done += len(cjobs)
                        success += result["success_count"]
                        failed.extend(result["failed_jobs"])
                        if result["output_created"]:
                            outputs.append(out)

                if action in {"print","both"}:
                    if mode=="PDF ÚNICO" and outputs:
                        temp_copy=LOCAL_DATA/"temp"/"PROMO_IMPRESSAO_COM_COPIAS.pdf";temp_copy.parent.mkdir(parents=True,exist_ok=True)
                        print_source=pdf_with_copies(outputs[0],ordered_jobs,temp_copy)
                        self.print_document(print_source,"promo")
                    else:
                        for outp in outputs:
                            camp=next((c for c in {j.get("campanha") for j in ordered_jobs} if safe_name(str(c).replace("!!",""))==Path(outp).stem),None)
                            cjobs=[j for j in ordered_jobs if j.get("campanha")==camp] if camp else ordered_jobs
                            temp_copy=LOCAL_DATA/"temp"/(Path(outp).stem+"_COPIAS.pdf");temp_copy.parent.mkdir(parents=True,exist_ok=True)
                            self.print_document(pdf_with_copies(outp,cjobs,temp_copy),"promo")
                elapsed=time.time()-started
                error_log=self._write_error_log(failed,target,mode) if failed else None
                self.after(0,lambda:self.finish_generation(outputs,ordered_jobs,elapsed,success,failed,error_log,is_retry,action))
            except Exception as e:
                msg=str(e)
                self.after(0,lambda msg=msg:self.generation_error(msg))

        threading.Thread(target=worker,daemon=True).start()

    def _write_error_log(self,failed,target,mode):
        if not failed:
            return None
        folder=target.parent if mode=="PDF ÚNICO" else target
        folder.mkdir(parents=True,exist_ok=True)
        path=folder/("SR_STUDIO_ERROS_"+datetime.now().strftime("%Y%m%d_%H%M%S")+".txt")
        lines=[
            "SR STUDIO - LOG DE ERROS",
            "="*55,
            f"Data: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}",
            f"Planilha: {self.file_path.get()}",
            "",
        ]
        for i,j in enumerate(failed,1):
            lines += [
                f"{i}. {j.get('produto','')}",
                f"   Campanha: {j.get('campanha','')}",
                f"   Modelo: {'1 preço' if j.get('tipo')==1 else 'Promo + Clube' if j.get('tipo')==2 else 'Clube Exclusivo'}"
                    + (" + limite" if str(j.get('limite','')).strip() else ""),
                f"   Erro: {j.get('generation_error','Erro não identificado.')}",
                ""
            ]
        path.write_text("\n".join(lines),encoding="utf-8")
        return path

    def finish_generation(self,outputs,jobs,elapsed,success_count,failed,error_log,is_retry=False,action="save"):
        self.busy=False
        self.gen_spinner.stop()
        self.gen_spinner.pack_forget()
        self.gen_progress["value"]=100
        self.generate_btn.config(state="normal",text="SALVAR PDF")
        self.promo_print_btn.config(state="normal"); self.promo_both_btn.config(state="normal")
        self.cancel_btn.config(state="disabled")
        self.review_btn.config(state="normal")

        self.last_failed_jobs=failed
        self.last_error_log=error_log

        if failed:
            self.status_text.set(f"⚠ {success_count} gerados • {len(failed)} com erro.")
        else:
            self.status_text.set("✓ PDF gerado com sucesso.")

        campaigns=", ".join(sorted({j["campanha"].replace("!!","") for j in jobs}))
        history=load_json(HISTORY_FILE,[])
        history.append({
            "data":datetime.now().strftime("%d/%m/%Y %H:%M"),
            "planilha":Path(self.file_path.get()).name,
            "campanhas":campaigns,
            "cartazes":success_count,
            "erros":len(failed),
            "arquivos":[str(x) for x in outputs],
            "erro_log":str(error_log) if error_log else "",
            "tempo_segundos":round(elapsed,1),
        })
        save_json(HISTORY_FILE,history[-200:])
        successful=[j for j in jobs if j not in failed]
        if successful:record_product_jobs(successful,"Promoções",";".join(map(str,outputs)))
        if not failed and successful:
            try:mark_campaign_status_from_jobs(successful,"CARTAZES GERADOS")
            except Exception:pass
        if outputs:record_reprint("Promoções",outputs,success_count,campaigns,{"planilha":Path(self.file_path.get()).name})
        self.show_result(outputs,len(jobs),success_count,failed,campaigns,elapsed,error_log,is_retry,action)
        if not failed and success_count>=len(jobs):
            self.after(120,lambda:self.clear_promotion_generation(automatic=True))

    def generation_error(self,e):
        self.busy=False
        self.gen_spinner.stop()
        self.gen_spinner.pack_forget()
        self.generate_btn.config(state="normal",text="SALVAR PDF")
        self.promo_print_btn.config(state="normal"); self.promo_both_btn.config(state="normal")
        self.cancel_btn.config(state="disabled")
        self.review_btn.config(state="normal")
        self.status_text.set(str(e))
        if "cancelada" not in str(e).lower():
            messagebox.showerror("Erro no SR Studio",str(e))

    def retry_failed(self):
        if not self.last_failed_jobs:
            messagebox.showinfo("SR Studio","Não há cartazes com erro para gerar novamente.")
            return
        base=self.last_generation_base
        if base is None:
            return
        if base.suffix.lower()==".pdf":
            target=base.with_name(base.stem+"_REPROCESSADOS_ERROS.pdf")
        else:
            target=base/"REPROCESSADOS_ERROS.pdf"
        self._run_generation(self.last_failed_jobs,target,"PDF ÚNICO",is_retry=True,action="save")

    def show_result(self,outputs,total_requested,success_count,failed,campaigns,elapsed,error_log,is_retry=False,action="save"):
        w=tk.Toplevel(self)
        w.title("SR Studio - Resultado da geração")
        w.configure(bg=CARD)
        w.transient(self)
        center_toplevel(w,self,560 if failed else 540,500 if failed else 420)

        if failed:
            icon="⚠"; icon_color=ORANGE_TXT
            title="Geração concluída com pendências"
        else:
            icon="✓"; icon_color=GREEN_TXT
            title="Cartazes enviados para impressão" if action=="print" else "PDF salvo e impressão enviada" if action=="both" else "PDF gerado com sucesso"

        tk.Label(w,text=icon,bg=CARD,fg=icon_color,font=("Segoe UI",42,"bold")).pack(pady=(22,0))
        tk.Label(w,text=title,bg=CARD,fg=TEXT,font=("Segoe UI",17,"bold")).pack()
        tk.Label(
            w,
            text=f"{success_count} gerados • {len(failed)} com erro • {elapsed:.1f} segundos",
            bg=CARD,fg=MUTED,font=("Segoe UI",10)
        ).pack(pady=(5,10))
        tk.Label(w,text=campaigns,bg=CARD,fg=TEXT,font=("Segoe UI",9),
                 wraplength=480,justify="center").pack(padx=20,pady=(0,14))

        if outputs:
            first=Path(outputs[0])
            tk.Button(w,text="ABRIR PDF",command=lambda:os.startfile(first),
                      bg=BLUE,fg="white",relief="flat",font=("Segoe UI",10,"bold"),
                      pady=9).pack(fill="x",padx=75,pady=4)
            tk.Button(w,text="ABRIR PASTA",command=lambda:os.startfile(first.parent),
                      bg=LIGHT_BLUE,fg=LIGHT_BLUE_TXT,relief="flat",font=("Segoe UI",10,"bold"),
                      pady=9).pack(fill="x",padx=75,pady=4)
            tk.Button(w,text="IMPRIMIR PDF",command=lambda:print_pdf(first),
                      bg=GREEN,fg=GREEN_TXT,relief="flat",font=("Segoe UI",9,"bold"),
                      pady=8).pack(fill="x",padx=75,pady=4)

        if failed:
            sample="\n".join(
                f"• {j.get('produto','')}: {j.get('generation_error','Erro')}"
                for j in failed[:4]
            )
            tk.Label(w,text=sample,bg=RED,fg=RED_TXT,font=("Segoe UI",8),
                     wraplength=450,justify="left",padx=10,pady=8).pack(fill="x",padx=45,pady=(8,4))
            if error_log:
                tk.Button(w,text="LOG DE ERROS",command=lambda:os.startfile(error_log),
                          bg=ORANGE,fg=ORANGE_TXT,relief="flat",font=("Segoe UI",9,"bold"),
                          pady=8).pack(fill="x",padx=75,pady=4)
            tk.Button(w,text="GERAR NOVAMENTE APENAS OS COM ERRO",
                      command=lambda:(w.destroy(),self.retry_failed()),
                      bg=BLUE,fg="white",relief="flat",font=("Segoe UI",9,"bold"),
                      pady=9).pack(fill="x",padx=75,pady=4)

        tk.Button(w,text="NOVA PROMOÇÃO",command=lambda:(w.destroy(),self.navigate("promo")),
                  bg=ROW_ALT,fg=TEXT,relief="flat",font=("Segoe UI",9,"bold"),
                  pady=8).pack(fill="x",padx=75,pady=4)

    def show_history(self):
        self.clear_content()
        self.page_title.config(text="Histórico")
        frame=tk.Frame(self.content,bg=APP_BG)
        frame.pack(fill="both",expand=True,padx=28,pady=22)
        tk.Label(frame,text="Histórico",bg=APP_BG,fg=TEXT,
                 font=("Segoe UI",20,"bold")).pack(anchor="w",pady=(0,10))

        card=tk.Frame(frame,bg=CARD,highlightbackground=LINE,highlightthickness=1)
        card.pack(fill="both",expand=True)
        cols=("data","planilha","campanhas","cartazes","erros","tempo")
        tree=ttk.Treeview(card,columns=cols,show="headings")
        for col,title,width in [
            ("data","Data",125),("planilha","Planilha",240),("campanhas","Campanhas",280),
            ("cartazes","Cartazes",75),("erros","Erros",60),("tempo","Tempo",75)
        ]:
            tree.heading(col,text=title); tree.column(col,width=width,anchor="w")
        tree.pack(fill="both",expand=True,padx=12,pady=12)
        hist=load_json(HISTORY_FILE,[])
        if not hist:
            tree.insert("","end",values=("—","Nenhuma geração registrada","Use Promoções, Atacado ou Geração Manual para começar.","—","—","—"))
        for idx,h in enumerate(reversed(hist[-100:])):
            tree.insert("", "end", iid=str(idx), values=(
                h.get("data",""),h.get("planilha",""),h.get("campanhas",""),
                h.get("cartazes",""),h.get("erros",0),str(h.get("tempo_segundos",""))+"s"
            ))
        def open_selected_folder():
            sel=tree.selection()
            if not sel: return
            h=list(reversed(hist[-100:]))[int(sel[0])]
            files=h.get("arquivos",[])
            if files:
                try: os.startfile(Path(files[0]).parent)
                except Exception: pass
        buttons=tk.Frame(frame,bg=APP_BG); buttons.pack(fill="x",pady=(10,0))
        tk.Button(buttons,text="ABRIR PASTA",command=open_selected_folder,
                  bg=BLUE,fg="white",relief="flat",font=("Segoe UI",9,"bold"),
                  padx=14,pady=8).pack(side="right")
        def open_error_log():
            sel=tree.selection()
            if not sel: return
            h=list(reversed(hist[-100:]))[int(sel[0])]
            p=h.get("erro_log","")
            if p and Path(p).exists():
                os.startfile(p)
            else:
                messagebox.showinfo("SR Studio","Esta geração não possui log de erros.")
        tk.Button(buttons,text="LOG DE ERROS",command=open_error_log,
                  bg=ORANGE,fg=ORANGE_TXT,relief="flat",font=("Segoe UI",9,"bold"),
                  padx=14,pady=8).pack(side="right",padx=6)

    def show_models(self):
        self.clear_content();self.page_title.config(text="Modelos")
        for k,b in self.nav_buttons.items():b.config(bg=self.palette["SIDEBAR_HOVER"] if k=="modelos" else self.palette["SIDEBAR"],fg="white" if k=="modelos" else "#DCE7F7")
        pal=self.palette
        canvas=tk.Canvas(self.content,bg=pal["APP_BG"],highlightthickness=0);sb=ttk.Scrollbar(self.content,orient="vertical",command=canvas.yview);canvas.configure(yscrollcommand=sb.set);canvas.pack(side="left",fill="both",expand=True);sb.pack(side="right",fill="y")
        frame=tk.Frame(canvas,bg=pal["APP_BG"]);win=canvas.create_window((0,0),window=frame,anchor="nw");frame.bind("<Configure>",lambda e:canvas.configure(scrollregion=canvas.bbox("all")));canvas.bind("<Configure>",lambda e:canvas.itemconfigure(win,width=e.width));frame.configure(padx=28,pady=22)
        mh=tk.Frame(frame,bg=pal["APP_BG"]);mh.pack(fill="x",pady=(0,10))
        tk.Label(mh,text="Modelos",bg=pal["APP_BG"],fg=pal["TEXT"],font=("Segoe UI",20,"bold")).pack(side="left")
        info=tk.Label(mh,text="ⓘ",bg=pal["APP_BG"],fg=pal["BLUE2"],font=("Segoe UI",10,"bold"));info.pack(side="left",padx=8);add_tooltip(info,"Modelos oficiais do PowerPoint. Miniaturas e campos são verificados automaticamente.")

        grid=tk.Frame(frame,bg=pal["APP_BG"]);grid.pack(fill="x")
        for i in range(3):grid.grid_columnconfigure(i,weight=1,uniform="models")
        specs=[
            ("Promoção • 1 preço",MODEL1,{"SR_PRODUTO","SR_PRECO_PROMO","SR_VALIDADE","SR_CAMPANHA","SR_UNIDADE"}),
            ("Promoção • 1 preço + limite",MODEL1_LIMIT,{"SR_PRODUTO","SR_PRECO_PROMO","SR_VALIDADE","SR_CAMPANHA","SR_UNIDADE","SR_LIMITE"}),
            ("Promo + Clube • 2 preços",MODEL2,{"SR_PRODUTO","SR_PRECO_PROMO","SR_PRECO_CLUBE","SR_VALIDADE","SR_CAMPANHA","SR_UNIDADE_PROMO","SR_UNIDADE_CLUBE"}),
            ("Promo + Clube • 2 preços + limite",MODEL2_LIMIT,{"SR_PRODUTO","SR_PRECO_PROMO","SR_PRECO_CLUBE","SR_VALIDADE","SR_CAMPANHA","SR_UNIDADE_PROMO","SR_UNIDADE_CLUBE","SR_LIMITE"}),
            ("Clube Exclusivo",CLUB_MODEL,{"SR_CLUBE_PRODUTO","SR_CLUBE_PRECO","SR_CLUBE_VALIDADE"}),
            ("Clube Exclusivo + limite",CLUB_MODEL_LIMIT,{"SR_CLUBE_PRODUTO","SR_CLUBE_PRECO","SR_CLUBE_VALIDADE","SR_CLUBE_LIMITE"}),
            ("Atacado",ATACADO_MODEL,{"SR_ATACADO_NOME","SR_ATACADO_VAREJO","SR_ATACADO_PRECO","SR_ATACADO_TOTAL","SR_ATACADO_QUANTIDADE","SR_ATACADO_QUANTIDADE_2"}),
            ("Cartaz Venda",MODEL_SALE,{"SR_VENDA_PRODUTO","SR_VENDA_PRECO","SR_VENDA_UNIDADE"}),
        ]
        self._model_images=[]
        cache_dir=LOCAL_DATA/"model_thumbnails";cache_dir.mkdir(parents=True,exist_ok=True);thumb_tasks=[]
        def has_fields(path,required):
            if not Path(path).exists():return False,required
            found=set()
            try:
                with zipfile.ZipFile(path) as z:
                    for fn in z.namelist():
                        if fn.startswith("ppt/slides/slide") and fn.endswith(".xml"):
                            text=z.read(fn).decode("utf-8","ignore")
                            found.update(re.findall(r'name="(SR_[^"]+)"',text))
                return required.issubset(found),required-found
            except Exception:return False,required
        def thumb_path(path):return cache_dir/(cache_key(file_signature(path))+".png")
        def load_thumb(path,label):
            dest=thumb_path(path);bundled=ASSETS/"model_thumbs"/(Path(path).stem+".png")
            def show(imgpath):
                if not label.winfo_exists():return
                try:
                    img=tk.PhotoImage(file=str(imgpath));factor=max(1,math.ceil(max(img.width()/215,img.height()/245)));img=img.subsample(factor,factor);self._model_images.append(img);label.config(image=img,text="")
                except Exception:label.config(text="Miniatura indisponível")
            if dest.exists():show(dest);return
            if bundled.exists():show(bundled)
            thumb_tasks.append((path,dest,show))
        def create_card(i,title,path,required):
            ok,missing=has_fields(path,required)
            card=tk.Frame(grid,bg=pal["CARD"],highlightbackground=pal["LINE"],highlightthickness=1);card.grid(row=i//3,column=i%3,sticky="nsew",padx=6,pady=6)
            img=tk.Label(card,text="Carregando miniatura...",bg=pal["ROW_ALT"],fg=pal["MUTED"],font=("Segoe UI",8),height=12);img.pack(fill="x",padx=10,pady=(10,6));load_thumb(path,img)
            tk.Label(card,text=title,bg=pal["CARD"],fg=pal["TEXT"],font=("Segoe UI",10,"bold"),wraplength=250,justify="left").pack(anchor="w",padx=12)
            status=tk.Label(card,text="✓ MODELO OK" if ok else "! PRECISA DE CORREÇÃO",bg=pal["GREEN"] if ok else pal["ORANGE"],fg=pal["GREEN_TXT"] if ok else pal["ORANGE_TXT"],font=("Segoe UI",7,"bold"),padx=7,pady=3);status.pack(anchor="w",padx=12,pady=(6,3));add_tooltip(status,"Todos os campos SR necessários foram encontrados." if ok else "Campos ausentes: "+", ".join(sorted(missing)))
            tk.Label(card,text=Path(path).name,bg=pal["CARD"],fg=pal["MUTED"],font=("Segoe UI",7),wraplength=250,justify="left").pack(anchor="w",padx=12,pady=(2,5))
            buttons=tk.Frame(card,bg=pal["CARD"]);buttons.pack(fill="x",padx=10,pady=(3,10))
            def replace_model():
                src=filedialog.askopenfilename(title="Selecionar modelo PowerPoint",filetypes=[("PowerPoint","*.pptx")])
                if not src:return
                if messagebox.askyesno("Substituir modelo",f"Substituir o modelo {title}?"):
                    backup_model_version(path,"antes_substituir");shutil.copy2(src,path);unlock_sr_model_fields(path);self.toast.show("Modelo substituído e versão anterior salva.","ok");self.show_models();self.refresh_health_async()
            def restore_model():
                original=ORIGINAL_MODELS/Path(path).name
                if original.exists() and messagebox.askyesno("Restaurar modelo",f"Restaurar o modelo original de {title}?"):
                    backup_model_version(path,"antes_original");shutil.copy2(original,path);unlock_sr_model_fields(path);self.toast.show("Modelo original restaurado; versão anterior preservada.","ok");self.show_models();self.refresh_health_async()
            def versions_dialog():
                vers=model_versions(path);w=tk.Toplevel(self);w.title(f"Versões - {title}");w.configure(bg=pal["APP_BG"]);center_toplevel(w,self,760,470)
                box=tk.Frame(w,bg=pal["CARD"],highlightbackground=pal["LINE"],highlightthickness=1);box.pack(fill="both",expand=True,padx=16,pady=16)
                tree=ttk.Treeview(box,columns=("data","arquivo"),show="headings",selectmode="browse");tree.heading("data",text="Data");tree.heading("arquivo",text="Backup");tree.column("data",width=150);tree.column("arquivo",width=520);tree.pack(fill="both",expand=True,padx=10,pady=10)
                for i,v in enumerate(vers):tree.insert("","end",iid=str(i),values=(datetime.fromtimestamp(v.stat().st_mtime).strftime("%d/%m/%Y %H:%M"),v.name))
                if not vers:tree.insert("","end",values=("—","Nenhuma versão anterior salva ainda."))
                def restore_version():
                    sel=tree.selection()
                    if not sel or not sel[0].isdigit() or int(sel[0])>=len(vers):return
                    v=vers[int(sel[0])]
                    if messagebox.askyesno("Restaurar versão",f"Restaurar esta versão?\n\n{v.name}"):
                        restore_model_version(path,v);unlock_sr_model_fields(path);w.destroy();self.show_models();self.refresh_health_async()
                tk.Button(box,text="RESTAURAR VERSÃO SELECIONADA",command=restore_version,bg=pal["BLUE"],fg="white",relief="flat",font=("Segoe UI",8,"bold"),padx=12,pady=7).pack(anchor="e",padx=10,pady=(0,10))
            def autocorrect():
                changed=unlock_sr_model_fields(path);self.toast.show(f"Correção automática concluída • {changed} trava(s) removida(s).","ok");self.show_models()
            tk.Button(buttons,text="SUBSTITUIR",command=replace_model,bg=pal["BLUE"],fg="white",relief="flat",font=("Segoe UI",7,"bold"),padx=8,pady=5).pack(side="left")
            tk.Button(buttons,text="RESTAURAR",command=restore_model,bg=pal["LIGHT_BLUE"],fg=pal["LIGHT_BLUE_TXT"],relief="flat",font=("Segoe UI",7,"bold"),padx=8,pady=5).pack(side="left",padx=4)
            tk.Button(buttons,text="CORRIGIR",command=autocorrect,bg=pal["ORANGE"],fg=pal["ORANGE_TXT"],relief="flat",font=("Segoe UI",7,"bold"),padx=8,pady=5).pack(side="right")
            tk.Button(buttons,text="VERSÕES",command=versions_dialog,bg=pal["PURPLE"],fg=pal["PURPLE_TXT"],relief="flat",font=("Segoe UI",7,"bold"),padx=8,pady=5).pack(side="right",padx=4)
        for i,spec in enumerate(specs):create_card(i,*spec)
        def thumb_worker():
            if os.name!="nt":return
            for path,dest,show in thumb_tasks:
                try:
                    cmd=[find_powershell(),"-NoProfile","-ExecutionPolicy","Bypass","-File",str(APP_DIR/"ModelThumbnail.ps1"),"-Model",str(path),"-OutputPng",str(dest)]
                    r=subprocess.run(cmd,capture_output=True,text=True,timeout=90,creationflags=getattr(subprocess,"CREATE_NO_WINDOW",0))
                    if r.returncode==0 and dest.exists():self.after(0,lambda d=dest,show=show:show(d))
                except Exception:pass
        if thumb_tasks:threading.Thread(target=thumb_worker,daemon=True).start()


if __name__=="__main__":
    run_startup_splash()
    App().mainloop()
